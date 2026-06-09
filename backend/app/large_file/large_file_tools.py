# ==============================================================================
# 大型Excel文件操作工具集
# 使用 openpyxl 直接操作服务器上的文件
# 使用 claude_agent_sdk 的 @tool 装饰器定义 MCP 工具
# ==============================================================================
import json
import time
import threading
import re
from decimal import Decimal
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, Dict, List, Any, Union
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils.cell import coordinate_from_string

from claude_agent_sdk import tool, create_sdk_mcp_server

from .storage import large_file_storage
from .schemas import PREVIEW_ROW_COUNT
from ..utils.logger import get_logger

logger = get_logger('large_file.tools')


# ==============================================================================
# 工作簿缓存管理器
# 在一次指令处理期间缓存工作簿，避免重复打开/保存
# ==============================================================================
class WorkbookCache:
    """工作簿缓存管理器 - 单例模式"""
    _instance = None
    _lock = threading.Lock()
    
    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        self._cache: Dict[str, Dict] = {}  # {file_id: {wb, path, dirty, last_access}}
        self._cache_lock = threading.Lock()
    
    def get_workbook(self, file_id: str, read_only: bool = False) -> Optional[Workbook]:
        """获取工作簿（优先从缓存获取）"""
        with self._cache_lock:
            if file_id in self._cache and not read_only:
                entry = self._cache[file_id]
                entry['last_access'] = time.time()
                logger.debug(f'从缓存获取工作簿: file_id={file_id}')
                return entry['wb']
        
        # 缓存中没有，打开新的
        file_path = large_file_storage.get_file_path(file_id)
        if not file_path or not file_path.exists():
            logger.warning(f'文件不存在: file_id={file_id}, path={file_path}')
            return None
        
        try:
            logger.info(f'打开工作簿: {file_path}')
            start_time = time.time()
            wb = load_workbook(file_path, read_only=read_only, data_only=False)
            elapsed = time.time() - start_time
            logger.info(f'工作簿打开成功: file_id={file_id}, 耗时={elapsed:.2f}秒')
            
            # 只缓存非只读模式的工作簿
            if not read_only:
                with self._cache_lock:
                    self._cache[file_id] = {
                        'wb': wb,
                        'path': file_path,
                        'dirty': False,
                        'last_access': time.time()
                    }
            
            return wb
        except Exception as e:
            import traceback
            logger.error(f'打开工作簿失败: file_id={file_id}, 错误: {str(e)}')
            logger.debug(f'错误详情:\n{traceback.format_exc()}')
            return None
    
    def mark_dirty(self, file_id: str):
        """标记工作簿为已修改"""
        with self._cache_lock:
            if file_id in self._cache:
                self._cache[file_id]['dirty'] = True
    
    def save_workbook(self, file_id: str) -> bool:
        """保存工作簿（仅当有修改时）"""
        with self._cache_lock:
            if file_id not in self._cache:
                logger.warning(f'工作簿不在缓存中: file_id={file_id}')
                return False
            
            entry = self._cache[file_id]
            if not entry['dirty']:
                logger.debug(f'工作簿未修改，跳过保存: file_id={file_id}')
                return True
            
            wb = entry['wb']
            file_path = entry['path']
        
        try:
            logger.info(f'保存工作簿: file_id={file_id}')
            start_time = time.time()
            wb.save(file_path)
            elapsed = time.time() - start_time
            logger.info(f'工作簿保存成功: file_id={file_id}, 耗时={elapsed:.2f}秒')
            
            with self._cache_lock:
                if file_id in self._cache:
                    self._cache[file_id]['dirty'] = False
            return True
        except Exception as e:
            import traceback
            logger.error(f'保存工作簿失败: file_id={file_id}, 错误: {str(e)}')
            logger.debug(f'错误详情:\n{traceback.format_exc()}')
            return False
    
    def save_and_close(self, file_id: str) -> bool:
        """保存并关闭工作簿"""
        result = self.save_workbook(file_id)
        
        with self._cache_lock:
            if file_id in self._cache:
                try:
                    self._cache[file_id]['wb'].close()
                except:
                    pass
                del self._cache[file_id]
                logger.info(f'工作簿已关闭并移出缓存: file_id={file_id}')
        
        return result
    
    def close_all(self):
        """保存并关闭所有缓存的工作簿"""
        with self._cache_lock:
            file_ids = list(self._cache.keys())
        
        for file_id in file_ids:
            self.save_and_close(file_id)


# 全局工作簿缓存实例
workbook_cache = WorkbookCache()


# ==============================================================================
# 辅助函数
# ==============================================================================

def _col_letter_to_num(col: str) -> int:
    """列字母转数字（A=1, B=2, ...）"""
    return column_index_from_string(col)


def _col_num_to_letter(col: int) -> str:
    """列数字转字母"""
    return get_column_letter(col)


def _parse_color(color: str) -> str:
    """解析颜色值，确保为 ARGB 格式"""
    if color.startswith('#'):
        color = color[1:]
    if len(color) == 6:
        return 'FF' + color.upper()
    return color.upper()


def _get_workbook(file_id: str, read_only: bool = False) -> Optional[Workbook]:
    """获取工作簿实例（使用缓存）"""
    return workbook_cache.get_workbook(file_id, read_only)


def _mark_dirty(file_id: str):
    """标记工作簿已修改（不立即保存）"""
    workbook_cache.mark_dirty(file_id)


def _save_workbook(wb: Workbook, file_id: str) -> bool:
    """标记工作簿已修改（延迟保存，由 Agent 统一保存）"""
    # 只标记为脏，不立即保存
    _mark_dirty(file_id)
    logger.debug(f'工作簿已标记为待保存: file_id={file_id}')
    return True


def _log_tool_call(tool_name: str, args: Dict[str, Any]) -> None:
    """记录工具调用"""
    file_id = args.get('file_id', 'unknown')
    logger.info(f'工具调用: {tool_name}, file_id={file_id}')
    logger.debug(f'工具参数: {tool_name}, args={args}')


def _parse_json_param(value: Any, param_name: str) -> Any:
    """
    解析可能是 JSON 字符串的参数
    
    Claude SDK 有时会将 list/dict 参数作为 JSON 字符串传递，
    此函数自动检测并解析这种情况。
    
    Args:
        value: 参数值（可能是 str、list、dict 等）
        param_name: 参数名称（用于错误信息）
    
    Returns:
        解析后的值
    
    Raises:
        ValueError: 如果字符串无法解析为有效 JSON
    """
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            logger.debug(f'参数 {param_name} 从 JSON 字符串解析: {type(parsed).__name__}')
            return parsed
        except json.JSONDecodeError as e:
            raise ValueError(f'参数 {param_name} JSON 格式错误: {e}')
    return value


def _write_numeric_safe_cell(ws, row: int, col: int, value: Any):
    """写入数值时显式设置格式，避免被误显示为日期/时间。"""
    cell = ws.cell(row=row, column=col)
    if isinstance(value, bool):
        cell.value = value
        return
    if isinstance(value, Decimal):
        cell.value = round(float(value) + 1e-12, 2)
        cell.number_format = '0.00'
        return
    if isinstance(value, float):
        cell.value = round(float(value) + 1e-12, 2)
        cell.number_format = '0.00'
        return
    if isinstance(value, int):
        cell.value = int(value)
        cell.number_format = '0'
        return
    cell.value = value


def _create_result(
    success: bool, 
    message: str, 
    data: Optional[Dict] = None, 
    steps: Optional[list] = None,
    execution_time_ms: Optional[float] = None,
    sql_executed: Optional[str] = None
) -> Dict:
    """
    创建统一的返回结果
    
    Args:
        success: 是否成功
        message: 结果消息
        data: 返回数据
        steps: 执行步骤列表（用于向用户展示执行过程）
        execution_time_ms: 执行耗时（毫秒）
        sql_executed: 执行的 SQL 语句（用于完整追踪）
    """
    result = {
        'success': success,
        'message': message,
    }
    if data:
        result['data'] = data
    if steps:
        result['steps'] = steps  # 执行步骤，前端可展示
    if execution_time_ms is not None:
        result['execution_time_ms'] = round(execution_time_ms, 2)
    if sql_executed:
        # SQL 语句可能很长，截断显示但保留完整信息
        result['sql_executed'] = sql_executed[:500] + '...' if len(sql_executed) > 500 else sql_executed
    return {"content": [{"type": "text", "text": json.dumps(result, ensure_ascii=False)}]}


# ==============================================================================
# 结果文件与工作表命名
# ==============================================================================
def _sanitize_sheet_name(name: str) -> str:
    """清理工作表名（移除非法字符，限制长度）"""
    if not name:
        name = "结果"
    # Excel 禁用字符: : \ / ? * [ ]
    name = re.sub(r'[:\\\\/?*\\[\\]]', '_', name)
    name = name.strip() or "结果"
    return name[:31]


def _build_result_filename(file_id: str, filename: Optional[str] = None) -> str:
    """生成结果文件名"""
    if filename:
        return filename
    source_meta = large_file_storage.get_metadata(file_id)
    base_name = Path(source_meta.original_name).stem if source_meta else "分析结果"
    return f"{base_name}_分析结果.xlsx"


def _build_result_sheet_name(prefix: str, parts: Optional[List[str]] = None) -> str:
    """生成结果工作表名"""
    merged = [p for p in (parts or []) if p]
    base = "_".join([prefix] + merged) if merged else prefix
    return _sanitize_sheet_name(f"结果_{base}")


def _normalize_sheet_key(name: Optional[str]) -> str:
    """标准化工作表名用于模糊匹配。"""
    text = str(name or "").strip().lower()
    text = re.sub(r'[\s_\-()（）【】\[\]·]+', '', text)
    if text.endswith('工作表'):
        text = text[:-3]
    if text.endswith('表') and len(text) > 1:
        text = text[:-1]
    return text


def _list_available_source_sheets(file_id: str) -> List[str]:
    """返回可用源工作表（排除系统元表）。"""
    meta = large_file_storage.get_metadata(file_id)
    if not meta:
        return []
    return [s for s in (meta.sheet_names or []) if s and s != "__SHEETBOT_META__"]


def _resolve_source_sheet_name(file_id: str, requested_sheet: Optional[str]) -> Optional[str]:
    """
    将请求工作表名解析为实际存在的源工作表名。
    优先级：精确匹配 > 忽略大小写 > 归一化匹配 > 包含匹配。
    """
    sheets = _list_available_source_sheets(file_id)
    if not sheets:
        return None

    if not requested_sheet:
        return sheets[0]

    if requested_sheet in sheets:
        return requested_sheet

    req_text = str(requested_sheet).strip()
    req_lower = req_text.lower()
    for s in sheets:
        if s.lower() == req_lower:
            return s

    req_key = _normalize_sheet_key(req_text)
    if not req_key:
        return sheets[0]

    key_map = {_normalize_sheet_key(s): s for s in sheets}
    if req_key in key_map:
        return key_map[req_key]

    # 兼容“销售明细表 -> 销售明细”等场景
    contains_hits = [s for s in sheets if req_key in _normalize_sheet_key(s) or _normalize_sheet_key(s) in req_key]
    if len(contains_hits) == 1:
        return contains_hits[0]

    return None


def _find_top_level_clause_pos(sql: str, clauses: List[str]) -> Optional[int]:
    """查找 SQL 顶层子句位置（忽略括号和引号内部）。"""
    text = sql or ""
    lowered = text.lower()
    targets = [c.lower() for c in clauses if c]
    if not targets:
        return None

    depth = 0
    in_single = False
    in_double = False
    i = 0
    n = len(text)
    while i < n:
        ch = text[i]
        prev = text[i - 1] if i > 0 else ""
        if ch == "'" and not in_double and prev != "\\":
            in_single = not in_single
            i += 1
            continue
        if ch == '"' and not in_single and prev != "\\":
            in_double = not in_double
            i += 1
            continue
        if in_single or in_double:
            i += 1
            continue
        if ch == "(":
            depth += 1
            i += 1
            continue
        if ch == ")":
            depth = max(0, depth - 1)
            i += 1
            continue
        if depth == 0:
            for target in targets:
                if lowered.startswith(target, i):
                    before_ok = i == 0 or not (lowered[i - 1].isalnum() or lowered[i - 1] == "_")
                    end = i + len(target)
                    after_ok = end >= n or not (lowered[end].isalnum() or lowered[end] == "_")
                    if before_ok and after_ok:
                        return i
        i += 1
    return None


def _try_repair_group_by_binder_sql(sql: str, err_msg: str) -> Optional[str]:
    """修复 DuckDB 缺失 GROUP BY 的绑定错误（首选 GROUP BY ALL）。"""
    err = str(err_msg or "").lower()
    if "must appear in the group by clause" not in err:
        return None

    base_sql = str(sql or "").strip()
    if not base_sql:
        return None
    if _find_top_level_clause_pos(base_sql, ["group by"]) is not None:
        return None

    has_trailing_semicolon = base_sql.endswith(";")
    if has_trailing_semicolon:
        base_sql = base_sql[:-1].rstrip()

    insert_pos = _find_top_level_clause_pos(base_sql, ["having", "order by", "limit", "offset"])
    if insert_pos is None:
        repaired = f"{base_sql}\nGROUP BY ALL"
    else:
        prefix = base_sql[:insert_pos].rstrip()
        suffix = base_sql[insert_pos:].lstrip()
        repaired = f"{prefix}\nGROUP BY ALL\n{suffix}"

    if has_trailing_semicolon:
        repaired = f"{repaired};"
    return repaired


# ==============================================================================
# 读取工具
# ==============================================================================

@tool(
    "get_file_info",
    "获取大型Excel文件的基本信息（工作表列表、行数、列数等）",
    {"file_id": str}
)
async def get_file_info(args: Dict[str, Any]) -> Dict[str, Any]:
    """获取文件基本信息"""
    _log_tool_call("get_file_info", args)
    file_id = args["file_id"]
    
    meta = large_file_storage.get_metadata(file_id)
    if not meta:
        return _create_result(False, f'文件不存在: {file_id}')
    
    # 获取 DuckDB 内存中的结果表列表
    result_tables = []
    try:
        available_tables = duckdb_manager.list_available_tables(file_id)
        result_tables = [t['name'] for t in available_tables if t.get('type') == 'result']
    except Exception:
        pass  # 如果获取失败，忽略
    
    return _create_result(True, '获取文件信息成功', {
        'file_id': meta.file_id,
        'original_name': meta.original_name,
        'file_size': meta.file_size,
        'sheet_names': meta.sheet_names,
        'result_tables': result_tables,  # 新增：内存中的结果表列表
        'row_count': meta.row_count,
        'col_count': meta.col_count,
        'status': meta.status.value,
    })


@tool(
    "get_sheet_info",
    "获取工作表信息（行数、列数、表头等）",
    {"file_id": str, "sheet_name": str}
)
async def get_sheet_info(args: Dict[str, Any]) -> Dict[str, Any]:
    """获取工作表信息"""
    _log_tool_call("get_sheet_info", args)
    file_id = args["file_id"]
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id, read_only=True)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title
        
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else '')
        
        wb.close()
        
        return _create_result(True, '获取工作表信息成功', {
            'sheet_name': sheet_name,
            'row_count': ws.max_row,
            'col_count': ws.max_column,
            'headers': headers,
        })
    except Exception as e:
        wb.close()
        return _create_result(False, f'获取工作表信息失败: {e}')


@tool(
    "get_cell_value",
    "获取单元格的值",
    {"file_id": str, "row": int, "col": int, "sheet_name": str}
)
async def get_cell_value(args: Dict[str, Any]) -> Dict[str, Any]:
    """获取单元格值"""
    _log_tool_call("get_cell_value", args)
    file_id = args["file_id"]
    row = args["row"]
    col = args["col"]
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id, read_only=True)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        cell = ws.cell(row=row, column=col)
        value = cell.value
        wb.close()
        
        return _create_result(True, '获取单元格值成功', {
            'row': row,
            'col': col,
            'value': value,
        })
    except Exception as e:
        wb.close()
        return _create_result(False, f'获取单元格值失败: {e}')


@tool(
    "get_range_values",
    "获取范围内的值（二维数组）",
    {"file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int, "sheet_name": str}
)
async def get_range_values(args: Dict[str, Any]) -> Dict[str, Any]:
    """获取范围内的值"""
    _log_tool_call("get_range_values", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    sheet_name = args.get("sheet_name")
    
    # 调试日志：记录 get_range_values 参数
    logger.info(f'get_range_values 参数: start=({start_row},{start_col}), end=({end_row},{end_col}), row_count={end_row - start_row + 1}')
    
    wb = _get_workbook(file_id, read_only=True)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        
        data = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, 
                                 min_col=start_col, max_col=end_col):
            row_data = []
            for cell in row:
                value = cell.value
                # 将 datetime 转换为 ISO 格式字符串，确保 JSON 可序列化
                if hasattr(value, 'isoformat'):
                    value = value.isoformat()
                row_data.append(value)
            data.append(row_data)
        
        wb.close()
        
        logger.info(f'get_range_values 成功: rows={len(data)}, cols={len(data[0]) if data else 0}')
        
        return _create_result(True, '获取范围值成功', {
            'start_row': start_row,
            'start_col': start_col,
            'end_row': end_row,
            'end_col': end_col,
            'data': data,
        })
    except Exception as e:
        wb.close()
        logger.error(f'get_range_values 失败: {e}')
        return _create_result(False, f'获取范围值失败: {e}')


# ==============================================================================
# 写入工具
# ==============================================================================

@tool(
    "set_cell_value",
    "设置单个单元格的值。注意：如需设置多个单元格，请使用 set_range_values 批量设置，效率更高。",
    {"file_id": str, "row": int, "col": int, "value": str, "sheet_name": str}
)
async def set_cell_value(args: Dict[str, Any]) -> Dict[str, Any]:
    """设置单个单元格值（批量设置请用 set_range_values）"""
    _log_tool_call("set_cell_value", args)
    file_id = args["file_id"]
    row = args["row"]
    col = args["col"]
    value = args["value"]
    sheet_name = args.get("sheet_name")
    
    logger.debug(f'设置单元格: file_id={file_id}, sheet={sheet_name}, cell={_col_num_to_letter(col)}{row}, value={value}')
    
    wb = _get_workbook(file_id)
    if not wb:
        logger.error(f'无法打开文件: file_id={file_id}')
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        _write_numeric_safe_cell(ws, row, col, value)
        logger.debug(f'单元格值已设置: {_col_num_to_letter(col)}{row} = {value}')
        
        if _save_workbook(wb, file_id):
            logger.info(f'设置单元格成功: file_id={file_id}, cell={_col_num_to_letter(col)}{row}')
            return _create_result(True, f'已设置单元格 {_col_num_to_letter(col)}{row} = {value}')
        else:
            logger.error(f'保存文件失败: file_id={file_id}')
            return _create_result(False, '保存文件失败')
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.error(f'设置单元格值失败: file_id={file_id}, 错误: {str(e)}')
        logger.debug(f'错误详情:\n{error_details}')
        return _create_result(False, f'设置单元格值失败: {e}')
    finally:
        wb.close()


@tool(
    "set_cell_formula",
    "设置单元格的公式（公式必须以=开头）",
    {"file_id": str, "row": int, "col": int, "formula": str, "sheet_name": str}
)
async def set_cell_formula(args: Dict[str, Any]) -> Dict[str, Any]:
    """设置单元格公式"""
    _log_tool_call("set_cell_formula", args)
    file_id = args["file_id"]
    row = args["row"]
    col = args["col"]
    formula = args["formula"]
    sheet_name = args.get("sheet_name")
    
    if not formula.startswith('='):
        formula = '=' + formula
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.cell(row=row, column=col, value=formula)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已设置公式 {_col_num_to_letter(col)}{row} = {formula}')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'设置公式失败: {e}')
    finally:
        wb.close()


@tool(
    "set_range_values",
    "【推荐】批量设置范围内多个单元格的值（一次调用设置整行/整列/区域，比逐个设置快10倍以上）。values 为二维数组，每行一个数组。",
    {"file_id": str, "start_row": int, "start_col": int, "values": list, "sheet_name": str}
)
async def set_range_values(args: Dict[str, Any]) -> Dict[str, Any]:
    """【推荐】批量设置范围内多个单元格的值，比逐个设置快10倍以上"""
    _log_tool_call("set_range_values", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    values = args["values"]
    sheet_name = args.get("sheet_name")
    
    # 调试日志：记录 set_range_values 调用参数
    _val_preview = str(values)[:500] if values else "None"
    logger.info(f'set_range_values 详情: start=({start_row},{start_col}), values_type={type(values).__name__}, preview={_val_preview}')
    
    # 解析可能是 JSON 字符串的参数
    try:
        values = _parse_json_param(values, 'values')
        logger.info(f'set_range_values 解析后: type={type(values).__name__}, is_list={isinstance(values, list)}')
        if isinstance(values, list) and len(values) > 0:
            logger.info(f'set_range_values 第一行: type={type(values[0]).__name__}, content={str(values[0])[:200]}')
    except ValueError as e:
        logger.error(f'set_range_values JSON 解析失败: {e}')
        return _create_result(False, str(e))
    
    # 验证 values 是二维数组
    if not isinstance(values, list):
        logger.error(f'set_range_values values 不是列表: type={type(values).__name__}')
        return _create_result(False, f'values 必须是二维数组，实际类型: {type(values).__name__}')
    if values and not isinstance(values[0], list):
        # 如果第一行不是列表，可能是一维数组，转换为二维
        logger.warning(f'set_range_values: values 第一行不是列表，尝试转换')
        values = [values]
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 记录前几个单元格的写入情况
        cell_samples = []
        for r_idx, row_data in enumerate(values):
            for c_idx, value in enumerate(row_data):
                _write_numeric_safe_cell(ws, start_row + r_idx, start_col + c_idx, value)
                # 记录前 5 个单元格
                if len(cell_samples) < 5:
                    cell_samples.append(f'({start_row + r_idx},{start_col + c_idx})={repr(value)[:50]}')
        
        logger.info(f'set_range_values 写入样本: {cell_samples}')
        
        if _save_workbook(wb, file_id):
            end_row = start_row + len(values) - 1
            end_col = start_col + (len(values[0]) if values else 0) - 1
            logger.info(f'set_range_values 成功: rows={len(values)}, cols={len(values[0]) if values else 0}, range={_col_num_to_letter(start_col)}{start_row}:{_col_num_to_letter(end_col)}{end_row}')
            return _create_result(True, f'已设置范围 {_col_num_to_letter(start_col)}{start_row}:{_col_num_to_letter(end_col)}{end_row}')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        logger.error(f'set_range_values 失败: {e}')
        return _create_result(False, f'批量设置值失败: {e}')
    finally:
        wb.close()


# ==============================================================================
# 样式工具
# ==============================================================================

@tool(
    "set_cell_style",
    "设置单个单元格样式。注意：如需设置多个单元格，请使用 set_range_style 批量设置，效率更高。",
    {
        "file_id": str, "row": int, "col": int,
        "font_color": str, "background_color": str, "bold": bool,
        "italic": bool, "font_size": int, "horizontal_align": str,
        "vertical_align": str, "sheet_name": str
    }
)
async def set_cell_style(args: Dict[str, Any]) -> Dict[str, Any]:
    """设置单个单元格样式（批量设置请用 set_range_style）"""
    _log_tool_call("set_cell_style", args)
    file_id = args["file_id"]
    row = args["row"]
    col = args["col"]
    font_color = args.get("font_color")
    background_color = args.get("background_color")
    bold = args.get("bold")
    italic = args.get("italic")
    font_size = args.get("font_size")
    horizontal_align = args.get("horizontal_align")
    vertical_align = args.get("vertical_align")
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        cell = ws.cell(row=row, column=col)
        
        # 字体样式
        font_kwargs = {}
        if font_color:
            font_kwargs['color'] = _parse_color(font_color)
        if bold is not None:
            font_kwargs['bold'] = bold
        if italic is not None:
            font_kwargs['italic'] = italic
        if font_size:
            font_kwargs['size'] = font_size
        
        if font_kwargs:
            old_font = cell.font
            font_kwargs.setdefault('name', old_font.name)
            font_kwargs.setdefault('size', old_font.size)
            font_kwargs.setdefault('bold', old_font.bold)
            font_kwargs.setdefault('italic', old_font.italic)
            cell.font = Font(**font_kwargs)
        
        # 背景色
        if background_color:
            cell.fill = PatternFill(
                start_color=_parse_color(background_color), 
                end_color=_parse_color(background_color),
                fill_type='solid'
            )
        
        # 对齐
        if horizontal_align or vertical_align:
            align_kwargs = {}
            if horizontal_align:
                align_kwargs['horizontal'] = horizontal_align
            if vertical_align:
                align_kwargs['vertical'] = vertical_align
            cell.alignment = Alignment(**align_kwargs)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已设置单元格 {_col_num_to_letter(col)}{row} 的样式')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'设置样式失败: {e}')
    finally:
        wb.close()


@tool(
    "set_range_style",
    "【推荐】批量设置范围内所有单元格的样式（一次调用设置整行/整列/区域，比逐个设置快10倍以上）。设置整行样式时：start_row=end_row=目标行号，start_col=1，end_col=总列数。",
    {
        "file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int,
        "font_color": str, "background_color": str, "bold": bool,
        "italic": bool, "font_size": int, "horizontal_align": str,
        "vertical_align": str, "sheet_name": str
    }
)
async def set_range_style(args: Dict[str, Any]) -> Dict[str, Any]:
    """【推荐】批量设置范围内所有单元格的样式，比逐个设置快10倍以上"""
    _log_tool_call("set_range_style", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    font_color = args.get("font_color")
    background_color = args.get("background_color")
    bold = args.get("bold")
    italic = args.get("italic")
    font_size = args.get("font_size")
    horizontal_align = args.get("horizontal_align")
    vertical_align = args.get("vertical_align")
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 预构建样式对象
        font = None
        if any([font_color, bold is not None, italic is not None, font_size]):
            font_kwargs = {}
            if font_color:
                font_kwargs['color'] = _parse_color(font_color)
            if bold is not None:
                font_kwargs['bold'] = bold
            if italic is not None:
                font_kwargs['italic'] = italic
            if font_size:
                font_kwargs['size'] = font_size
            font = Font(**font_kwargs)
        
        fill = None
        if background_color:
            fill = PatternFill(
                start_color=_parse_color(background_color),
                end_color=_parse_color(background_color),
                fill_type='solid'
            )
        
        alignment = None
        if horizontal_align or vertical_align:
            align_kwargs = {}
            if horizontal_align:
                align_kwargs['horizontal'] = horizontal_align
            if vertical_align:
                align_kwargs['vertical'] = vertical_align
            alignment = Alignment(**align_kwargs)
        
        # 应用样式
        for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                 min_col=start_col, max_col=end_col):
            for cell in row:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已设置范围 {_col_num_to_letter(start_col)}{start_row}:{_col_num_to_letter(end_col)}{end_row} 的样式')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'批量设置样式失败: {e}')
    finally:
        wb.close()


# ==============================================================================
# 行列操作
# ==============================================================================

@tool(
    "insert_rows",
    "插入行",
    {"file_id": str, "row": int, "count": int, "sheet_name": str}
)
async def insert_rows(args: Dict[str, Any]) -> Dict[str, Any]:
    """插入行"""
    _log_tool_call("insert_rows", args)
    file_id = args["file_id"]
    row = args["row"]
    count = args.get("count", 1)
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.insert_rows(row, count)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已在第 {row} 行前插入 {count} 行')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'插入行失败: {e}')
    finally:
        wb.close()


@tool(
    "delete_rows",
    "删除行",
    {"file_id": str, "row": int, "count": int, "sheet_name": str}
)
async def delete_rows(args: Dict[str, Any]) -> Dict[str, Any]:
    """删除行"""
    _log_tool_call("delete_rows", args)
    file_id = args["file_id"]
    row = args["row"]
    count = args.get("count", 1)
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.delete_rows(row, count)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已删除第 {row} 行起的 {count} 行')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'删除行失败: {e}')
    finally:
        wb.close()


@tool(
    "insert_columns",
    "插入列",
    {"file_id": str, "col": int, "count": int, "sheet_name": str}
)
async def insert_columns(args: Dict[str, Any]) -> Dict[str, Any]:
    """插入列"""
    _log_tool_call("insert_columns", args)
    file_id = args["file_id"]
    col = args["col"]
    count = args.get("count", 1)
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.insert_cols(col, count)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已在第 {_col_num_to_letter(col)} 列前插入 {count} 列')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'插入列失败: {e}')
    finally:
        wb.close()


@tool(
    "delete_columns",
    "删除列",
    {"file_id": str, "col": int, "count": int, "sheet_name": str}
)
async def delete_columns(args: Dict[str, Any]) -> Dict[str, Any]:
    """删除列"""
    _log_tool_call("delete_columns", args)
    file_id = args["file_id"]
    col = args["col"]
    count = args.get("count", 1)
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.delete_cols(col, count)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已删除第 {_col_num_to_letter(col)} 列起的 {count} 列')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'删除列失败: {e}')
    finally:
        wb.close()


@tool(
    "set_column_width",
    "设置列宽",
    {"file_id": str, "col": int, "width": float, "sheet_name": str}
)
async def set_column_width(args: Dict[str, Any]) -> Dict[str, Any]:
    """设置列宽"""
    _log_tool_call("set_column_width", args)
    file_id = args["file_id"]
    col = args["col"]
    width = args["width"]
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.column_dimensions[_col_num_to_letter(col)].width = width
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已设置第 {_col_num_to_letter(col)} 列宽度为 {width}')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'设置列宽失败: {e}')
    finally:
        wb.close()


@tool(
    "set_row_height",
    "设置行高",
    {"file_id": str, "row": int, "height": float, "sheet_name": str}
)
async def set_row_height(args: Dict[str, Any]) -> Dict[str, Any]:
    """设置行高"""
    _log_tool_call("set_row_height", args)
    file_id = args["file_id"]
    row = args["row"]
    height = args["height"]
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.row_dimensions[row].height = height
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已设置第 {row} 行高度为 {height}')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'设置行高失败: {e}')
    finally:
        wb.close()


# ==============================================================================
# 数据操作
# ==============================================================================

@tool(
    "sort_range",
    "对范围内的数据排序",
    {
        "file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int,
        "sort_col": int, "ascending": bool, "has_header": bool, "sheet_name": str
    }
)
async def sort_range(args: Dict[str, Any]) -> Dict[str, Any]:
    """对范围内的数据排序"""
    _log_tool_call("sort_range", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    sort_col = args["sort_col"]
    ascending = args.get("ascending", True)
    has_header = args.get("has_header", True)
    sheet_name = args.get("sheet_name")
    
    logger.info(f'排序操作: file_id={file_id}, range={_col_num_to_letter(start_col)}{start_row}:{_col_num_to_letter(end_col)}{end_row}, sort_col={_col_num_to_letter(sort_col)}, ascending={ascending}')
    
    wb = _get_workbook(file_id)
    if not wb:
        logger.error(f'无法打开文件: file_id={file_id}')
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 读取数据
        data_start = start_row + 1 if has_header else start_row
        data = []
        logger.debug(f'开始读取数据: range={data_start}:{end_row}')
        for row in ws.iter_rows(min_row=data_start, max_row=end_row,
                                 min_col=start_col, max_col=end_col):
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        logger.info(f'数据读取完成: 共 {len(data)} 行')
        
        # 计算排序列在范围内的索引
        sort_idx = sort_col - start_col
        
        # 排序
        def sort_key(x):
            val = x[sort_idx]
            if val is None:
                return (1, '')  # None 排最后
            elif isinstance(val, (int, float)):
                return (0, val)
            else:
                return (0, str(val))
        
        logger.debug(f'开始排序: 排序列索引={sort_idx}')
        data.sort(key=sort_key, reverse=not ascending)
        logger.debug(f'排序完成')
        
        # 写回
        logger.debug(f'开始写回数据')
        for r_idx, row_data in enumerate(data):
            for c_idx, value in enumerate(row_data):
                _write_numeric_safe_cell(ws, data_start + r_idx, start_col + c_idx, value)
        
        if _save_workbook(wb, file_id):
            logger.info(f'排序操作成功: file_id={file_id}, 排序了 {len(data)} 行')
            return _create_result(True, f'已按第 {_col_num_to_letter(sort_col)} 列{"升序" if ascending else "降序"}排序')
        else:
            logger.error(f'保存文件失败: file_id={file_id}')
            return _create_result(False, '保存文件失败')
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.error(f'排序失败: file_id={file_id}, 错误: {str(e)}')
        logger.debug(f'错误详情:\n{error_details}')
        return _create_result(False, f'排序失败: {e}')
    finally:
        wb.close()


@tool(
    "find_replace",
    "查找替换",
    {
        "file_id": str, "find_text": str, "replace_text": str,
        "match_case": bool, "match_entire_cell": bool, "sheet_name": str
    }
)
async def find_replace(args: Dict[str, Any]) -> Dict[str, Any]:
    """查找替换"""
    _log_tool_call("find_replace", args)
    file_id = args["file_id"]
    find_text = args["find_text"]
    replace_text = args["replace_text"]
    match_case = args.get("match_case", False)
    match_entire_cell = args.get("match_entire_cell", False)
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        import re
        sheets = [wb[sheet_name]] if sheet_name else wb.worksheets
        replace_count = 0
        
        for ws in sheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    cell_str = str(cell.value)
                    find_str = find_text
                    
                    if not match_case:
                        cell_str_lower = cell_str.lower()
                        find_str_lower = find_str.lower()
                    else:
                        cell_str_lower = cell_str
                        find_str_lower = find_str
                    
                    if match_entire_cell:
                        if cell_str_lower == find_str_lower:
                            cell.value = replace_text
                            replace_count += 1
                    else:
                        if find_str_lower in cell_str_lower:
                            pattern = re.compile(re.escape(find_text), re.IGNORECASE if not match_case else 0)
                            new_value = pattern.sub(replace_text, cell_str)
                            if new_value != cell_str:
                                cell.value = new_value
                                replace_count += 1
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已完成替换，共替换 {replace_count} 处')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'查找替换失败: {e}')
    finally:
        wb.close()


@tool(
    "remove_duplicates",
    "删除重复行",
    {
        "file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int,
        "check_columns": list, "has_header": bool, "sheet_name": str
    }
)
async def remove_duplicates(args: Dict[str, Any]) -> Dict[str, Any]:
    """删除重复行"""
    _log_tool_call("remove_duplicates", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    check_columns = args.get("check_columns")
    has_header = args.get("has_header", True)
    sheet_name = args.get("sheet_name")
    
    # 解析可能是 JSON 字符串的参数
    if check_columns:
        try:
            check_columns = _parse_json_param(check_columns, 'check_columns')
        except ValueError as e:
            return _create_result(False, str(e))
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        
        data_start = start_row + 1 if has_header else start_row
        
        # 读取数据
        rows_data = []
        for row in ws.iter_rows(min_row=data_start, max_row=end_row,
                                 min_col=start_col, max_col=end_col):
            row_data = [cell.value for cell in row]
            rows_data.append(row_data)
        
        # 确定检查列
        if check_columns:
            check_indices = [c - start_col for c in check_columns]
        else:
            check_indices = list(range(end_col - start_col + 1))
        
        # 去重
        seen = set()
        unique_rows = []
        removed_count = 0
        
        for row_data in rows_data:
            key = tuple(row_data[i] for i in check_indices if i < len(row_data))
            if key not in seen:
                seen.add(key)
                unique_rows.append(row_data)
            else:
                removed_count += 1
        
        # 清空原数据区域
        for row in range(data_start, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col, value=None)
        
        # 写回去重后的数据
        for r_idx, row_data in enumerate(unique_rows):
            for c_idx, value in enumerate(row_data):
                _write_numeric_safe_cell(ws, data_start + r_idx, start_col + c_idx, value)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已删除 {removed_count} 行重复数据')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'删除重复行失败: {e}')
    finally:
        wb.close()


# ==============================================================================
# 工作表操作
# ==============================================================================

@tool(
    "add_sheet",
    "添加新工作表",
    {"file_id": str, "sheet_name": str, "position": int}
)
async def add_sheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """添加新工作表"""
    _log_tool_call("add_sheet", args)
    file_id = args["file_id"]
    sheet_name = args["sheet_name"]
    position = args.get("position")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        if sheet_name in wb.sheetnames:
            return _create_result(False, f'工作表 "{sheet_name}" 已存在')
        
        if position is not None:
            wb.create_sheet(sheet_name, position)
        else:
            wb.create_sheet(sheet_name)
        
        # 更新元数据
        meta = large_file_storage.get_metadata(file_id)
        if meta:
            meta.sheet_names = [name for name in (wb.sheetnames or []) if name and name != "__SHEETBOT_META__"]
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已创建工作表 "{sheet_name}"')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'添加工作表失败: {e}')
    finally:
        wb.close()


@tool(
    "rename_sheet",
    "重命名工作表",
    {"file_id": str, "old_name": str, "new_name": str}
)
async def rename_sheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """重命名工作表"""
    _log_tool_call("rename_sheet", args)
    file_id = args["file_id"]
    old_name = args["old_name"]
    new_name = args["new_name"]
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        if old_name not in wb.sheetnames:
            return _create_result(False, f'工作表 "{old_name}" 不存在')
        if new_name in wb.sheetnames:
            return _create_result(False, f'工作表 "{new_name}" 已存在')
        
        wb[old_name].title = new_name
        
        # 更新元数据
        meta = large_file_storage.get_metadata(file_id)
        if meta:
            meta.sheet_names = [name for name in (wb.sheetnames or []) if name and name != "__SHEETBOT_META__"]
            if old_name in meta.sheet_row_counts:
                meta.sheet_row_counts[new_name] = meta.sheet_row_counts.pop(old_name)
        
        if _save_workbook(wb, file_id):
            # 同步重命名 DuckDB 内存表（未加载时静默跳过）
            from .large_file_duckdb import duckdb_manager
            duckdb_manager.rename_sheet(file_id, old_name, new_name)
            return _create_result(True, f'已将工作表 "{old_name}" 重命名为 "{new_name}"')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'重命名工作表失败: {e}')
    finally:
        wb.close()


@tool(
    "delete_sheet",
    "删除工作表",
    {"file_id": str, "sheet_name": str}
)
async def delete_sheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """删除工作表"""
    _log_tool_call("delete_sheet", args)
    file_id = args["file_id"]
    sheet_name = args["sheet_name"]
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        if sheet_name not in wb.sheetnames:
            return _create_result(False, f'工作表 "{sheet_name}" 不存在')
        if len(wb.sheetnames) <= 1:
            return _create_result(False, '不能删除唯一的工作表')
        
        del wb[sheet_name]
        
        # 更新元数据
        meta = large_file_storage.get_metadata(file_id)
        if meta:
            meta.sheet_names = [name for name in (wb.sheetnames or []) if name and name != "__SHEETBOT_META__"]
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已删除工作表 "{sheet_name}"')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'删除工作表失败: {e}')
    finally:
        wb.close()


@tool(
    "copy_sheet",
    "复制工作表",
    {"file_id": str, "source_name": str, "target_name": str}
)
async def copy_sheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """复制工作表"""
    _log_tool_call("copy_sheet", args)
    file_id = args["file_id"]
    source_name = args["source_name"]
    target_name = args["target_name"]
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        if source_name not in wb.sheetnames:
            return _create_result(False, f'工作表 "{source_name}" 不存在')
        if target_name in wb.sheetnames:
            return _create_result(False, f'工作表 "{target_name}" 已存在')
        
        source = wb[source_name]
        target = wb.copy_worksheet(source)
        target.title = target_name
        
        # 更新元数据
        meta = large_file_storage.get_metadata(file_id)
        if meta:
            meta.sheet_names = [name for name in (wb.sheetnames or []) if name and name != "__SHEETBOT_META__"]
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已复制工作表 "{source_name}" 为 "{target_name}"')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'复制工作表失败: {e}')
    finally:
        wb.close()


# ==============================================================================
# 合并单元格
# ==============================================================================

@tool(
    "merge_cells",
    "合并单元格",
    {"file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int, "sheet_name": str}
)
async def merge_cells(args: Dict[str, Any]) -> Dict[str, Any]:
    """合并单元格"""
    _log_tool_call("merge_cells", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=end_row, end_column=end_col)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已合并单元格 {_col_num_to_letter(start_col)}{start_row}:{_col_num_to_letter(end_col)}{end_row}')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'合并单元格失败: {e}')
    finally:
        wb.close()


@tool(
    "unmerge_cells",
    "取消合并单元格",
    {"file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int, "sheet_name": str}
)
async def unmerge_cells(args: Dict[str, Any]) -> Dict[str, Any]:
    """取消合并单元格"""
    _log_tool_call("unmerge_cells", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.unmerge_cells(start_row=start_row, start_column=start_col,
                         end_row=end_row, end_column=end_col)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已取消合并单元格 {_col_num_to_letter(start_col)}{start_row}:{_col_num_to_letter(end_col)}{end_row}')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'取消合并单元格失败: {e}')
    finally:
        wb.close()


# ==============================================================================
# 条件格式
# ==============================================================================

@tool(
    "add_conditional_format",
    "添加条件格式",
    {
        "file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int,
        "condition_type": str, "condition_value": str,
        "font_color": str, "background_color": str, "bold": bool, "sheet_name": str
    }
)
async def add_conditional_format(args: Dict[str, Any]) -> Dict[str, Any]:
    """添加条件格式"""
    _log_tool_call("add_conditional_format", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    condition_type = args["condition_type"]
    condition_value = args["condition_value"]
    font_color = args.get("font_color")
    background_color = args.get("background_color")
    bold = args.get("bold", False)
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 构建样式
        font = Font(color=_parse_color(font_color), bold=bold) if font_color else Font(bold=bold) if bold else None
        fill = PatternFill(
            start_color=_parse_color(background_color),
            end_color=_parse_color(background_color),
            fill_type='solid'
        ) if background_color else None
        
        # 范围字符串
        range_str = f'{_col_num_to_letter(start_col)}{start_row}:{_col_num_to_letter(end_col)}{end_row}'
        
        # 创建规则
        rule = CellIsRule(
            operator=condition_type,
            formula=[str(condition_value)],
            font=font,
            fill=fill
        )
        
        ws.conditional_formatting.add(range_str, rule)
        
        if _save_workbook(wb, file_id):
            return _create_result(True, f'已为范围 {range_str} 添加条件格式')
        else:
            return _create_result(False, '保存文件失败')
    except Exception as e:
        return _create_result(False, f'添加条件格式失败: {e}')
    finally:
        wb.close()


# ==============================================================================
# 统计分析
# ==============================================================================

@tool(
    "calculate_statistics",
    "计算范围内的统计信息（计数、求和、平均、最大、最小）",
    {"file_id": str, "start_row": int, "start_col": int, "end_row": int, "end_col": int, "sheet_name": str}
)
async def calculate_statistics(args: Dict[str, Any]) -> Dict[str, Any]:
    """计算范围内的统计信息"""
    _log_tool_call("calculate_statistics", args)
    file_id = args["file_id"]
    start_row = args["start_row"]
    start_col = args["start_col"]
    end_row = args["end_row"]
    end_col = args["end_col"]
    sheet_name = args.get("sheet_name")
    
    wb = _get_workbook(file_id, read_only=True)
    if not wb:
        return _create_result(False, f'无法打开文件: {file_id}')
    
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        
        values = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                 min_col=start_col, max_col=end_col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    values.append(cell.value)
        
        wb.close()
        
        if not values:
            return _create_result(True, '范围内没有数值数据', {
                'count': 0,
                'sum': 0,
                'average': None,
                'max': None,
                'min': None,
            })
        
        return _create_result(True, '统计计算完成', {
            'count': len(values),
            'sum': sum(values),
            'average': sum(values) / len(values),
            'max': max(values),
            'min': min(values),
        })
    except Exception as e:
        wb.close()
        return _create_result(False, f'统计计算失败: {e}')


# ==============================================================================
# DuckDB 高性能工具（数据查询和聚合）
# ==============================================================================

from .large_file_duckdb import duckdb_manager


@tool(
    "query_data",
    "【高性能】使用 SQL 查询 Excel 数据。支持完整 SQL 语法。使用 {table} 引用当前工作表，使用 {table:工作表名} 实现跨表 JOIN 查询（包括结果工作表）。",
    {"file_id": str, "sql": str, "sheet_name": str}
)
async def query_data(args: Dict[str, Any]) -> Dict[str, Any]:
    """使用 SQL 查询 Excel 数据（DuckDB 引擎）"""
    _log_tool_call("query_data", args)
    file_id = args["file_id"]
    sql = args["sql"]
    sheet_name = args.get("sheet_name")
    steps = []  # 执行步骤记录
    start_time = time.time()
    actual_sql = None
    
    try:
        # ========================================
        # 检查是否为结果表查询（两种情况）
        # ========================================
        # 1. SQL 中使用 {table:结果_xxx} 格式引用结果表
        # 2. sheet_name 参数以 "结果_" 开头（使用 {table} 占位符）
        import re
        result_table_refs = re.findall(r'\{table:(结果_[^}]+)\}', sql)  # 提取所有结果表引用
        is_result_sheet = sheet_name and sheet_name.startswith('结果_')
        uses_result_table = bool(result_table_refs) or is_result_sheet
        resolved_sheet_name = sheet_name
        
        # 如果是结果表查询，验证并从 DuckDB 内存加载
        if uses_result_table:
            steps.append("🔗 检测到结果表引用，从内存加载...")
            
            # 验证 SQL 中引用的所有结果表是否存在（方式 A）
            for ref_name in result_table_refs:
                result_table_name = duckdb_manager.get_result_table_name(file_id, ref_name)
                if not result_table_name:
                    return _create_result(False, f'结果表 [{ref_name}] 不在内存中，可能已过期或未创建')
                logger.info(f'[DuckDB] SQL 引用的结果表已验证: {ref_name} -> {result_table_name}')
            
            # 验证 sheet_name 参数指定的结果表是否存在（方式 B）
            if is_result_sheet:
                result_table_name = duckdb_manager.get_result_table_name(file_id, sheet_name)
                if not result_table_name:
                    return _create_result(False, f'结果表 [{sheet_name}] 不在内存中，可能已过期或未创建')
                logger.info(f'[DuckDB] 使用内存结果表: {result_table_name}')
        # 否则，确保源数据已加载
        else:
            resolved_sheet_name = _resolve_source_sheet_name(file_id, sheet_name)
            if not resolved_sheet_name:
                candidates = _list_available_source_sheets(file_id)
                return _create_result(
                    False,
                    f'工作表 "{sheet_name}" 不存在，可选工作表: {candidates[:12]}',
                )
            if resolved_sheet_name != sheet_name:
                steps.append(f'🔁 工作表名自动纠偏: {sheet_name} -> {resolved_sheet_name}')
                logger.info(
                    '[DuckDB] 工作表名纠偏: file_id=%s requested=%s resolved=%s',
                    file_id, sheet_name, resolved_sheet_name
                )

        if not uses_result_table and not duckdb_manager.is_loaded(file_id, resolved_sheet_name):
            load_start = time.time()
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={resolved_sheet_name}')
            steps.append("📂 正在将 Excel 数据加载到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, resolved_sheet_name)
            load_time = (time.time() - load_start) * 1000
            steps.append(f"✅ 数据加载完成 (耗时 {load_time:.0f}ms)")
        
        # 解析 SQL 占位符（用于显示实际执行的 SQL）
        try:
            actual_sql = duckdb_manager._resolve_table_placeholders(file_id, sql, resolved_sheet_name)
        except Exception:
            actual_sql = sql  # 解析失败时使用原始 SQL
        
        # 执行查询
        query_start = time.time()
        logger.info(f'[DuckDB] 执行查询: file_id={file_id}, sql={sql[:100]}...')
        sql_preview = sql[:200] + '...' if len(sql) > 200 else sql
        steps.append(f"🔍 正在执行 SQL: {sql_preview}")
        result = duckdb_manager.query(file_id, sql, resolved_sheet_name)
        query_time = (time.time() - query_start) * 1000
        steps.append(f"✅ 查询完成: {len(result)} 行数据 (耗时 {query_time:.0f}ms)")
        
        total_time = (time.time() - start_time) * 1000
        logger.info(f'[DuckDB] query_data 完成: 返回 {len(result)} 行, 总耗时 {total_time:.0f}ms')
        return _create_result(
            True, 
            f'查询成功，返回 {len(result)} 行数据', 
            {
                'data': result,
                'row_count': len(result),
                'col_count': len(result[0]) if result else 0
            }, 
            steps=steps,
            execution_time_ms=total_time,
            sql_executed=actual_sql
        )
    except Exception as e:
        repaired_sql = _try_repair_group_by_binder_sql(sql, str(e))
        if repaired_sql and repaired_sql.strip() != sql.strip():
            try:
                logger.warning(
                    '[DuckDB] query_data 命中 GROUP BY 绑定错误，自动重试修复: file_id=%s',
                    file_id
                )
                steps.append("⚠️ 检测到 GROUP BY 语义缺失，自动注入 GROUP BY ALL 重试...")
                actual_sql = duckdb_manager._resolve_table_placeholders(file_id, repaired_sql, resolved_sheet_name)
                retry_start = time.time()
                result = duckdb_manager.query(file_id, repaired_sql, resolved_sheet_name)
                retry_time = (time.time() - retry_start) * 1000
                steps.append(f"✅ 自动修复后查询成功: {len(result)} 行 (耗时 {retry_time:.0f}ms)")
                total_time = (time.time() - start_time) * 1000
                return _create_result(
                    True,
                    f'查询成功，返回 {len(result)} 行数据',
                    {
                        'data': result,
                        'row_count': len(result),
                        'col_count': len(result[0]) if result else 0,
                    },
                    steps=steps,
                    execution_time_ms=total_time,
                    sql_executed=actual_sql,
                )
            except Exception as retry_err:
                logger.error(f'[DuckDB] query_data 自动修复重试失败: {retry_err}')
                steps.append(f"❌ 自动修复重试失败: {retry_err}")

        total_time = (time.time() - start_time) * 1000
        logger.error(f'[DuckDB] query_data 失败: {e}')
        steps.append(f"❌ 查询失败: {e}")
        return _create_result(False, f'查询失败: {e}', steps=steps, execution_time_ms=total_time, sql_executed=actual_sql)


@tool(
    "get_unique_values",
    "【高性能】获取指定列的所有唯一值。适用于获取分类字段的可选值列表。",
    {"file_id": str, "column": str, "sheet_name": str, "limit": int}
)
async def get_unique_values_duckdb(args: Dict[str, Any]) -> Dict[str, Any]:
    """获取指定列的唯一值（DuckDB 引擎）"""
    _log_tool_call("get_unique_values", args)
    file_id = args["file_id"]
    column = args["column"]
    sheet_name = args.get("sheet_name")
    limit = args.get("limit", 1000)
    steps = []
    
    try:
        # 确保文件已加载
        if not duckdb_manager.is_loaded(file_id, sheet_name):
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={sheet_name}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, sheet_name)
            steps.append("✅ 数据加载完成")
        
        logger.info(f'[DuckDB] 获取唯一值: file_id={file_id}, column={column}')
        steps.append(f"🔍 正在提取列 [{column}] 的唯一值...")
        values = duckdb_manager.get_unique_values(file_id, column, sheet_name, limit)
        steps.append(f"✅ 找到 {len(values)} 个唯一值")
        
        logger.info(f'[DuckDB] get_unique_values 完成: 列={column}, 唯一值数={len(values)}')
        return _create_result(True, f'获取到 {len(values)} 个唯一值', {
            'column': column,
            'values': values,
            'count': len(values)
        }, steps=steps)
    except Exception as e:
        logger.error(f'[DuckDB] get_unique_values 失败: {e}')
        steps.append(f"❌ 获取唯一值失败: {e}")
        return _create_result(False, f'获取唯一值失败: {e}', steps=steps)


@tool(
    "get_column_statistics",
    "【高性能】获取指定列的统计信息，包括计数、唯一值数、最小值、最大值、平均值、求和等。",
    {"file_id": str, "column": str, "sheet_name": str}
)
async def get_column_statistics(args: Dict[str, Any]) -> Dict[str, Any]:
    """获取列统计信息（DuckDB 引擎）"""
    _log_tool_call("get_column_statistics", args)
    file_id = args["file_id"]
    column = args["column"]
    sheet_name = args.get("sheet_name")
    steps = []
    
    try:
        if not duckdb_manager.is_loaded(file_id, sheet_name):
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={sheet_name}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, sheet_name)
            steps.append("✅ 数据加载完成")
        
        logger.info(f'[DuckDB] 获取列统计: file_id={file_id}, column={column}')
        steps.append(f"📊 正在计算列 [{column}] 的统计信息...")
        stats = duckdb_manager.get_statistics(file_id, column, sheet_name)
        steps.append(f"✅ 统计完成: {stats.get('count')} 条记录")
        
        logger.info(f'[DuckDB] get_column_statistics 完成: 列={column}, count={stats.get("count")}')
        return _create_result(True, f'列 {column} 的统计信息', stats, steps=steps)
    except Exception as e:
        logger.error(f'[DuckDB] get_column_statistics 失败: {e}')
        steps.append(f"❌ 获取统计信息失败: {e}")
        return _create_result(False, f'获取统计信息失败: {e}', steps=steps)


@tool(
    "create_pivot_table",
    "【高性能】创建数据透视表。使用 DuckDB PIVOT 语法，自动计算聚合结果。",
    {"file_id": str, "row_field": str, "column_field": str, "value_field": str, "agg_func": str, "sheet_name": str}
)
async def create_pivot_table(args: Dict[str, Any]) -> Dict[str, Any]:
    """创建数据透视表（DuckDB 引擎）"""
    _log_tool_call("create_pivot_table", args)
    file_id = args["file_id"]
    row_field = args["row_field"]
    column_field = args["column_field"]
    value_field = args["value_field"]
    agg_func = args.get("agg_func", "SUM")
    sheet_name = args.get("sheet_name")
    steps = []
    
    try:
        if not duckdb_manager.is_loaded(file_id, sheet_name):
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={sheet_name}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, sheet_name)
            steps.append("✅ 数据加载完成")
        
        logger.info(f'[DuckDB] 创建透视表: row={row_field}, col={column_field}, val={value_field}, agg={agg_func}')
        steps.append(f"📊 正在创建透视表: 行={row_field}, 列={column_field}, 值={value_field}({agg_func})...")
        result = duckdb_manager.create_pivot_table(
            file_id, row_field, column_field, value_field, agg_func, sheet_name
        )
        steps.append(f"✅ 透视表创建完成: {result['row_count']} 行 x {result['col_count']} 列")
        
        logger.info(f'[DuckDB] create_pivot_table 完成: {result["row_count"]} 行 x {result["col_count"]} 列')
        return _create_result(True, f'透视表创建成功: {result["row_count"]} 行 x {result["col_count"]} 列', result, steps=steps)
    except Exception as e:
        logger.error(f'[DuckDB] create_pivot_table 失败: {e}')
        steps.append(f"❌ 创建透视表失败: {e}")
        return _create_result(False, f'创建透视表失败: {e}', steps=steps)


@tool(
    "group_by_aggregate",
    "【高性能】执行分组聚合操作。支持多列分组和多个聚合表达式。",
    {"file_id": str, "group_columns": list, "agg_expressions": dict, "sheet_name": str, "having": str, "order_by": str}
)
async def group_by_aggregate(args: Dict[str, Any]) -> Dict[str, Any]:
    """执行分组聚合（DuckDB 引擎）"""
    _log_tool_call("group_by_aggregate", args)
    file_id = args["file_id"]
    group_columns = args["group_columns"]
    agg_expressions = args["agg_expressions"]
    sheet_name = args.get("sheet_name")
    having = args.get("having")
    order_by = args.get("order_by")
    steps = []
    
    # 解析 JSON 字符串参数
    try:
        group_columns = _parse_json_param(group_columns, 'group_columns')
        agg_expressions = _parse_json_param(agg_expressions, 'agg_expressions')
    except ValueError as e:
        return _create_result(False, str(e))
    
    try:
        if not duckdb_manager.is_loaded(file_id, sheet_name):
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={sheet_name}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, sheet_name)
            steps.append("✅ 数据加载完成")
        
        logger.info(f'[DuckDB] 分组聚合: group_by={group_columns}, agg={list(agg_expressions.keys())}')
        steps.append(f"📊 正在执行分组聚合: 分组={group_columns}...")
        result = duckdb_manager.group_by(
            file_id, group_columns, agg_expressions, sheet_name, having, order_by
        )
        steps.append(f"✅ 聚合完成: 返回 {len(result)} 行")
        
        logger.info(f'[DuckDB] group_by_aggregate 完成: 返回 {len(result)} 行')
        return _create_result(True, f'聚合成功，返回 {len(result)} 行', {
            'data': result,
            'row_count': len(result)
        }, steps=steps)
    except Exception as e:
        logger.error(f'group_by_aggregate 失败: {e}')
        return _create_result(False, f'分组聚合失败: {e}')


@tool(
    "export_query_to_sheet",
    "【高性能】将 SQL 查询结果导出到结果文件的新工作表。支持 {table:工作表名} 跨表 JOIN（包括结果工作表的二次加工）。",
    {"file_id": str, "sql": str, "target_sheet": str, "source_sheet": str, "start_row": int, "start_col": int}
)
async def export_query_to_sheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """将查询结果导出到结果文件的新工作表（支持二次加工）"""
    _log_tool_call("export_query_to_sheet", args)
    file_id = args["file_id"]
    sql = args["sql"]
    source_sheet = args.get("source_sheet")
    steps = []
    start_time = time.time()
    actual_sql = None
    
    try:
        # ========================================
        # 检查是否为结果表查询（两种情况）
        # ========================================
        import re
        result_table_refs = re.findall(r'\{table:(结果_[^}]+)\}', sql)
        is_result_sheet = source_sheet and source_sheet.startswith('结果_')
        uses_result_table = bool(result_table_refs) or is_result_sheet
        
        # 如果是结果表查询，验证并从 DuckDB 内存加载
        if uses_result_table:
            steps.append("🔗 检测到结果表引用，从内存加载...")
            
            # 验证 SQL 中引用的所有结果表是否存在（方式 A）
            for ref_name in result_table_refs:
                result_table_name = duckdb_manager.get_result_table_name(file_id, ref_name)
                if not result_table_name:
                    return _create_result(False, f'结果表 [{ref_name}] 不在内存中，可能已过期或未创建')
                logger.info(f'[DuckDB] SQL 引用的结果表已验证: {ref_name} -> {result_table_name}')
            
            # 验证 source_sheet 参数指定的结果表是否存在（方式 B）
            if is_result_sheet:
                result_table_name = duckdb_manager.get_result_table_name(file_id, source_sheet)
                if not result_table_name:
                    return _create_result(False, f'结果表 [{source_sheet}] 不在内存中，可能已过期或未创建')
                logger.info(f'[DuckDB] 使用内存结果表: {result_table_name}')
        # 否则，确保源数据已加载
        elif not duckdb_manager.is_loaded(file_id, source_sheet):
            load_start = time.time()
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={source_sheet}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, source_sheet)
            load_time = (time.time() - load_start) * 1000
            steps.append(f"✅ 数据加载完成 (耗时 {load_time:.0f}ms)")
        
        # 解析 SQL 占位符
        try:
            actual_sql = duckdb_manager._resolve_table_placeholders(file_id, sql, source_sheet)
        except Exception:
            actual_sql = sql
        
        # 执行查询获取数据
        query_start = time.time()
        logger.info(f'[DuckDB] 执行查询并导出到结果文件: sql={sql[:80]}...')
        sql_preview = sql[:200] + '...' if len(sql) > 200 else sql
        steps.append(f"🔍 正在执行 SQL: {sql_preview}")
        result_df = duckdb_manager.query_df(file_id, sql, source_sheet)
        columns = list(result_df.columns)
        data = result_df.values.tolist()
        query_time = (time.time() - query_start) * 1000
        steps.append(f"✅ 查询完成: {len(data)} 行数据 (耗时 {query_time:.0f}ms)")

        # 获取或创建结果文件
        result_filename = _build_result_filename(file_id)
        result_meta = await large_file_storage.get_or_create_result_file(file_id, result_filename)
        if not result_meta:
            return _create_result(False, "无法创建结果文件")

        sheet_name = _build_result_sheet_name("SQL查询")
        steps.append(f"📝 正在写入结果文件工作表 [{sheet_name}]...")
        write_start = time.time()
        final_name = await large_file_storage.append_sheet_to_result_file(
            result_meta.file_id,
            sheet_name,
            columns,
            data
        )
        if not final_name:
            return _create_result(False, "写入结果文件失败")
        write_time = (time.time() - write_start) * 1000
        steps.append(f"✅ 结果已写入文件: {len(data)} 行 x {len(columns)} 列 (耗时 {write_time:.0f}ms)")
        
        # 注册结果表到 DuckDB（支持二次加工）
        try:
            reg_start = time.time()
            duckdb_manager.register_result_table(file_id, final_name, columns, data)
            reg_time = (time.time() - reg_start) * 1000
            steps.append(f"🔗 结果已注册到内存，可用 {{table:{final_name}}} 进行二次加工 (耗时 {reg_time:.0f}ms)")
        except Exception as reg_error:
            logger.warning(f'注册结果表失败（不影响导出）: {reg_error}')
            steps.append(f"⚠️ 结果表注册失败（不影响已导出的数据）: {reg_error}")

        total_time = (time.time() - start_time) * 1000
        
        # 添加操作日志
        large_file_storage.add_operation_log(
            source_file_id=file_id,
            sheet_name=final_name,
            operation_type='SQL查询',
            logic=actual_sql or sql,
            logic_description='基于 SQL 查询生成结果表',
            row_count=len(data),
            execution_time_ms=total_time
        )
        # 同步操作日志到文件
        await large_file_storage.sync_operation_log_to_file(file_id)
        
        logger.info(f'[DuckDB] export_query_to_sheet 完成: result_file={result_meta.file_id}, sheet={final_name}, 总耗时={total_time:.0f}ms')
        return _create_result(
            True, 
            f'已导出 {len(data)} 行数据到结果文件工作表 {final_name}', 
            {
                'result_file_id': result_meta.file_id,
                'filename': result_meta.original_name,
                'sheet_name': final_name,
                'row_count': len(data),
                'col_count': len(columns),
                'columns': columns,
                'can_reprocess': True  # 标记支持二次加工
            }, 
            steps=steps,
            execution_time_ms=total_time,
            sql_executed=actual_sql
        )
    except Exception as e:
        total_time = (time.time() - start_time) * 1000
        logger.error(f'[DuckDB] export_query_to_sheet 失败: {e}')
        import traceback
        logger.debug(f'错误详情:\n{traceback.format_exc()}')
        steps.append(f"❌ 导出失败: {e}")
        return _create_result(False, f'导出失败: {e}', steps=steps, execution_time_ms=total_time, sql_executed=actual_sql)


@tool(
    "export_pivot_to_sheet",
    "【高性能】创建透视表并导出到新工作表。结合 DuckDB 高效计算和 openpyxl 格式化输出。结果支持二次加工。",
    {"file_id": str, "row_field": str, "column_field": str, "value_field": str, "agg_func": str, "target_sheet": str, "source_sheet": str, "start_row": int, "start_col": int, "include_formulas": bool}
)
async def export_pivot_to_sheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """创建透视表并导出到工作表（写入结果文件，支持二次加工）"""
    _log_tool_call("export_pivot_to_sheet", args)
    file_id = args["file_id"]
    row_field = args["row_field"]
    column_field = args["column_field"]
    value_field = args["value_field"]
    agg_func = args.get("agg_func", "SUM")
    source_sheet = args.get("source_sheet")
    include_formulas = args.get("include_formulas", False)
    steps = []
    start_time = time.time()
    
    try:
        # 确保源数据已加载
        if not duckdb_manager.is_loaded(file_id, source_sheet):
            load_start = time.time()
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={source_sheet}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, source_sheet)
            load_time = (time.time() - load_start) * 1000
            steps.append(f"✅ 数据加载完成 (耗时 {load_time:.0f}ms)")
        
        # 创建透视表
        pivot_start = time.time()
        logger.info(f'[DuckDB] 创建透视表并导出到结果文件: row={row_field}, col={column_field}, val={value_field}')
        steps.append(f"📊 正在计算透视表: 行={row_field}, 列={column_field}, 值={value_field}({agg_func})...")
        pivot_result = duckdb_manager.create_pivot_table(
            file_id, row_field, column_field, value_field, agg_func, source_sheet
        )
        
        columns = pivot_result['columns']
        data = pivot_result['data']
        pivot_time = (time.time() - pivot_start) * 1000
        steps.append(f"✅ 透视表计算完成: {len(data)} 行 x {len(columns)} 列 (耗时 {pivot_time:.0f}ms)")
        
        # 写入结果文件
        result_filename = _build_result_filename(file_id)
        result_meta = await large_file_storage.get_or_create_result_file(file_id, result_filename)
        if not result_meta:
            return _create_result(False, "无法创建结果文件")

        sheet_name = _build_result_sheet_name("透视表", [row_field, column_field, value_field])
        steps.append(f"📝 正在写入结果文件工作表 [{sheet_name}]...")
        write_start = time.time()
        final_name = await large_file_storage.append_sheet_to_result_file(
            result_meta.file_id,
            sheet_name,
            columns,
            data
        )
        if not final_name:
            return _create_result(False, "写入结果文件失败")
        write_time = (time.time() - write_start) * 1000
        steps.append(f"✅ 透视表已写入文件 (耗时 {write_time:.0f}ms)")

        if include_formulas:
            steps.append("ℹ️ 结果由 DuckDB 计算生成，未写入 Excel 公式")

        # 注册结果表到 DuckDB（支持二次加工）
        try:
            reg_start = time.time()
            duckdb_manager.register_result_table(file_id, final_name, columns, data)
            reg_time = (time.time() - reg_start) * 1000
            steps.append(f"🔗 结果已注册到内存，可用 {{table:{final_name}}} 进行二次加工 (耗时 {reg_time:.0f}ms)")
        except Exception as reg_error:
            logger.warning(f'注册结果表失败（不影响导出）: {reg_error}')
            steps.append(f"⚠️ 结果表注册失败（不影响已导出的数据）: {reg_error}")

        total_time = (time.time() - start_time) * 1000
        
        # 添加操作日志
        pivot_logic = f'行字段: {row_field}, 列字段: {column_field}, 值字段: {value_field}, 聚合函数: {agg_func}'
        large_file_storage.add_operation_log(
            source_file_id=file_id,
            sheet_name=final_name,
            operation_type='透视表',
            logic=pivot_logic,
            logic_description='按行列字段聚合生成透视结果',
            row_count=len(data),
            execution_time_ms=total_time
        )
        # 同步操作日志到文件
        await large_file_storage.sync_operation_log_to_file(file_id)
        
        logger.info(f'[DuckDB] export_pivot_to_sheet 完成: result_file={result_meta.file_id}, sheet={final_name}, 总耗时={total_time:.0f}ms')
        return _create_result(
            True, 
            f'透视表已导出到结果文件工作表 {final_name}: {len(data)} 行 x {len(columns)} 列', 
            {
                'result_file_id': result_meta.file_id,
                'filename': result_meta.original_name,
                'sheet_name': final_name,
                'row_count': len(data),
                'col_count': len(columns),
                'columns': columns,
                'row_field': row_field,
                'column_field': column_field,
                'value_field': value_field,
                'agg_func': agg_func,
                'can_reprocess': True  # 标记支持二次加工
            }, 
            steps=steps,
            execution_time_ms=total_time
        )
    except Exception as e:
        total_time = (time.time() - start_time) * 1000
        logger.error(f'export_pivot_to_sheet 失败: {e}')
        steps.append(f"❌ 导出透视表失败: {e}")
        return _create_result(False, f'导出透视表失败: {e}', steps=steps, execution_time_ms=total_time)


@tool(
    "export_statistics_to_sheet",
    "【高性能】计算数值列的统计信息并导出到新工作表。支持指定列或自动检测所有数值列。统计包括计数、唯一值数、最小值、最大值、平均值、总和。",
    {"file_id": str, "columns": list, "sheet_name": str}
)
async def export_statistics_to_sheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    计算统计信息并导出到新工作表
    
    Args:
        file_id: 文件ID
        columns: 要统计的列名列表（可选，为空则自动检测所有数值列）
        sheet_name: 源工作表名称
    """
    _log_tool_call("export_statistics_to_sheet", args)
    file_id = args["file_id"]
    columns = args.get("columns", [])
    sheet_name = args.get("sheet_name")
    steps = []
    start_time = time.time()
    
    # 解析 JSON 字符串参数
    try:
        columns = _parse_json_param(columns, 'columns') if columns else []
    except ValueError as e:
        return _create_result(False, str(e))
    
    try:
        # 确保文件已加载
        if not duckdb_manager.is_loaded(file_id, sheet_name):
            load_start = time.time()
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={sheet_name}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, sheet_name)
            load_time = (time.time() - load_start) * 1000
            steps.append(f"✅ 数据加载完成 (耗时 {load_time:.0f}ms)")
        
        # 获取表名用于查询
        table_name = duckdb_manager._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            return _create_result(False, f'工作表未加载: {sheet_name}')
        
        # 如果没有指定列，自动检测数值列
        if not columns:
            steps.append("🔍 正在检测数值列...")
            # 获取所有列的类型
            describe_result = duckdb_manager.conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            numeric_types = ['INTEGER', 'BIGINT', 'DOUBLE', 'FLOAT', 'DECIMAL', 'NUMERIC', 'REAL', 'SMALLINT', 'TINYINT']
            columns = [col[0] for col in describe_result if any(t in col[1].upper() for t in numeric_types)]
            if not columns:
                return _create_result(False, '未找到数值列，无法计算统计信息')
            steps.append(f"✅ 检测到 {len(columns)} 个数值列: {', '.join(columns[:5])}{'...' if len(columns) > 5 else ''}")
        
        # 计算每列的统计信息
        stats_start = time.time()
        steps.append(f"📊 正在计算 {len(columns)} 个列的统计信息...")
        
        stats_data = []
        for col in columns:
            try:
                # 使用 SQL 计算统计信息
                sql = f'''
                SELECT 
                    '{col}' as 列名,
                    COUNT("{col}") as 计数,
                    COUNT(DISTINCT "{col}") as 唯一值数,
                    MIN("{col}") as 最小值,
                    MAX("{col}") as 最大值,
                    ROUND(AVG("{col}")::DOUBLE, 2) as 平均值,
                    ROUND(SUM("{col}")::DOUBLE, 2) as 总和
                FROM "{table_name}"
                '''
                result = duckdb_manager.conn.execute(sql).fetchone()
                stats_data.append(list(result))
            except Exception as col_error:
                logger.warning(f'统计列 {col} 失败: {col_error}')
                stats_data.append([col, 0, 0, None, None, None, None])
        
        stats_time = (time.time() - stats_start) * 1000
        steps.append(f"✅ 统计计算完成 (耗时 {stats_time:.0f}ms)")
        
        # 准备导出数据
        result_columns = ['列名', '计数', '唯一值数', '最小值', '最大值', '平均值', '总和']
        
        # 写入结果文件
        result_filename = _build_result_filename(file_id)
        result_meta = await large_file_storage.get_or_create_result_file(file_id, result_filename)
        if not result_meta:
            return _create_result(False, "无法创建结果文件")
        
        result_sheet_name = _build_result_sheet_name("统计信息")
        steps.append(f"📝 正在写入结果文件工作表 [{result_sheet_name}]...")
        write_start = time.time()
        final_name = await large_file_storage.append_sheet_to_result_file(
            result_meta.file_id,
            result_sheet_name,
            result_columns,
            stats_data
        )
        if not final_name:
            return _create_result(False, "写入结果文件失败")
        write_time = (time.time() - write_start) * 1000
        steps.append(f"✅ 统计结果已写入文件 (耗时 {write_time:.0f}ms)")
        
        # 注册结果表到 DuckDB（支持二次加工）
        try:
            reg_start = time.time()
            duckdb_manager.register_result_table(file_id, final_name, result_columns, stats_data)
            reg_time = (time.time() - reg_start) * 1000
            steps.append(f"🔗 结果已注册到内存，可用 {{table:{final_name}}} 进行二次加工 (耗时 {reg_time:.0f}ms)")
        except Exception as reg_error:
            logger.warning(f'注册结果表失败（不影响导出）: {reg_error}')
        
        total_time = (time.time() - start_time) * 1000
        
        # 添加操作日志
        stats_logic = f'统计列: {", ".join(columns)}'
        large_file_storage.add_operation_log(
            source_file_id=file_id,
            sheet_name=final_name,
            operation_type='统计信息',
            logic=stats_logic,
            logic_description='统计列的计数/极值/均值等指标',
            row_count=len(stats_data),
            execution_time_ms=total_time
        )
        # 同步操作日志到文件
        await large_file_storage.sync_operation_log_to_file(file_id)
        
        logger.info(f'[DuckDB] export_statistics_to_sheet 完成: sheet={final_name}, 列数={len(columns)}, 总耗时={total_time:.0f}ms')
        
        return _create_result(
            True,
            f'统计信息已导出到结果文件工作表 [{final_name}]: {len(columns)} 个列的统计数据',
            {
                'result_file_id': result_meta.file_id,
                'filename': result_meta.original_name,
                'sheet_name': final_name,
                'row_count': len(stats_data),
                'col_count': len(result_columns),
                'columns': result_columns,
                'statistics_columns': columns,
                'can_reprocess': True
            },
            steps=steps,
            execution_time_ms=total_time
        )
    except Exception as e:
        total_time = (time.time() - start_time) * 1000
        logger.error(f'export_statistics_to_sheet 失败: {e}')
        import traceback
        logger.debug(f'错误详情:\n{traceback.format_exc()}')
        steps.append(f"❌ 导出统计信息失败: {e}")
        return _create_result(False, f'导出统计信息失败: {e}', steps=steps, execution_time_ms=total_time)


@tool(
    "get_data_preview",
    "【高性能】获取数据预览。比 get_range_values 快 10 倍以上，适合大文件。",
    {"file_id": str, "sheet_name": str, "limit": int}
)
async def get_data_preview(args: Dict[str, Any]) -> Dict[str, Any]:
    """获取数据预览（DuckDB 引擎）"""
    _log_tool_call("get_data_preview", args)
    file_id = args["file_id"]
    sheet_name = args.get("sheet_name")
    limit = args.get("limit", 500)
    steps = []
    
    try:
        # 检查是否是结果表（以"结果_"开头）
        is_result_sheet = sheet_name and sheet_name.startswith('结果_')
        
        if is_result_sheet:
            # 结果表：从 DuckDB 内存中获取
            result_table_name = duckdb_manager.get_result_table_name(file_id, sheet_name)
            if result_table_name:
                steps.append("🔗 从内存结果表加载...")
                logger.info(f'[DuckDB] 从结果表获取预览: table={result_table_name}')
                preview = duckdb_manager.get_preview_from_table(result_table_name, limit=limit, offset=0)
                steps.append(f"✅ 预览完成: {preview['preview_rows']} 行 / 总 {preview['row_count']} 行")
                return _create_result(True, f'获取预览成功: {preview["preview_rows"]} 行 / 总 {preview["row_count"]} 行', preview, steps=steps)
            else:
                # 结果表不在内存中
                return _create_result(False, f'结果表 {sheet_name} 不存在或已被清理，请重新创建')
        
        # 源数据表：需要从文件加载
        resolved_sheet_name = _resolve_source_sheet_name(file_id, sheet_name)
        if not resolved_sheet_name:
            candidates = _list_available_source_sheets(file_id)
            return _create_result(
                False,
                f'工作表 "{sheet_name}" 不存在，可选工作表: {candidates[:12]}',
            )
        if resolved_sheet_name != sheet_name:
            steps.append(f'🔁 工作表名自动纠偏: {sheet_name} -> {resolved_sheet_name}')
            logger.info(
                '[DuckDB] 工作表名纠偏: file_id=%s requested=%s resolved=%s',
                file_id, sheet_name, resolved_sheet_name
            )

        if not duckdb_manager.is_loaded(file_id, resolved_sheet_name):
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={resolved_sheet_name}')
            steps.append("📂 正在加载数据...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, resolved_sheet_name)
            steps.append("✅ 数据加载完成")
        
        logger.info(f'[DuckDB] 获取数据预览: file_id={file_id}, limit={limit}')
        steps.append(f"🔍 正在读取数据预览（前 {limit} 行）...")
        preview = duckdb_manager.get_preview(file_id, resolved_sheet_name, limit)
        steps.append(f"✅ 预览完成: {preview['preview_rows']} 行 / 总 {preview['row_count']} 行")
        
        logger.info(f'[DuckDB] get_data_preview 完成: {preview["preview_rows"]} 行 / 总 {preview["row_count"]} 行')
        return _create_result(True, f'获取预览成功: {preview["preview_rows"]} 行 / 总 {preview["row_count"]} 行', preview, steps=steps)
    except Exception as e:
        logger.error(f'[DuckDB] get_data_preview 失败: {e}')
        steps.append(f"❌ 获取预览失败: {e}")
        return _create_result(False, f'获取预览失败: {e}', steps=steps)


@tool(
    "export_query_to_new_file",
    "【高性能】将 SQL 查询结果导出到结果文件的新工作表。支持 {table:工作表名} 跨表 JOIN 查询。",
    {"file_id": str, "sql": str, "source_sheet": str, "filename": str}
)
async def export_query_to_new_file(args: Dict[str, Any]) -> Dict[str, Any]:
    """将查询结果导出到结果文件的新工作表"""
    _log_tool_call("export_query_to_new_file", args)
    file_id = args["file_id"]
    sql = args["sql"]
    source_sheet = args.get("source_sheet")
    filename = args.get("filename")
    steps = []
    
    try:
        # ========================================
        # 检查是否为结果表查询（两种情况）
        # ========================================
        import re
        result_table_refs = re.findall(r'\{table:(结果_[^}]+)\}', sql)
        is_result_sheet = source_sheet and source_sheet.startswith('结果_')
        uses_result_table = bool(result_table_refs) or is_result_sheet
        
        # 如果是结果表查询，验证并从 DuckDB 内存加载
        if uses_result_table:
            steps.append("🔗 检测到结果表引用，从内存加载...")
            
            # 验证 SQL 中引用的所有结果表是否存在（方式 A）
            for ref_name in result_table_refs:
                result_table_name = duckdb_manager.get_result_table_name(file_id, ref_name)
                if not result_table_name:
                    return _create_result(False, f'结果表 [{ref_name}] 不在内存中，可能已过期或未创建')
                logger.info(f'[DuckDB] SQL 引用的结果表已验证: {ref_name} -> {result_table_name}')
            
            # 验证 source_sheet 参数指定的结果表是否存在（方式 B）
            if is_result_sheet:
                result_table_name = duckdb_manager.get_result_table_name(file_id, source_sheet)
                if not result_table_name:
                    return _create_result(False, f'结果表 [{source_sheet}] 不在内存中，可能已过期或未创建')
                logger.info(f'[DuckDB] 使用内存结果表: {result_table_name}')
        # 否则，确保源数据已加载
        elif not duckdb_manager.is_loaded(file_id, source_sheet):
            logger.info(f'[DuckDB] 文件未缓存，开始加载: file_id={file_id}, sheet={source_sheet}')
            steps.append("📂 正在加载数据到高性能引擎...")
            file_path = large_file_storage.get_file_path(file_id)
            if not file_path:
                return _create_result(False, f'文件不存在: {file_id}')
            duckdb_manager.load_excel(str(file_path), file_id, source_sheet)
            steps.append("✅ 数据加载完成")
        
        # 执行查询获取数据
        logger.info(f'[DuckDB] 执行查询并导出: sql={sql[:80]}...')
        steps.append("🔍 正在执行 SQL 查询...")
        result_df = duckdb_manager.query_df(file_id, sql, source_sheet)
        columns = list(result_df.columns)
        data = result_df.values.tolist()
        steps.append(f"✅ 查询完成: {len(data)} 行数据")
        
        # 获取或创建结果文件
        result_filename = _build_result_filename(file_id, filename)
        result_meta = await large_file_storage.get_or_create_result_file(file_id, result_filename)
        if not result_meta:
            return _create_result(False, "无法创建结果文件")

        sheet_name = _build_result_sheet_name("SQL查询")
        steps.append(f"📝 正在写入结果文件工作表 [{sheet_name}]...")
        final_name = await large_file_storage.append_sheet_to_result_file(
            result_meta.file_id,
            sheet_name,
            columns,
            data
        )
        if not final_name:
            return _create_result(False, "写入结果文件失败")

        steps.append(f"✅ 结果已写入文件: {len(data)} 行 x {len(columns)} 列")
        
        # 注册结果表到 DuckDB（支持二次/三次加工）
        try:
            duckdb_manager.register_result_table(file_id, final_name, columns, data)
            steps.append(f"🔗 结果已注册到内存，可用 {{table:{final_name}}} 进行二次加工")
        except Exception as reg_error:
            logger.warning(f'注册结果表失败（不影响导出）: {reg_error}')
        
        # 添加操作日志
        large_file_storage.add_operation_log(
            source_file_id=file_id,
            sheet_name=final_name,
            operation_type='SQL查询',
            logic=sql,
            logic_description='基于 SQL 查询导出到新文件',
            row_count=len(data),
            execution_time_ms=0  # 这个工具没有计时，简化处理
        )
        # 同步操作日志到文件
        await large_file_storage.sync_operation_log_to_file(file_id)
        
        logger.info(f'[DuckDB] export_query_to_new_file 完成: result_file={result_meta.file_id}, sheet={final_name}')
        
        return _create_result(True, f'已导出 {len(data)} 行数据到结果文件工作表 {final_name}', {
            'result_file_id': result_meta.file_id,
            'filename': result_meta.original_name,
            'sheet_name': final_name,
            'row_count': len(data),
            'col_count': len(columns),
            'columns': columns,
            'can_reprocess': True  # 支持二次/三次加工
        }, steps=steps)
            
    except Exception as e:
        logger.error(f'[DuckDB] export_query_to_new_file 失败: {e}')
        import traceback
        logger.debug(f'错误详情:\n{traceback.format_exc()}')
        steps.append(f"❌ 导出失败: {e}")
        return _create_result(False, f'导出失败: {e}', steps=steps)


# ==============================================================================
# 工具列表
# ==============================================================================

# 所有工具函数列表
all_tools = [
    # DuckDB 高性能工具（优先使用）
    query_data, get_unique_values_duckdb, get_column_statistics,
    create_pivot_table, group_by_aggregate,
    export_query_to_sheet, export_pivot_to_sheet, export_statistics_to_sheet, export_query_to_new_file, get_data_preview,
    # openpyxl 读取（仅在需要时使用）
    get_file_info, get_sheet_info, get_cell_value, get_range_values,
    # openpyxl 写入
    set_cell_value, set_cell_formula, set_range_values,
    # openpyxl 样式
    set_cell_style, set_range_style,
    # openpyxl 行列
    insert_rows, delete_rows, insert_columns, delete_columns,
    set_column_width, set_row_height,
    # openpyxl 数据
    sort_range, find_replace, remove_duplicates,
    # openpyxl 工作表
    add_sheet, rename_sheet, delete_sheet, copy_sheet,
    # openpyxl 合并
    merge_cells, unmerge_cells,
    # openpyxl 条件格式
    add_conditional_format,
    # openpyxl 统计（低效，推荐用 get_column_statistics）
    calculate_statistics,
]

# 创建 MCP 服务器
large_file_mcp = create_sdk_mcp_server(
    name="large-file-tools",
    version="1.0.0",
    tools=all_tools
)

# 工具名称列表（用于 allowed_tools 配置）
LARGE_FILE_TOOL_NAMES = [
    # DuckDB 高性能工具（优先使用）
    'mcp__large-file-tools__query_data',
    'mcp__large-file-tools__get_unique_values',
    'mcp__large-file-tools__get_column_statistics',
    'mcp__large-file-tools__create_pivot_table',
    'mcp__large-file-tools__group_by_aggregate',
    'mcp__large-file-tools__export_query_to_sheet',
    'mcp__large-file-tools__export_pivot_to_sheet',
    'mcp__large-file-tools__export_statistics_to_sheet',  # 统计信息导出
    'mcp__large-file-tools__export_query_to_new_file',
    'mcp__large-file-tools__get_data_preview',
    # openpyxl 读取
    'mcp__large-file-tools__get_file_info',
    'mcp__large-file-tools__get_sheet_info',
    'mcp__large-file-tools__get_cell_value',
    'mcp__large-file-tools__get_range_values',
    # openpyxl 写入
    'mcp__large-file-tools__set_cell_value',
    'mcp__large-file-tools__set_cell_formula',
    'mcp__large-file-tools__set_range_values',
    # openpyxl 样式
    'mcp__large-file-tools__set_cell_style',
    'mcp__large-file-tools__set_range_style',
    # openpyxl 行列
    'mcp__large-file-tools__insert_rows',
    'mcp__large-file-tools__delete_rows',
    'mcp__large-file-tools__insert_columns',
    'mcp__large-file-tools__delete_columns',
    'mcp__large-file-tools__set_column_width',
    'mcp__large-file-tools__set_row_height',
    # openpyxl 数据
    'mcp__large-file-tools__sort_range',
    'mcp__large-file-tools__find_replace',
    'mcp__large-file-tools__remove_duplicates',
    # openpyxl 工作表
    'mcp__large-file-tools__add_sheet',
    'mcp__large-file-tools__rename_sheet',
    'mcp__large-file-tools__delete_sheet',
    'mcp__large-file-tools__copy_sheet',
    # openpyxl 合并
    'mcp__large-file-tools__merge_cells',
    'mcp__large-file-tools__unmerge_cells',
    # openpyxl 条件格式
    'mcp__large-file-tools__add_conditional_format',
    # openpyxl 统计
    'mcp__large-file-tools__calculate_statistics',
]
