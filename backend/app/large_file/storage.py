# ==============================================================================
# 大型文件存储管理
# 集成 DuckDB 高性能缓存层
# ==============================================================================
import os
import uuid
import asyncio
from io import BytesIO
from decimal import Decimal
from datetime import datetime, timedelta, date, time
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple, Callable
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from .schemas import (
    FileMetadata, FileStatus, 
    PREVIEW_ROW_COUNT,
)
from ..utils.logger import get_logger

logger = get_logger('large_file.storage')

# 系统内置元数据表，不对外暴露（分析视图等不展示）
META_SHEET_NAME = '__SHEETBOT_META__'

# DuckDB 管理器延迟导入，避免循环依赖
_duckdb_manager = None

def _get_duckdb_manager():
    """延迟获取 DuckDB 管理器"""
    global _duckdb_manager
    if _duckdb_manager is None:
        from .large_file_duckdb import duckdb_manager
        _duckdb_manager = duckdb_manager
    return _duckdb_manager


def _serialize_cell_value(val: Any) -> Any:
    """将单元格值转换为 JSON 可序列化的类型"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, time):
        return val.strftime('%H:%M:%S')
    if isinstance(val, timedelta):
        return str(val)
    return val


class LargeFileStorage:
    """大型文件存储管理器"""
    
    _instance: Optional['LargeFileStorage'] = None
    _lock = asyncio.Lock()
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        
        # 存储根目录：项目根目录/uploads
        self.base_dir = Path(__file__).parent.parent.parent.parent / 'uploads'
        self.base_dir.mkdir(parents=True, exist_ok=True)
        
        # 文件元数据缓存 {file_id: FileMetadata}
        self._metadata: Dict[str, FileMetadata] = {}

        # 源文件 -> 结果文件 映射（单一结果文件策略）
        self._result_file_map: Dict[str, str] = {}
        
        # 内存结果缓存：{source_file_id: {sheet_name: {table_name, columns, row_count, created_at}}}
        # 会话期间有效，关闭工作表时销毁
        self._memory_result_cache: Dict[str, Dict[str, Dict[str, Any]]] = {}
        
        # 操作日志缓存：{source_file_id: [{sheet_name, operation_type, logic, created_at, ...}]}
        # 用于追踪每个结果工作表的计算逻辑
        self._operation_log_cache: Dict[str, List[Dict[str, Any]]] = {}
        
        # 操作日志工作表名称（固定，放在第一个位置）
        self.OPERATION_LOG_SHEET_NAME = "_操作日志"
        
        logger.info(f'LargeFileStorage initialized, base_dir: {self.base_dir}')
    
    def _get_date_dir(self) -> Path:
        """获取当天的存储目录（uploads/excel_files/YYYY-MM-DD）"""
        date_str = datetime.now().strftime('%Y-%m-%d')
        date_dir = self.base_dir / 'excel_files' / date_str
        date_dir.mkdir(parents=True, exist_ok=True)
        return date_dir
    
    def register_placeholder(
        self,
        original_name: str,
        file_size: int,
        source_file_id: Optional[str] = None,
    ) -> FileMetadata:
        """
        预注册文件占位元数据（不写入磁盘）。
        用于 prepare 异步流程：先分配 file_id 返回给前端，后台再执行写入+加载。
        """
        file_id = str(uuid.uuid4())
        ext = Path(original_name).suffix or '.xlsx'
        date_dir = self._get_date_dir()
        file_path = date_dir / f'{file_id}{ext}'

        meta = FileMetadata(
            file_id=file_id,
            original_name=original_name,
            file_path=str(file_path),
            file_size=file_size,
            source_file_id=source_file_id,
            status=FileStatus.READY,
            sheet_names=[],
            sheet_row_counts={},
            row_count=0,
            col_count=0,
            duckdb_ready=False,
            duckdb_load_stage="正在读取文件...",
            duckdb_load_progress=3,
        )
        self._metadata[file_id] = meta
        return meta

    async def save_file(
        self,
        file_content: bytes,
        original_name: str,
        source_file_id: Optional[str] = None,
        preload_duckdb: bool = True,
        file_id_override: Optional[str] = None,
    ) -> FileMetadata:
        """
        保存上传的文件

        Args:
            file_content: 文件内容
            original_name: 原始文件名
            source_file_id: 结果文件关联的源文件ID
            file_id_override: 复用已有的 file_id（由 register_placeholder 预分配）

        Returns:
            FileMetadata: 文件元数据
        """
        file_size = len(file_content)
        file_size_mb = file_size / 1024 / 1024

        # 复用已注册的占位 metadata，或新建
        if file_id_override and file_id_override in self._metadata:
            file_id = file_id_override
            existing = self._metadata[file_id]
            file_path = Path(existing.file_path)
            date_dir = file_path.parent
        else:
            file_id = str(uuid.uuid4())
            date_dir = self._get_date_dir()
            ext = Path(original_name).suffix or '.xlsx'
            file_path = date_dir / f'{file_id}{ext}'
            early_meta = FileMetadata(
                file_id=file_id,
                original_name=original_name,
                file_path=str(file_path),
                file_size=file_size,
                source_file_id=source_file_id,
                status=FileStatus.READY,
                sheet_names=[],
                sheet_row_counts={},
                row_count=0,
                col_count=0,
                duckdb_ready=False,
                duckdb_load_stage="正在写入文件...",
                duckdb_load_progress=5,
            )
            self._metadata[file_id] = early_meta

        logger.info(f'开始保存文件: original_name={original_name}, size={file_size_mb:.2f} MB')
        self.update_duckdb_load_stage(file_id, "正在写入文件...", 5)

        loop = asyncio.get_running_loop()

        # ---- 文件写入（线程池，不阻塞事件循环） ----
        def _sync_write():
            with open(file_path, 'wb') as f:
                f.write(file_content)

        try:
            await loop.run_in_executor(None, _sync_write)
            logger.info(f'文件写入完成: {file_path}, 大小={file_size_mb:.2f} MB')
        except Exception as e:
            logger.error(f'文件写入失败: {file_path}, 错误: {str(e)}')
            raise

        # ---- 解析文件结构（openpyxl，CPU 密集 → 线程池） ----
        self.update_duckdb_load_stage(file_id, "正在解析文件结构...", 10)
        try:
            sheet_names, sheet_row_counts, row_count, col_count = await loop.run_in_executor(
                None, self._get_file_info_sync, file_path
            )
            logger.info(f'文件信息获取成功: sheets={sheet_names}, rows={row_count}, cols={col_count}')
        except Exception as e:
            logger.error(f'获取文件信息失败: {file_path}, 错误: {str(e)}')
            raise
        
        # 更新元数据（补全 sheet 信息）
        metadata = self._metadata[file_id]
        metadata.sheet_names = sheet_names
        metadata.sheet_row_counts = sheet_row_counts
        metadata.row_count = row_count
        metadata.col_count = col_count
        self.update_duckdb_load_stage(file_id, "正在准备数据...", 25)
        logger.info(f'文件保存完成: file_id={file_id}, name={original_name}, size={file_size_mb:.2f} MB, rows={row_count}, cols={col_count}')
        
        # 记录 DuckDB 加载开始时间
        metadata.duckdb_load_started_at = datetime.now()
        
        # 异步预加载到 DuckDB（非阻塞）
        # 结果文件在占位阶段不预加载，避免将“结果占位”误当作有效分析数据。
        if preload_duckdb:
            asyncio.create_task(self._preload_to_duckdb(file_id, str(file_path)))
        
        return metadata

    def _sync_result_file_map(self, source_file_id: str, result_file_id: str) -> None:
        """同步结果文件映射"""
        if source_file_id:
            self._result_file_map[source_file_id] = result_file_id

    def get_result_file_id(self, source_file_id: str) -> Optional[str]:
        """获取源文件对应的结果文件ID"""
        result_id = self._result_file_map.get(source_file_id)
        if not result_id:
            return None
        meta = self._metadata.get(result_id)
        if not meta or meta.status == FileStatus.DELETED:
            self._result_file_map.pop(source_file_id, None)
            return None
        return result_id

    async def get_or_create_result_file(
        self,
        source_file_id: str,
        filename: Optional[str] = None
    ) -> Optional[FileMetadata]:
        """获取或创建结果文件（单一结果文件策略）"""
        existing_id = self.get_result_file_id(source_file_id)
        if existing_id:
            return self.get_metadata(existing_id)

        source_meta = self.get_metadata(source_file_id)
        if not source_meta:
            logger.warning(f'源文件不存在，无法创建结果文件: {source_file_id}')
            return None

        if not filename:
            base_name = Path(source_meta.original_name).stem
            filename = f'{base_name}_分析结果.xlsx'
        elif not filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            # 确保文件名有扩展名
            filename = f'{filename}.xlsx'

        # 创建占位工作簿（避免空文件）
        wb = Workbook()
        ws = wb.active
        ws.title = "结果占位"
        file_content = BytesIO()
        wb.save(file_content)
        file_content.seek(0)
        wb.close()

        new_meta = await self.save_file(
            file_content.getvalue(),
            filename,
            source_file_id=source_file_id,
            preload_duckdb=False
        )
        self._sync_result_file_map(source_file_id, new_meta.file_id)
        return new_meta

    def _ensure_unique_sheet_name(self, sheet_names: List[str], desired: str) -> str:
        """确保工作表名唯一"""
        if desired not in sheet_names:
            return desired
        index = 1
        while True:
            candidate = f"{desired}_{index:02d}"
            if candidate not in sheet_names:
                return candidate
            index += 1

    async def _update_file_metadata(self, file_id: str, file_path: Path) -> None:
        """更新文件元数据（行列数、工作表列表、大小）"""
        if file_id not in self._metadata:
            return
        loop = asyncio.get_running_loop()
        sheet_names, sheet_row_counts, row_count, col_count = await loop.run_in_executor(
            None, self._get_file_info_sync, file_path
        )
        meta = self._metadata[file_id]
        meta.sheet_names = sheet_names
        meta.sheet_row_counts = sheet_row_counts
        meta.row_count = row_count
        meta.col_count = col_count
        meta.file_size = file_path.stat().st_size
        meta.last_accessed = datetime.now()

    async def append_sheet_to_result_file(
        self,
        result_file_id: str,
        sheet_name: str,
        columns: List[str],
        data: List[List[Any]]
    ) -> Optional[str]:
        """追加结果工作表到结果文件"""
        file_path = self.get_file_path(result_file_id)
        if not file_path or not file_path.exists():
            logger.warning(f'结果文件不存在: {result_file_id}')
            return None

        wb = load_workbook(file_path, read_only=False, data_only=False)
        existing_names = list(wb.sheetnames)
        final_name = self._ensure_unique_sheet_name(existing_names, sheet_name)
        ws = wb.create_sheet(final_name)

        # 写入表头
        for c_idx, col_name in enumerate(columns):
            ws.cell(row=1, column=c_idx + 1, value=col_name)

        avg_col_indexes = {
            idx for idx, col_name in enumerate(columns)
            if isinstance(col_name, str) and '平均' in col_name
        }

        def _round2(num: float) -> float:
            return round(float(num) + 1e-12, 2)

        def _write_numeric_cell(cell, numeric_value: Any) -> None:
            """统一写入数值，避免被工作簿默认日期样式误渲染。"""
            if isinstance(numeric_value, bool):
                cell.value = numeric_value
                return
            if isinstance(numeric_value, Decimal):
                cell.value = _round2(numeric_value)
                cell.number_format = '0.00'
                return
            if isinstance(numeric_value, float):
                cell.value = _round2(numeric_value)
                cell.number_format = '0.00'
                return
            if isinstance(numeric_value, int):
                cell.value = int(numeric_value)
                cell.number_format = '0'
                return
            cell.value = numeric_value

        # 写入数据
        for r_idx, row_data in enumerate(data):
            row_head = str(row_data[0]).strip() if row_data else ''
            is_total_row = row_head in ('总计', '合计', 'Total', 'TOTAL')
            for c_idx, value in enumerate(row_data):
                if hasattr(value, 'item'):
                    value = value.item()
                if is_total_row and c_idx in avg_col_indexes:
                    value = None

                cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
                if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
                    _write_numeric_cell(cell, value)
                else:
                    cell.value = value

        # 如果只有占位工作表，且没有数据，移除占位表
        if len(existing_names) == 1 and existing_names[0] == "结果占位":
            wb.remove(wb[existing_names[0]])

        wb.save(file_path)
        wb.close()
        await self._update_file_metadata(result_file_id, file_path)
        return final_name
    
    # ==========================================================================
    # 操作日志管理
    # ==========================================================================
    
    def add_operation_log(
        self,
        source_file_id: str,
        sheet_name: str,
        operation_type: str,
        logic: str,
        logic_description: str = "",
        row_count: int = 0,
        execution_time_ms: float = 0
    ) -> None:
        """
        添加操作日志记录
        
        Args:
            source_file_id: 源文件ID
            sheet_name: 结果工作表名称
            operation_type: 操作类型（SQL查询、透视表、统计信息等）
            logic: 计算逻辑（SQL语句、透视表配置等）
            logic_description: 逻辑说明（面向业务人员的可读描述）
            row_count: 结果数据行数
            execution_time_ms: 执行耗时（毫秒）
        """
        if source_file_id not in self._operation_log_cache:
            self._operation_log_cache[source_file_id] = []
        
        log_entry = {
            'seq': len(self._operation_log_cache[source_file_id]) + 1,
            'sheet_name': sheet_name,
            'operation_type': operation_type,
            'logic': logic,
            'logic_description': logic_description,
            'created_at': datetime.now(),
            'row_count': row_count,
            'execution_time_ms': round(execution_time_ms, 2)
        }
        self._operation_log_cache[source_file_id].append(log_entry)
        logger.info(f'添加操作日志: source_file_id={source_file_id}, sheet={sheet_name}, type={operation_type}')
    
    def remove_operation_log(self, source_file_id: str, sheet_name: str) -> bool:
        """
        移除操作日志记录（关闭工作表时调用）
        
        Args:
            source_file_id: 源文件ID
            sheet_name: 结果工作表名称
            
        Returns:
            是否成功移除
        """
        if source_file_id not in self._operation_log_cache:
            return False
        
        logs = self._operation_log_cache[source_file_id]
        original_len = len(logs)
        self._operation_log_cache[source_file_id] = [
            log for log in logs if log['sheet_name'] != sheet_name
        ]
        
        # 重新编号
        for i, log in enumerate(self._operation_log_cache[source_file_id]):
            log['seq'] = i + 1
        
        removed = original_len > len(self._operation_log_cache[source_file_id])
        if removed:
            logger.info(f'移除操作日志: source_file_id={source_file_id}, sheet={sheet_name}')
        return removed
    
    def get_operation_logs(self, source_file_id: str) -> List[Dict[str, Any]]:
        """获取源文件的所有操作日志"""
        return self._operation_log_cache.get(source_file_id, [])
    
    async def sync_operation_log_to_file(self, source_file_id: str) -> bool:
        """
        同步操作日志到结果文件
        
        在结果文件中创建/更新 _操作日志 工作表
        
        Args:
            source_file_id: 源文件ID
            
        Returns:
            是否成功同步
        """
        result_file_id = self.get_result_file_id(source_file_id)
        if not result_file_id:
            return False
        
        file_path = self.get_file_path(result_file_id)
        if not file_path or not file_path.exists():
            return False
        
        logs = self._operation_log_cache.get(source_file_id, [])
        
        try:
            wb = load_workbook(file_path, read_only=False, data_only=False)
            
            # 如果操作日志表已存在，删除后重建
            if self.OPERATION_LOG_SHEET_NAME in wb.sheetnames:
                wb.remove(wb[self.OPERATION_LOG_SHEET_NAME])
            
            # 如果没有日志，不创建空表
            if not logs:
                wb.save(file_path)
                wb.close()
                return True
            
            # 创建操作日志工作表
            ws = wb.create_sheet(self.OPERATION_LOG_SHEET_NAME, 0)  # 放在第一个位置
            
            # 写入表头
            headers = ['序号', '结果工作表', '操作类型', '逻辑说明', '计算逻辑', '生成时间', '数据行数', '耗时(ms)']
            for c_idx, header in enumerate(headers):
                cell = ws.cell(row=1, column=c_idx + 1, value=header)
                # 设置表头样式
                cell.font = cell.font.copy(bold=True)
            
            # 写入数据
            for r_idx, log in enumerate(logs):
                ws.cell(row=r_idx + 2, column=1, value=log['seq'])
                ws.cell(row=r_idx + 2, column=2, value=log['sheet_name'])
                ws.cell(row=r_idx + 2, column=3, value=log['operation_type'])
                ws.cell(row=r_idx + 2, column=4, value=log.get('logic_description') or log['operation_type'])
                # 逻辑可能很长，截断显示
                logic_display = log['logic'][:500] + '...' if len(log['logic']) > 500 else log['logic']
                ws.cell(row=r_idx + 2, column=5, value=logic_display)
                ws.cell(row=r_idx + 2, column=6, value=log['created_at'].strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=r_idx + 2, column=7, value=log['row_count'])
                ws.cell(row=r_idx + 2, column=8, value=log['execution_time_ms'])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8   # 序号
            ws.column_dimensions['B'].width = 20  # 结果工作表
            ws.column_dimensions['C'].width = 12  # 操作类型
            ws.column_dimensions['D'].width = 24  # 逻辑说明
            ws.column_dimensions['E'].width = 80  # 计算逻辑
            ws.column_dimensions['F'].width = 20  # 生成时间
            ws.column_dimensions['G'].width = 12  # 数据行数
            ws.column_dimensions['H'].width = 12  # 耗时
            
            wb.save(file_path)
            wb.close()
            
            # 更新元数据（包含操作日志工作表）
            await self._update_file_metadata(result_file_id, file_path)
            
            logger.info(f'操作日志已同步到文件: source_file_id={source_file_id}, logs_count={len(logs)}')
            return True
            
        except Exception as e:
            logger.error(f'同步操作日志失败: {e}')
            return False
    
    def clear_operation_logs(self, source_file_id: str) -> int:
        """
        清空源文件的所有操作日志
        
        Args:
            source_file_id: 源文件ID
            
        Returns:
            清除的日志数量
        """
        if source_file_id not in self._operation_log_cache:
            return 0
        count = len(self._operation_log_cache[source_file_id])
        del self._operation_log_cache[source_file_id]
        logger.info(f'清空操作日志: source_file_id={source_file_id}, count={count}')
        return count

    async def remove_sheet_from_result_file(
        self,
        source_file_id: str,
        sheet_name: str
    ) -> Dict[str, Any]:
        """从结果文件中移除指定工作表（同步更新操作日志）"""
        result_file_id = self.get_result_file_id(source_file_id)
        if not result_file_id:
            return {"success": False, "message": "结果文件不存在"}

        file_path = self.get_file_path(result_file_id)
        if not file_path or not file_path.exists():
            return {"success": False, "message": "结果文件不存在或已删除"}

        wb = load_workbook(file_path, read_only=False, data_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "message": "工作表不存在"}

        # 计算有效工作表数量（排除操作日志工作表）
        data_sheets = [s for s in wb.sheetnames if s != self.OPERATION_LOG_SHEET_NAME]
        
        # 如果删除最后一个数据工作表，直接删除结果文件
        if len(data_sheets) == 1 and sheet_name in data_sheets:
            wb.close()
            await self.delete_file(result_file_id)
            for key, value in list(self._result_file_map.items()):
                if value == result_file_id:
                    self._result_file_map.pop(key, None)
            # 清空操作日志
            self.clear_operation_logs(source_file_id)
            return {"success": True, "result_deleted": True, "sheet_names": []}

        wb.remove(wb[sheet_name])
        wb.save(file_path)
        wb.close()
        
        # 同步删除操作日志记录
        self.remove_operation_log(source_file_id, sheet_name)
        # 同步操作日志到文件
        await self.sync_operation_log_to_file(source_file_id)
        
        await self._update_file_metadata(result_file_id, file_path)
        meta = self._metadata.get(result_file_id)
        # 返回数据工作表列表（排除操作日志）
        remaining_sheets = [s for s in (meta.sheet_names if meta else []) if s != self.OPERATION_LOG_SHEET_NAME]
        return {
            "success": True,
            "result_deleted": False,
            "sheet_names": remaining_sheets
        }
    
    def update_duckdb_load_stage(self, file_id: str, stage: str, progress: int = None):
        """更新 DuckDB 加载阶段"""
        if file_id in self._metadata:
            self._metadata[file_id].duckdb_load_stage = stage
            if progress is not None:
                self._metadata[file_id].duckdb_load_progress = progress
            logger.info(f'DuckDB 加载阶段更新: file_id={file_id}, stage={stage}, progress={progress}%')
    
    async def _preload_to_duckdb(self, file_id: str, file_path: str):
        """预加载文件到 DuckDB 缓存"""
        try:
            import time
            duckdb_mgr = _get_duckdb_manager()
            meta = self._metadata.get(file_id)
            file_size_mb = meta.file_size / 1024 / 1024 if meta else 0
            row_count = meta.row_count if meta else 0
            
            logger.info(f'开始预加载到 DuckDB: file_id={file_id}, size={file_size_mb:.1f}MB, rows={row_count}')
            
            # 更新阶段：开始加载（用户友好文案，不暴露技术栈）
            self.update_duckdb_load_stage(file_id, f"正在读取文件 ({file_size_mb:.1f} MB)...", 10)
            
            # 在线程池中执行，避免阻塞事件循环
            loop = asyncio.get_event_loop()
            
            # 用于跟踪加载状态
            load_start_time = time.time()
            load_completed = False
            
            async def update_progress_periodically():
                """定期更新进度（每 3 秒更新一次经过的时间）"""
                while not load_completed:
                    elapsed = time.time() - load_start_time
                    # 根据经过时间估算进度（假设大文件需要 2-3 分钟）
                    # 进度从 30% 开始，最多到 90%
                    estimated_progress = min(30 + int(elapsed / 2), 90)  # 每 2 秒增加 1%
                    self.update_duckdb_load_stage(
                        file_id, 
                        f"正在解析数据 ({row_count:,} 行)，已耗时 {int(elapsed)} 秒...", 
                        estimated_progress
                    )
                    await asyncio.sleep(3)  # 每 3 秒更新一次
            
            def do_load():
                """执行实际的加载操作"""
                nonlocal load_completed
                try:
                    duckdb_mgr.load_excel(file_path, file_id)
                finally:
                    load_completed = True
            
            # 并行执行：1) 进度更新任务 2) 实际加载任务
            progress_task = asyncio.create_task(update_progress_periodically())
            
            try:
                # 在线程池中执行加载
                await loop.run_in_executor(None, do_load)
            finally:
                # 确保进度更新任务停止
                load_completed = True
                progress_task.cancel()
                try:
                    await progress_task
                except asyncio.CancelledError:
                    pass
            
            # 加载完成
            total_elapsed = time.time() - load_start_time
            self.update_duckdb_load_stage(file_id, f"已完成 (耗时 {int(total_elapsed)} 秒)", 100)
            self.set_duckdb_ready(file_id, True)
            logger.info(f'DuckDB 预加载完成: file_id={file_id}, 耗时={total_elapsed:.1f}秒')
        except Exception as e:
            # 预加载失败不影响正常使用，工具会在需要时重新加载
            self.update_duckdb_load_stage(file_id, f"处理出错: {str(e)[:50]}", None)
            logger.warning(f'DuckDB 预加载失败（将在使用时重新加载）: file_id={file_id}, error={e}')
    
    def set_first_page_ready(self, file_id: str):
        """标记第一张工作表已加载，前端可以提前渲染"""
        if file_id in self._metadata:
            self._metadata[file_id].first_page_ready = True
            logger.info(f'首页就绪: file_id={file_id}')

    def set_duckdb_ready(self, file_id: str, ready: bool = True):
        """设置全部工作表加载完成"""
        if file_id in self._metadata:
            self._metadata[file_id].duckdb_ready = ready
            self._metadata[file_id].first_page_ready = True
            if ready:
                self._metadata[file_id].duckdb_load_finished_at = datetime.now()
    
    def _get_file_info_sync(self, file_path: Path) -> Tuple[List[str], Dict[str, int], int, int]:
        """
        获取 Excel 文件基本信息（同步，适合在线程池中调用）
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = [s for s in (wb.sheetnames or []) if s and s != META_SHEET_NAME]

            sheet_row_counts = {}
            total_row_count = 0
            max_col_count = 0

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                rows = ws.max_row or 0
                cols = ws.max_column or 0
                sheet_row_counts[sheet_name] = rows
                total_row_count += rows
                max_col_count = max(max_col_count, cols)

            wb.close()
            return sheet_names, sheet_row_counts, total_row_count, max_col_count
        except Exception as e:
            logger.error(f'获取文件信息失败: {file_path}, 错误: {str(e)}')
            return [], {}, 0, 0
    
    def get_metadata(self, file_id: str) -> Optional[FileMetadata]:
        """获取文件元数据"""
        meta = self._metadata.get(file_id)
        if meta:
            meta.last_accessed = datetime.now()
        return meta
    
    def get_file_path(self, file_id: str) -> Optional[Path]:
        """获取文件路径"""
        meta = self.get_metadata(file_id)
        if meta and meta.status != FileStatus.DELETED:
            return Path(meta.file_path)
        return None
    
    def _detect_title_row(self, file_path: Path, sheet_name: str) -> int:
        """
        检测标题行（与 DuckDB 管理器中的逻辑一致）
        
        Returns:
            标题行数量（0 或 1）
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # 读取第一行
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            # 检查第一行：如果只有第一列有内容，其他列大部分为空，则认为是标题行
            first_col_value = first_row[0] if first_row else None
            other_cols_non_empty = sum(1 for val in first_row[1:] if val is not None and str(val).strip())
            
            # 标题行判断条件：
            # 1. 第一列有内容
            # 2. 其他列非空数量 <= 2（允许少量合并单元格）
            is_title_row = False
            if first_col_value and isinstance(first_col_value, str):
                first_col_str = str(first_col_value).strip()
                # 检查是否包含标题关键词
                title_keywords = ['表', '数据', '明细', '统计', '汇总', '分析', '报表']
                has_title_keyword = any(keyword in first_col_str for keyword in title_keywords)
                
                # 如果第一列有内容，且其他列非空数量 <= 2，认为是标题行
                if other_cols_non_empty <= 2:
                    is_title_row = True
                    logger.info(f'[预览] 检测到标题行: A1="{first_col_str}", 其他列非空数={other_cols_non_empty}')
            
            wb.close()
            return 1 if is_title_row else 0
        except Exception as e:
            logger.warning(f'[预览] 检测标题行失败: {e}，默认不跳过')
            return 0
    
    async def get_preview(
        self, 
        file_id: str, 
        sheet_name: Optional[str] = None,
        max_rows: int = PREVIEW_ROW_COUNT,
        include_styles: bool = True
    ) -> Optional[Dict[str, Any]]:
        """
        获取文件预览数据（TOP N 行）
        
        Args:
            file_id: 文件ID
            sheet_name: 工作表名称，None 则使用第一个
            max_rows: 最大行数
            include_styles: 是否包含样式信息（大文件时可关闭以提升性能）
            
        Returns:
            预览数据字典
        """
        import time as _time
        start_time = _time.time()
        
        logger.debug(f'获取预览: file_id={file_id}, sheet_name={sheet_name}, max_rows={max_rows}, include_styles={include_styles}')
        file_path = self.get_file_path(file_id)
        if not file_path or not file_path.exists():
            logger.warning(f'文件不存在: file_id={file_id}, path={file_path}')
            return None
        
        try:
            # 检测标题行
            if sheet_name is None:
                wb_temp = load_workbook(file_path, read_only=True)
                sheet_name = wb_temp.active.title
                wb_temp.close()
            
            title_row_count = self._detect_title_row(file_path, sheet_name)
            start_row = 1 + title_row_count  # 如果有标题行，从第二行开始读取
            
            # 第一次读取：获取计算值（data_only=True, read_only=True - 快速模式）
            t1 = _time.time()
            logger.debug(f'打开工作簿获取计算值: {file_path}')
            wb_data = load_workbook(file_path, read_only=True, data_only=True)
            
            if sheet_name and sheet_name in wb_data.sheetnames:
                ws_data = wb_data[sheet_name]
            else:
                ws_data = wb_data.active
                sheet_name = ws_data.title
            
            # 读取计算值（跳过标题行）
            data_values = []
            row_idx = 0
            for row in ws_data.iter_rows(max_row=max_rows + start_row):
                row_idx += 1
                if row_idx < start_row:
                    continue  # 跳过标题行
                data_values.append([cell.value for cell in row])
            
            # 避免访问 max_row/max_column 触发全量扫描，优先使用元数据
            # 对于大文件，访问这些属性可能需要扫描整个工作表（5-10秒）
            meta = self._metadata.get(file_id)
            if meta and sheet_name in meta.sheet_row_counts:
                # 使用已缓存的元数据（上传时已计算，无需扫描）
                total_rows = meta.sheet_row_counts[sheet_name]
                total_cols = meta.col_count
                logger.debug(f'使用元数据获取行列数: rows={total_rows}, cols={total_cols}')
            else:
                # 如果元数据不可用，使用读取到的数据估算（避免全量扫描）
                # 注意：这只是估算值，实际行数可能更多
                total_rows = len(data_values) + max_rows if data_values else 0
                total_cols = len(data_values[0]) if data_values else 0
                logger.debug(f'使用估算值（避免全量扫描）: rows={total_rows}, cols={total_cols}')
            
            wb_data.close()
            logger.debug(f'计算值读取完成: 耗时={_time.time() - t1:.2f}秒')
            
            # 初始化数据结构
            headers = []
            data = []
            styles = []
            
            if include_styles:
                # 第二次读取：获取公式和样式（read_only=False, data_only=False）
                # 注意：read_only=False 才能读取样式，但会加载整个文件
                t2 = _time.time()
                logger.debug(f'打开工作簿获取公式和样式: {file_path}')
                wb = load_workbook(file_path, read_only=False, data_only=False)
                logger.debug(f'工作簿打开完成（含样式）: 耗时={_time.time() - t2:.2f}秒')
                
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                
                # 读取数据和样式（跳过标题行）
                row_counter = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows + start_row), 1):
                    if row_idx < start_row:
                        continue  # 跳过标题行
                    
                    row_values = []
                    row_styles = []
                    
                    for col_idx, cell in enumerate(row):
                        # 获取值：优先使用计算值
                        formula_val = cell.value
                        data_val = data_values[row_idx][col_idx] if row_idx < len(data_values) and col_idx < len(data_values[row_idx]) else None
                        
                        is_formula = isinstance(formula_val, str) and formula_val.startswith('=')
                        if is_formula:
                            if data_val is not None:
                                row_values.append(_serialize_cell_value(data_val))
                            else:
                                row_values.append(formula_val)
                        else:
                            final_val = data_val if data_val is not None else formula_val
                            row_values.append(_serialize_cell_value(final_val))
                        
                        # 获取样式
                        cell_style = {}
                        try:
                            if cell.font:
                                if cell.font.bold:
                                    cell_style['bold'] = True
                                if cell.font.italic:
                                    cell_style['italic'] = True
                                if cell.font.color and cell.font.color.rgb:
                                    rgb = cell.font.color.rgb
                                    if isinstance(rgb, str) and rgb != '00000000':
                                        if len(rgb) == 8:
                                            cell_style['fontColor'] = f'#{rgb[2:]}'
                                        elif len(rgb) == 6:
                                            cell_style['fontColor'] = f'#{rgb}'
                                if cell.font.size:
                                    cell_style['fontSize'] = cell.font.size
                            
                            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                                rgb = cell.fill.fgColor.rgb
                                if isinstance(rgb, str) and rgb != '00000000':
                                    if len(rgb) == 8:
                                        cell_style['backgroundColor'] = f'#{rgb[2:]}'
                                    elif len(rgb) == 6:
                                        cell_style['backgroundColor'] = f'#{rgb}'
                            
                            if cell.alignment:
                                if cell.alignment.horizontal:
                                    cell_style['horizontalAlign'] = cell.alignment.horizontal
                                if cell.alignment.vertical:
                                    cell_style['verticalAlign'] = cell.alignment.vertical
                        except Exception as style_err:
                            logger.debug(f'读取单元格样式失败: {style_err}')
                        
                        row_styles.append(cell_style if cell_style else None)
                    
                    # 第一行（跳过标题行后的第一行）作为表头
                    if row_counter == 0:
                        headers = [str(v) if v is not None else f'列{i+1}' for i, v in enumerate(row_values)]
                        styles.append(row_styles)
                    else:
                        data.append(row_values)
                        styles.append(row_styles)
                    
                    row_counter += 1
                
                wb.close()
            else:
                # 快速模式：不读取样式，直接使用 data_values
                # data_values 已经跳过了标题行
                for row_idx, row_data in enumerate(data_values):
                    row_values = [_serialize_cell_value(v) for v in row_data]
                    
                    # 第一行作为表头
                    if row_idx == 0:
                        headers = [str(v) if v is not None else f'列{i+1}' for i, v in enumerate(row_values)]
                        styles.append([None] * len(row_values))  # 无样式
                    else:
                        data.append(row_values)
                        styles.append([None] * len(row_values))  # 无样式
            
            _has_styles = any(any(s for s in row if s) for row in styles) if styles else False
            total_time = _time.time() - start_time
            
            preview_data = {
                'file_id': file_id,
                'sheet_name': sheet_name,
                'sheet_names': self._metadata[file_id].sheet_names,
                'headers': headers,
                'data': data,
                'styles': styles,
                'total_rows': total_rows,
                'total_cols': total_cols,
                'preview_rows': len(data),
            }
            
            logger.info(f'预览生成成功: file_id={file_id}, sheet={sheet_name}, 预览行数={len(data)}, 总行数={total_rows}, 包含样式={_has_styles}, 总耗时={total_time:.2f}秒')
            return preview_data
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f'获取预览失败: file_id={file_id}, 错误: {str(e)}')
            logger.debug(f'错误详情:\n{error_details}')
            return None
    
    async def delete_file(self, file_id: str) -> bool:
        """删除文件"""
        meta = self._metadata.get(file_id)
        if not meta:
            return False
        
        try:
            # 从 DuckDB 缓存中卸载
            try:
                duckdb_mgr = _get_duckdb_manager()
                duckdb_mgr.unload_file(file_id)
                logger.debug(f'DuckDB 缓存已清理: file_id={file_id}')
            except Exception as e:
                logger.warning(f'清理 DuckDB 缓存失败: file_id={file_id}, error={e}')
            
            file_path = Path(meta.file_path)
            if file_path.exists():
                file_path.unlink()
            
            meta.status = FileStatus.DELETED
            del self._metadata[file_id]
            # 清理结果文件映射
            for key, value in list(self._result_file_map.items()):
                if value == file_id:
                    self._result_file_map.pop(key, None)
            
            logger.info(f'File deleted: {file_id}')
            return True
        except Exception as e:
            logger.error(f'Failed to delete file {file_id}: {e}')
            return False
    
    def update_status(self, file_id: str, status: FileStatus, error: Optional[str] = None):
        """更新文件状态"""
        meta = self._metadata.get(file_id)
        if meta:
            meta.status = status
            if error:
                meta.error_message = error
            logger.info(f'File {file_id} status updated to {status}')
    
    def list_files(self) -> List[FileMetadata]:
        """列出所有文件"""
        return [
            meta for meta in self._metadata.values()
            if meta.status != FileStatus.DELETED
        ]
    
    async def get_preview_fast(
        self, 
        file_id: str, 
        sheet_name: Optional[str] = None,
        max_rows: int = PREVIEW_ROW_COUNT,
        offset: int = 0
    ) -> Optional[Dict[str, Any]]:
        """
        使用 DuckDB 快速获取预览数据（不含样式）
        
        比 openpyxl 快 10 倍以上，适合大文件
        """
        import time as _time
        start_time = _time.time()
        
        file_path = self.get_file_path(file_id)
        if not file_path or not file_path.exists():
            logger.warning(f'文件不存在: file_id={file_id}')
            return None
        
        try:
            duckdb_mgr = _get_duckdb_manager()
            
            # 确保文件已加载
            if not duckdb_mgr.is_loaded(file_id, sheet_name):
                logger.info(f'[DuckDB快速预览] 文件未在 DuckDB 缓存中，正在加载: file_id={file_id}, sheet={sheet_name}')
                duckdb_mgr.load_excel(str(file_path), file_id, sheet_name)
                logger.info(f'[DuckDB快速预览] 文件加载完成: file_id={file_id}, sheet={sheet_name}')
            else:
                logger.debug(f'[DuckDB快速预览] 文件已在 DuckDB 缓存中: file_id={file_id}, sheet={sheet_name}')
            
            # 获取预览
            logger.debug(f'[DuckDB快速预览] 开始获取预览: file_id={file_id}, sheet={sheet_name}, max_rows={max_rows}, offset={offset}')
            preview = duckdb_mgr.get_preview(file_id, sheet_name, max_rows, offset=offset)
            logger.debug(f'[DuckDB快速预览] 预览获取成功: rows={preview.get("preview_rows", 0)}, total_rows={preview.get("row_count", 0)}')
            
            # 转换为标准格式
            headers = preview['columns']
            data = preview['data']
            
            # 生成空样式数组
            styles = [[None] * len(headers)] * (len(data) + 1)
            
            total_time = _time.time() - start_time
            
            meta = self._metadata.get(file_id)
            sheet_names = meta.sheet_names if meta else []
            
            preview_data = {
                'file_id': file_id,
                'sheet_name': sheet_name or 'default',
                'sheet_names': sheet_names,
                'headers': headers,
                'data': data,
                'styles': styles,
                'total_rows': preview['row_count'],
                'total_cols': preview['col_count'],
                'preview_rows': preview['preview_rows'],
                'offset': preview.get('offset', max(0, int(offset))),
                'limit': preview.get('limit', max_rows),
                'has_more': preview.get('has_more', False),
            }
            
            logger.info(f'DuckDB 快速预览完成: file_id={file_id}, 预览行数={len(data)}, 总行数={preview["row_count"]}, 耗时={total_time:.2f}秒')
            return preview_data
            
        except Exception as e:
            import traceback
            logger.error(f'DuckDB 快速预览失败，回退到 openpyxl: file_id={file_id}, error={e}')
            logger.debug(f'错误详情:\n{traceback.format_exc()}')
            # 回退到传统方法
            return await self.get_preview(file_id, sheet_name, max_rows, include_styles=False)
    
    def is_duckdb_loaded(self, file_id: str, sheet_name: Optional[str] = None) -> bool:
        """检查文件是否已加载到 DuckDB"""
        try:
            duckdb_mgr = _get_duckdb_manager()
            return duckdb_mgr.is_loaded(file_id, sheet_name)
        except Exception:
            return False
    
    def get_result_files(self, source_file_id: str) -> List[Dict[str, Any]]:
        """
        获取源文件关联的结果文件列表
        
        结果文件是通过导出工具生成的新文件，source_file_id 指向源文件
        
        Args:
            source_file_id: 源文件ID
            
        Returns:
            结果文件列表，每个元素包含 {file_id, filename, row_count, col_count, created_at}
        """
        try:
            # 优先使用单一结果文件映射
            result_file_id = self.get_result_file_id(source_file_id)
            if result_file_id:
                meta = self.get_metadata(result_file_id)
                if meta:
                    return [{
                        'file_id': meta.file_id,
                        'filename': meta.original_name,
                        'row_count': meta.row_count,
                        'col_count': meta.col_count,
                        'created_at': meta.created_at.isoformat(),
                        'sheet_names': meta.sheet_names
                    }]

            # 兼容旧逻辑（无映射时回退）
            source_meta = self.get_metadata(source_file_id)
            if not source_meta:
                logger.warning(f'源文件不存在: {source_file_id}')
                return []

            result_metas = [
                meta for meta in self._metadata.values()
                if meta.status != FileStatus.DELETED and meta.source_file_id == source_file_id
            ]
            result_metas.sort(key=lambda m: m.created_at, reverse=True)

            result_files = []
            for meta in result_metas:
                result_files.append({
                    'file_id': meta.file_id,
                    'filename': meta.original_name,
                    'row_count': meta.row_count,
                    'col_count': meta.col_count,
                    'created_at': meta.created_at.isoformat(),
                    'sheet_names': meta.sheet_names
                })

            logger.debug(f'获取结果文件列表: source_file_id={source_file_id}, count={len(result_files)}')
            return result_files
        except Exception as e:
            logger.error(f'get_result_files 失败: {e}')
            return []
    
    # ==========================================================================
    # 内存结果管理（会话期间有效）
    # ==========================================================================
    
    def add_memory_result(
        self,
        source_file_id: str,
        sheet_name: str,
        table_name: str,
        columns: List[str],
        row_count: int
    ) -> None:
        """
        添加内存结果（不写文件）
        
        Args:
            source_file_id: 源文件ID
            sheet_name: 结果工作表名称
            table_name: DuckDB 内存表名
            columns: 列名列表
            row_count: 行数
        """
        if source_file_id not in self._memory_result_cache:
            self._memory_result_cache[source_file_id] = {}
        
        self._memory_result_cache[source_file_id][sheet_name] = {
            'table_name': table_name,
            'columns': columns,
            'row_count': row_count,
            'created_at': datetime.now()
        }
        logger.info(f'添加内存结果: source_file_id={source_file_id}, sheet={sheet_name}, table={table_name}, rows={row_count}')
    
    def get_memory_result(self, source_file_id: str, sheet_name: str) -> Optional[Dict[str, Any]]:
        """获取内存结果"""
        return self._memory_result_cache.get(source_file_id, {}).get(sheet_name)
    
    def list_memory_results(self, source_file_id: str) -> List[Dict[str, Any]]:
        """列出源文件的所有内存结果"""
        results = []
        for sheet_name, info in self._memory_result_cache.get(source_file_id, {}).items():
            results.append({
                'sheet_name': sheet_name,
                'columns': info['columns'],
                'row_count': info['row_count'],
                'created_at': info['created_at'].isoformat()
            })
        return results
    
    def remove_memory_result(self, source_file_id: str, sheet_name: str) -> Dict[str, Any]:
        """
        移除内存结果（关闭工作表时调用）
        
        同时释放 DuckDB 内存表
        
        Returns:
            {success: bool, message: str, remaining_sheets: List[str]}
        """
        if source_file_id not in self._memory_result_cache:
            return {
                'success': False,
                'message': f'源文件不存在内存结果: {source_file_id}',
                'remaining_sheets': []
            }
        
        if sheet_name not in self._memory_result_cache[source_file_id]:
            return {
                'success': False,
                'message': f'工作表不存在: {sheet_name}',
                'remaining_sheets': list(self._memory_result_cache[source_file_id].keys())
            }
        
        info = self._memory_result_cache[source_file_id][sheet_name]
        table_name = info['table_name']
        
        # 使用 DuckDB 管理器释放内存表
        try:
            duckdb_mgr = _get_duckdb_manager()
            # 先尝试使用新的 unload_sheet 方法（会同步清理缓存）
            duckdb_mgr.unload_sheet(f"result_{source_file_id}", sheet_name)
            # 如果 unload_sheet 没有删除（可能表名不匹配），直接删除
            duckdb_mgr.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            logger.info(f'已释放 DuckDB 内存表: {table_name}')
        except Exception as e:
            logger.warning(f'释放 DuckDB 内存表失败: {table_name}, error={e}')
        
        # 从缓存中移除
        del self._memory_result_cache[source_file_id][sheet_name]
        
        # 获取剩余工作表
        remaining_sheets = list(self._memory_result_cache.get(source_file_id, {}).keys())
        
        # 如果该源文件没有其他结果，清理整个条目
        if not self._memory_result_cache.get(source_file_id):
            if source_file_id in self._memory_result_cache:
                del self._memory_result_cache[source_file_id]
        
        logger.info(f'移除内存结果: source_file_id={source_file_id}, sheet={sheet_name}, remaining={remaining_sheets}')
        return {
            'success': True,
            'message': f'已释放工作表 "{sheet_name}" 的内存',
            'remaining_sheets': remaining_sheets
        }
    
    def clear_session_memory(self, source_file_id: str) -> Dict[str, Any]:
        """
        清空会话的所有内存结果（导出文件后调用）
        
        同时释放所有 DuckDB 内存表
        
        Args:
            source_file_id: 源文件ID
            
        Returns:
            {success: bool, cleared_sheets: int, cleared_tables: int}
        """
        cleared_sheets = 0
        cleared_tables = 0
        
        # 清理内存结果缓存
        if source_file_id in self._memory_result_cache:
            results = self._memory_result_cache[source_file_id]
            duckdb_mgr = _get_duckdb_manager()
            
            for sheet_name, info in results.items():
                table_name = info.get('table_name')
                if table_name:
                    try:
                        duckdb_mgr.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                        cleared_tables += 1
                        logger.debug(f'会话清理 - 删除内存表: {table_name}')
                    except Exception as e:
                        logger.warning(f'会话清理 - 删除内存表失败: {table_name}, error={e}')
                cleared_sheets += 1
            
            del self._memory_result_cache[source_file_id]
        
        # 清理 DuckDB 中的源数据表和结果表
        try:
            duckdb_mgr = _get_duckdb_manager()
            result = duckdb_mgr.clear_session_cache([source_file_id, f"result_{source_file_id}"])
            cleared_tables += result.get('cleared_tables', 0)
        except Exception as e:
            logger.warning(f'清理 DuckDB 会话缓存失败: {e}')
        
        # 清理操作日志
        cleared_logs = self.clear_operation_logs(source_file_id)
        
        logger.info(f'会话内存已清理: source_file_id={source_file_id}, sheets={cleared_sheets}, tables={cleared_tables}, logs={cleared_logs}')
        return {
            'success': True,
            'cleared_sheets': cleared_sheets,
            'cleared_tables': cleared_tables,
            'cleared_logs': cleared_logs
        }

    async def purge_by_source_file_id(self, source_file_id: str) -> Dict[str, Any]:
        """
        按源文件ID（文件管理 UserFile.id）彻底清理关联的大文件会话与结果文件。

        说明：
        - source_file_id 为文件管理中的文件ID
        - 先找到所有 metadata.source_file_id == source_file_id 的大文件会话 file_id
        - 逐个清理会话内存、删除结果文件、删除会话文件本身
        """
        related_file_ids = [
            meta.file_id
            for meta in self._metadata.values()
            if meta.status != FileStatus.DELETED and meta.source_file_id == source_file_id
        ]

        cleared_sessions = 0
        deleted_result_files = 0
        deleted_source_sessions = 0

        for file_id in related_file_ids:
            try:
                self.clear_session_memory(file_id)
                cleared_sessions += 1
            except Exception as e:
                logger.warning(f'清理会话内存失败: file_id={file_id}, error={e}')

            # 删除关联结果文件
            result_file_ids = []
            direct_result_id = self.get_result_file_id(file_id)
            if direct_result_id:
                result_file_ids.append(direct_result_id)
            # 兜底：扫描 metadata，删除所有 source_file_id 指向该会话的结果文件
            result_file_ids.extend([
                meta.file_id
                for meta in self._metadata.values()
                if meta.status != FileStatus.DELETED and meta.source_file_id == file_id
            ])
            for rid in list(dict.fromkeys(result_file_ids)):
                try:
                    deleted = await self.delete_file(rid)
                    if deleted:
                        deleted_result_files += 1
                except Exception as e:
                    logger.warning(f'删除结果文件失败: result_file_id={rid}, error={e}')

            try:
                deleted = await self.delete_file(file_id)
                if deleted:
                    deleted_source_sessions += 1
            except Exception as e:
                logger.warning(f'删除会话文件失败: file_id={file_id}, error={e}')

            self._result_file_map.pop(file_id, None)

        return {
            'source_file_id': source_file_id,
            'related_session_count': len(related_file_ids),
            'cleared_sessions': cleared_sessions,
            'deleted_result_files': deleted_result_files,
            'deleted_source_sessions': deleted_source_sessions,
        }
    
    async def save_memory_results_to_file(
        self,
        source_file_id: str,
        filename: Optional[str] = None,
        progress_callback: Optional[Callable[[str, str, Optional[float]], None]] = None
    ) -> Optional[FileMetadata]:
        """
        保存内存结果到文件（下载时触发）
        
        Args:
            source_file_id: 源文件ID
            filename: 文件名（可选）
            progress_callback: 进度回调函数 callback(stage, message, progress)
            
        Returns:
            文件元数据
        """
        if source_file_id not in self._memory_result_cache:
            logger.warning(f'没有内存结果需要保存: source_file_id={source_file_id}')
            return None
        
        results = self._memory_result_cache[source_file_id]
        if not results:
            logger.warning(f'内存结果为空: source_file_id={source_file_id}')
            return None
        
        duckdb_mgr = _get_duckdb_manager()
        total_sheets = len(results)
        
        def _notify_progress(stage: str, message: str, progress: float = None):
            """通知进度"""
            if progress_callback:
                try:
                    progress_callback(stage, message, progress)
                except Exception:
                    pass
            logger.info(f'[保存进度] {stage}: {message}')
        
        # 阶段1：准备文件
        _notify_progress('prepare', f'正在准备文件...', 0.0)
        if not filename:
            source_meta = self.get_metadata(source_file_id)
            base_name = Path(source_meta.original_name).stem if source_meta else "分析结果"
            filename = f'{base_name}_分析结果.xlsx'
        elif not filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            # 确保文件名有扩展名
            filename = f'{filename}.xlsx'
        
        # 创建临时工作簿
        wb = Workbook()
        ws_placeholder = wb.active
        ws_placeholder.title = "结果占位"
        file_content = BytesIO()
        wb.save(file_content)
        file_content.seek(0)
        wb.close()
        
        # 保存临时文件
        result_meta = await self.save_file(
            file_content.getvalue(),
            filename,
            source_file_id=source_file_id,
            preload_duckdb=False
        )
        result_file_path = self.get_file_path(result_meta.file_id)
        _notify_progress('prepare', f'文件准备完成', 0.1)
        
        # 阶段2：批量写入数据
        try:
            import pandas as pd

            def _round_numeric_dataframe(df):
                """统一数值精度，避免导出后出现长尾浮点。"""
                if df is None or df.empty:
                    return df
                rounded_df = df.copy()
                for col in rounded_df.columns:
                    series = rounded_df[col]
                    if pd.api.types.is_float_dtype(series):
                        rounded_df[col] = series.round(2)
                return rounded_df

            def _apply_numeric_number_formats(ws, df, data_start_row: int = 2):
                """统一设置数值格式，避免被误显示为日期/时间。"""
                if df is None or df.empty:
                    return
                for c_idx, col in enumerate(df.columns, start=1):
                    series = df[col]
                    is_int_col = pd.api.types.is_integer_dtype(series)
                    is_float_col = pd.api.types.is_float_dtype(series)
                    if not (is_int_col or is_float_col):
                        continue
                    for r_idx in range(data_start_row, data_start_row + len(df)):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        if isinstance(cell.value, bool):
                            continue
                        if isinstance(cell.value, int):
                            cell.number_format = '0'
                        elif isinstance(cell.value, float):
                            cell.value = round(float(cell.value) + 1e-12, 2)
                            cell.number_format = '0.00'
            
            _notify_progress('writing', f'正在写入 {total_sheets} 个工作表到文件...', 0.2)
            
            # 使用 pandas ExcelWriter 批量写入所有工作表（比逐个写入快）
            with pd.ExcelWriter(result_file_path, engine='openpyxl', mode='w') as writer:
                source_meta = self.get_metadata(source_file_id)
                logs = self._operation_log_cache.get(source_file_id, [])
                result_sheet_rows = []
                logs_by_sheet = {log.get('sheet_name'): log for log in logs}

                for sheet_name, info in results.items():
                    log = logs_by_sheet.get(sheet_name, {})
                    logic_text = log.get('logic') or ''
                    op_type = log.get('operation_type') or '分析结果'
                    result_sheet_rows.append({
                        '结果工作表': sheet_name,
                        '操作类型': op_type,
                        '逻辑说明': log.get('logic_description') or op_type,
                        'SQL/逻辑': logic_text,
                        '数据行数': info.get('row_count', 0),
                        '数据列数': len(info.get('columns', [])) if isinstance(info.get('columns'), list) else 0,
                        '执行耗时(ms)': log.get('execution_time_ms', 0),
                        '生成时间': (
                            log.get('created_at').strftime('%Y-%m-%d %H:%M:%S')
                            if log.get('created_at') else ''
                        ),
                        '内存表名': info.get('table_name', ''),
                    })

                # 分析元数据工作表（固定保留）
                summary_rows = [
                    {'字段': 'source_file_id', '值': source_file_id},
                    {'字段': 'source_file_name', '值': source_meta.original_name if source_meta else ''},
                    {'字段': 'export_file_name', '值': filename},
                    {'字段': 'exported_at', '值': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
                    {'字段': 'result_sheet_count', '值': total_sheets},
                    {'字段': 'result_total_rows', '值': sum(r.get('row_count', 0) for r in results.values())},
                    {'字段': 'operation_log_count', '值': len(logs)},
                ]
                summary_df = pd.DataFrame(summary_rows)
                detail_df = pd.DataFrame(result_sheet_rows)
                summary_df.to_excel(writer, sheet_name='分析元数据', index=False, startrow=0)
                if not detail_df.empty:
                    detail_start_row = len(summary_rows) + 3
                    detail_df.to_excel(writer, sheet_name='分析元数据', index=False, startrow=detail_start_row)
                meta_ws = writer.sheets.get('分析元数据')
                if meta_ws is not None:
                    _apply_numeric_number_formats(meta_ws, summary_df, data_start_row=2)
                    if not detail_df.empty:
                        _apply_numeric_number_formats(meta_ws, detail_df, data_start_row=detail_start_row + 2)

                # 操作日志工作表（兼容保留）
                if logs:
                    log_data = []
                    for log in logs:
                        logic_display = log['logic'][:500] + '...' if len(log['logic']) > 500 else log['logic']
                        log_data.append({
                            '序号': log['seq'],
                            '结果工作表': log['sheet_name'],
                            '操作类型': log['operation_type'],
                            '逻辑说明': log.get('logic_description') or log['operation_type'],
                            '计算逻辑': logic_display,
                            '生成时间': log['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
                            '数据行数': log['row_count'],
                            '耗时(ms)': log['execution_time_ms']
                        })
                    log_df = pd.DataFrame(log_data)
                    log_df.to_excel(writer, sheet_name=self.OPERATION_LOG_SHEET_NAME, index=False)
                    log_ws = writer.sheets.get(self.OPERATION_LOG_SHEET_NAME)
                    if log_ws is not None:
                        _apply_numeric_number_formats(log_ws, log_df, data_start_row=2)
                    logger.debug(f'操作日志工作表写入完成: {len(logs)} 条记录')
                
                # 写入每个结果表
                for idx, (sheet_name, info) in enumerate(results.items(), 1):
                    table_name = info['table_name']
                    row_count = info['row_count']
                    
                    progress = 0.2 + (idx / total_sheets) * 0.7  # 20% - 90%
                    _notify_progress('writing', f'正在写入工作表 [{sheet_name}] ({idx}/{total_sheets}): {row_count:,} 行...', progress)
                    
                    # 从 DuckDB 读取数据
                    df = duckdb_mgr.conn.execute(f'SELECT * FROM "{table_name}"').fetchdf()
                    export_df = _round_numeric_dataframe(df)
                    
                    # 使用 pandas 写入 Excel（批量写入，快5-10倍）
                    export_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets.get(sheet_name)
                    if ws is not None:
                        _apply_numeric_number_formats(ws, export_df, data_start_row=2)
                    
                    logger.debug(f'工作表写入完成: {sheet_name}, rows={len(export_df)}')
            
            # 阶段3：更新元数据
            _notify_progress('finalizing', f'正在更新文件信息...', 0.9)
            await self._update_file_metadata(result_meta.file_id, result_file_path)
            
            logs_info = f' (含操作日志)' if logs else ''
            _notify_progress('complete', f'文件保存完成: {total_sheets} 个工作表{logs_info}', 1.0)
            logger.info(f'内存结果已保存到文件: file_id={result_meta.file_id}, sheets={list(results.keys())}, total_sheets={total_sheets}, logs={len(logs)}')
            return result_meta
            
        except Exception as e:
            logger.error(f'保存内存结果到文件失败: {e}')
            import traceback
            logger.debug(traceback.format_exc())
            _notify_progress('error', f'保存失败: {str(e)}', None)
            return None


# 全局单例
large_file_storage = LargeFileStorage()
