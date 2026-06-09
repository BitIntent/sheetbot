# ==============================================================================
# DuckDB Excel 管理器
# 使用 DuckDB 高效读取和处理大型 Excel 文件
# 替代 openpyxl 的数据读取操作，保留 openpyxl 用于样式和公式
# ==============================================================================
import os
import threading
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple, Callable
import duckdb

from ..utils.logger import get_logger
from .sheet_normalizer import (
    dataframe_to_arrow_with_fallback,
    normalize_dataframe_for_duckdb,
)

logger = get_logger('large_file.duckdb')


# ==============================================================================
# DuckDB Excel 管理器
# ==============================================================================
class DuckDBExcelManager:
    """
    DuckDB Excel 管理器 - 单例模式
    
    职责：
    1. 高效读取 Excel 文件到内存表
    2. 执行 SQL 查询和聚合
    3. 生成透视表
    4. 导出数据到 Excel
    
    优势：
    - 读取速度比 openpyxl 快 10 倍以上
    - 支持 SQL 查询和聚合
    - 列式存储，内存占用低
    """
    _instance = None
    _lock = threading.Lock()
    
    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._initialized = True
        
        # 使用内存数据库
        self.conn = duckdb.connect(':memory:')
        
        # 安装并加载 Excel 扩展
        try:
            self.conn.execute("INSTALL excel")
            self.conn.execute("LOAD excel")
            logger.info("DuckDB Excel 扩展加载成功")
        except Exception as e:
            logger.error(f"DuckDB Excel 扩展加载失败: {e}")
            raise
        
        # 缓存已加载的文件信息 {file_id: {table_name, sheet_names, ...}}
        self._loaded_files: Dict[str, Dict[str, Any]] = {}
        self._cache_lock = threading.Lock()
        self._conn_lock = threading.RLock()
        # 同一 file_id 的全表加载互斥锁，避免并发 load_all_sheets 导致 DuckDB 连接冲突
        self._file_load_locks: Dict[str, threading.Lock] = {}
        # DuckDB 稳定性保护：连续失败到阈值后进行软重置
        self._consecutive_failures = 0
        self._failure_threshold = 3

    def _get_file_load_lock(self, file_id: str) -> threading.Lock:
        """获取（或创建）指定 file_id 的加载互斥锁。"""
        with self._cache_lock:
            lock = self._file_load_locks.get(file_id)
            if lock is None:
                lock = threading.Lock()
                self._file_load_locks[file_id] = lock
            return lock

    def execute_fetchdf(self, sql: str, params: Optional[Any] = None):
        """线程安全执行 SQL 并返回 DataFrame。"""
        with self._conn_lock:
            if params is None:
                return self.conn.execute(sql).fetchdf()
            return self.conn.execute(sql, params).fetchdf()

    def execute_fetchone(self, sql: str, params: Optional[Any] = None):
        """线程安全执行 SQL 并返回一行。"""
        with self._conn_lock:
            if params is None:
                return self.conn.execute(sql).fetchone()
            return self.conn.execute(sql, params).fetchone()

    def execute_fetchall(self, sql: str, params: Optional[Any] = None):
        """线程安全执行 SQL 并返回所有行。"""
        with self._conn_lock:
            if params is None:
                return self.conn.execute(sql).fetchall()
            return self.conn.execute(sql, params).fetchall()

    def _record_success(self):
        """记录一次成功操作，重置连续失败计数。"""
        self._consecutive_failures = 0

    def _record_failure(self, context: str, exc: Exception):
        """记录失败并在阈值后软重置连接，避免进入长期坏状态。"""
        self._consecutive_failures += 1
        logger.error(
            "DuckDB 操作失败: context=%s, consecutive_failures=%d, error=%s",
            context,
            self._consecutive_failures,
            exc,
        )
        if self._consecutive_failures >= self._failure_threshold:
            self._soft_reset_connection(context)

    def _soft_reset_connection(self, reason: str):
        """软重置 DuckDB 连接与缓存，避免坏连接拖死后续请求。"""
        with self._conn_lock:
            try:
                try:
                    self.conn.close()
                except Exception:
                    pass
                self.conn = duckdb.connect(':memory:')
                try:
                    self.conn.execute("INSTALL excel")
                    self.conn.execute("LOAD excel")
                except Exception:
                    # 扩展加载失败保留异常给后续真实调用暴露
                    pass
                with self._cache_lock:
                    self._loaded_files.clear()
                    self._file_load_locks.clear()
                logger.warning("DuckDB 连接已软重置: reason=%s", reason)
            finally:
                self._consecutive_failures = 0
    
    # ==========================================================================
    # 文件加载
    # ==========================================================================
    def load_excel(
        self, 
        file_path: str, 
        file_id: str, 
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        加载 Excel 文件到 DuckDB 内存表
        
        Args:
            file_path: Excel 文件路径
            file_id: 文件唯一标识
            sheet_name: 指定工作表名称，None 则加载第一个
            
        Returns:
            {table_name, row_count, col_count, columns}
        """
        # 生成表名（使用 file_id 作为表名，确保唯一）
        table_name = self._get_table_name(file_id, sheet_name)
        
        with self._cache_lock:
            # 检查是否已加载
            if file_id in self._loaded_files:
                cached = self._loaded_files[file_id]
                if sheet_name is None or sheet_name in cached.get('sheets', {}):
                    logger.debug(f'文件已缓存: file_id={file_id}')
                    return cached.get('sheets', {}).get(sheet_name or cached.get('default_sheet'))
        
        try:
            import time
            import os
            import pandas as pd
            
            file_size_mb = os.path.getsize(file_path) / 1024 / 1024
            logger.info(f'开始加载 Excel: file_path={file_path}, sheet={sheet_name}, size={file_size_mb:.2f}MB')
            
            start_time = time.time()
            
            # ================================================================
            # 方案 B：统一使用 pandas 读取 Excel，再导入 DuckDB
            # pandas + openpyxl 读取 Excel 比 DuckDB 原生 read_xlsx 快 10 倍以上
            # ================================================================
            
            # 确定要读取的工作表
            # pandas 的 sheet_name=None 会返回 dict，我们需要明确指定工作表
            target_sheet = sheet_name
            if target_sheet is None:
                # 获取第一个工作表名称
                sheet_names = self.get_sheet_names(file_path)
                target_sheet = sheet_names[0] if sheet_names else 0
                logger.info(f'未指定工作表，使用第一个: {target_sheet}')
            
            # 使用智能读取：检测公式并计算
            df = self._smart_read_excel(file_path, target_sheet)
            # 兼容：空工作表/仅图片工作表可能导致 0 列，DuckDB 无法直接导入
            if len(df.columns) == 0:
                logger.warning(f'工作表 {target_sheet} 读取结果为 0 列，创建占位列以兼容 DuckDB 导入')
                import pandas as pd
                df = pd.DataFrame({"__EMPTY__": []})
            
            read_time = time.time() - start_time
            logger.info(f'读取完成: {len(df)} 行 x {len(df.columns)} 列, 耗时 {read_time:.2f}秒')
            
            # 更新 sheet_name 为实际读取的工作表名
            if sheet_name is None:
                sheet_name = target_sheet
            
            # 导入到 DuckDB（使用 pyarrow 作为中间层，避免类型兼容问题）
            logger.info('正在导入数据到 DuckDB...')
            self._import_dataframe_to_duckdb(df, table_name, sheet_name=target_sheet)
            import_time = time.time() - start_time - read_time
            logger.info(f'DuckDB 导入完成: 耗时 {import_time:.2f}秒')
            
            total_time = time.time() - start_time
            
            # 获取表信息
            info = self._get_table_info(table_name)
            if info.get("columns") == ["__EMPTY__"]:
                info = {**info, "col_count": 0, "columns": []}
            
            # 缓存文件信息
            with self._cache_lock:
                if file_id not in self._loaded_files:
                    self._loaded_files[file_id] = {
                        'file_path': file_path,
                        'sheets': {},
                        'default_sheet': sheet_name
                    }
                self._loaded_files[file_id]['sheets'][sheet_name or 'default'] = {
                    'table_name': table_name,
                    **info
                }
            
            logger.info(f'Excel 加载完成: table={table_name}, rows={info["row_count"]}, cols={info["col_count"]}, 总耗时={total_time:.2f}秒')
            self._record_success()
            return {'table_name': table_name, **info}
            
        except Exception as e:
            logger.error(f'加载 Excel 失败: {e}')
            self._record_failure("load_excel", e)
            raise
    
    def _detect_title_row(self, file_path: str, sheet_name: str) -> int:
        """
        检测标题行
        
        标题行特征：
        1. 第一行只有第一列（A1）有内容，其他列为空或很少内容
        2. 第一行内容包含"表"、"数据"、"明细"等关键词
        
        Returns:
            标题行数量（0 或 1）
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # 读取第一行
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            
            # 检查第一行：如果只有第一列有内容，其他列大部分为空，则认为是标题行
            first_col_value = first_row[0] if first_row else None
            other_cols_non_empty = sum(1 for val in first_row[1:] if val is not None and str(val).strip())
            
            # 标题行判断条件（收紧）：
            # 1. 第一列有内容
            # 2. 其他列必须全部为空（避免把正常双列表头误判为标题行）
            # 3. 第一列内容包含常见标题关键词（辅助）
            is_title_row = False
            if first_col_value and isinstance(first_col_value, str):
                first_col_str = str(first_col_value).strip()
                # 检查是否包含标题关键词
                title_keywords = ['表', '数据', '明细', '统计', '汇总', '分析', '报表']
                has_title_keyword = any(keyword in first_col_str for keyword in title_keywords)
                
                # 仅在“第一列独占”时认定标题行；关键词用于增强可信度
                if other_cols_non_empty == 0 and (has_title_keyword or len(first_col_str) >= 4):
                    is_title_row = True
                    logger.info(f'检测到标题行: A1="{first_col_str}", 其他列非空数={other_cols_non_empty}')
            
            wb.close()
            return 1 if is_title_row else 0
        except Exception as e:
            logger.warning(f'检测标题行失败: {e}，默认不跳过')
            return 0
    
    def _smart_read_excel(self, file_path: str, sheet_name: str) -> 'pd.DataFrame':
        """
        智能读取 Excel 文件
        
        1. 检测并跳过标题行
        2. 检查是否有公式
        3. 如果有公式，尝试读取 cached value
        4. 如果 cached value 为空，计算公式
        5. 如果没有公式，直接用 pandas 读取
        
        优化：对超大文件使用流式读取
        """
        import pandas as pd
        from openpyxl import load_workbook
        
        # 获取文件大小
        file_size_mb = os.path.getsize(file_path) / 1024 / 1024
        logger.info(f'智能读取 Excel: {file_path}, sheet={sheet_name}, size={file_size_mb:.1f}MB')
        
        # 检测标题行
        title_row_count = self._detect_title_row(file_path, sheet_name)
        # pandas header 使用 0-based；openpyxl 行号使用 1-based
        pandas_header_idx = title_row_count
        header_row_excel = title_row_count + 1
        
        # 对于超大文件（>50MB），优先走 pandas 快速读取；
        # 若检测到公式列缓存值缺失，再回退到轻量公式补算，避免报表侧公式列为空。
        if file_size_mb > 50:
            logger.info(f'超大文件（{file_size_mb:.1f}MB），直接使用 pandas 读取')
            if title_row_count > 0:
                logger.info(f'跳过标题行（{title_row_count}行），表头从第{header_row_excel}行开始')
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', header=pandas_header_idx)
            try:
                formula_cols = self._detect_formula_columns(file_path, sheet_name, header_row_excel)
                if formula_cols:
                    logger.info(f'超大文件检测到公式列: {formula_cols}')
                    needs_calculation = False
                    for col in formula_cols:
                        if col not in df.columns:
                            continue
                        non_null = df[col].dropna()
                        if len(non_null) == 0:
                            needs_calculation = True
                            logger.warning(f'超大文件公式列 [{col}] 无缓存值，触发补算')
                            break
                        numeric_series = pd.to_numeric(non_null, errors='coerce')
                        if len(numeric_series) and numeric_series.notna().sum() > 0 and (numeric_series.fillna(0) == 0).all():
                            needs_calculation = True
                            logger.warning(f'超大文件公式列 [{col}] 缓存值疑似无效（全0），触发补算')
                            break
                    if needs_calculation:
                        df = self._calculate_formulas_directly(file_path, sheet_name, df)
                        logger.info('超大文件公式补算完成')
            except Exception as e:
                logger.warning(f'超大文件公式补算检查失败，继续使用原始读取结果: {e}')
            return df
        
        # 步骤 1：检查是否有公式（检查数据开始行）
        logger.info('检查公式...')
        wb_check = load_workbook(file_path, read_only=True, data_only=False)
        ws_check = wb_check[sheet_name] if sheet_name in wb_check.sheetnames else wb_check.active
        
        has_formula = False
        formula_cols = []
        formula_col_indices = []
        
        # 检查数据开始行（跳过标题行后的第一行数据）的公式
        data_start_row = header_row_excel + 1  # 数据开始行号
        rows = list(ws_check.iter_rows(min_row=data_start_row, max_row=data_start_row))
        if rows:
            for col_idx, cell in enumerate(rows[0], 1):
                if cell.value and str(cell.value).startswith('='):
                    has_formula = True
                    # 从表头行获取列名
                    header_cell = ws_check.cell(row=header_row_excel, column=col_idx)
                    formula_cols.append(header_cell.value)
                    formula_col_indices.append(col_idx)
        wb_check.close()
        
        if not has_formula:
            # 没有公式，直接用 pandas 读取
            logger.info('未检测到公式，使用 pandas 直接读取')
            if title_row_count > 0:
                logger.info(f'跳过标题行（{title_row_count}行），表头从第{header_row_excel}行开始')
            return pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', header=pandas_header_idx)
        
        logger.info(f'检测到公式列: {formula_cols}')
        
        # 步骤 2：尝试读取 cached value（使用流式读取）
        logger.info('读取公式缓存值...')
        wb_data = load_workbook(file_path, read_only=True, data_only=True)
        ws_data = wb_data[sheet_name] if sheet_name in wb_data.sheetnames else wb_data.active
        
        # 流式读取数据（避免一次性加载到内存）
        # 如果检测到标题行，跳过第一行
        data = []
        row_count = 0
        start_row = 1 + title_row_count  # 如果有标题行，从第二行开始读取
        for row_idx, row in enumerate(ws_data.iter_rows(values_only=True), 1):
            if row_idx < start_row:
                continue  # 跳过标题行
            data.append(row)
            row_count += 1
            if row_count % 100000 == 0:
                logger.info(f'已读取 {row_count} 行...')
        
        wb_data.close()
        logger.info(f'读取完成: {row_count} 行（已跳过{title_row_count}行标题）')
        
        if not data:
            return pd.DataFrame()
        
        # 第一行作为表头
        headers = data[0]
        df = pd.DataFrame(data[1:], columns=headers)
        
        # 检查公式列是否有有效值
        needs_calculation = False
        for col in formula_cols:
            if col in df.columns:
                # 检查是否全是 None 或 0
                non_null = df[col].dropna()
                if len(non_null) == 0 or (non_null == 0).all():
                    needs_calculation = True
                    logger.warning(f'公式列 [{col}] 无缓存值，需要计算')
                    break
                else:
                    sample = df[col].head(3).tolist()
                    logger.info(f'公式列 [{col}] 已有缓存值: {sample}')
        
        if not needs_calculation:
            logger.info('公式列已有缓存值，无需重新计算')
            return df

        # 步骤 3：使用 Python 直接计算公式（比 formulas 库快得多）
        logger.info('使用 Python 直接计算公式...')
        try:
            df = self._calculate_formulas_directly(file_path, sheet_name, df)
            logger.info('公式计算完成')
            
            # 验证计算结果
            for col in formula_cols:
                if col in df.columns:
                    sample = df[col].head(3).tolist()
                    logger.info(f'[计算后] 列 [{col}] 样本: {sample}')
            
            return df
        except Exception as e:
            logger.error(f'公式计算失败: {e}')
            logger.warning('回退到 cached value（可能为 0 或 None）')
            logger.warning('建议：用 Microsoft Excel 打开文件并保存，以缓存公式计算结果')
            return df

    def _detect_formula_columns(self, file_path: str, sheet_name: str, header_row: int) -> list:
        """
        轻量检测公式列：仅检查数据起始行，避免全表扫描。
        """
        from openpyxl import load_workbook
        wb_check = load_workbook(file_path, read_only=True, data_only=False)
        ws_check = wb_check[sheet_name] if sheet_name in wb_check.sheetnames else wb_check.active
        data_start_row = header_row + 1
        rows = list(ws_check.iter_rows(min_row=data_start_row, max_row=data_start_row))
        formula_cols = []
        if rows:
            for col_idx, cell in enumerate(rows[0], 1):
                if cell.value and str(cell.value).startswith('='):
                    header_cell = ws_check.cell(row=header_row, column=col_idx)
                    if header_cell.value:
                        formula_cols.append(header_cell.value)
        wb_check.close()
        return formula_cols
    
    def _calculate_formulas_directly(self, file_path: str, sheet_name: str, df: 'pd.DataFrame') -> 'pd.DataFrame':
        """
        直接用 Python/pandas 计算 Excel 公式
        
        比 formulas 库快 100 倍以上，适合大文件
        仅支持简单的算术公式，复杂公式会跳过
        """
        import pandas as pd
        import re
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        logger.info('读取公式定义...')
        
        # 检测标题行
        title_row_count = self._detect_title_row(file_path, sheet_name)
        header_row_num = 1 + title_row_count  # 表头行号
        data_start_row = header_row_num + 1   # 数据开始行号
        
        # 读取公式（从数据开始行）
        wb = load_workbook(file_path, read_only=True, data_only=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        # 获取列名映射：列字母 -> 列名（从表头行读取）
        headers = {}
        header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num))[0]
        for cell in header_row:
            col_letter = get_column_letter(cell.column)
            headers[col_letter] = cell.value
        
        # 获取公式定义（从数据开始行）
        formula_row = list(ws.iter_rows(min_row=data_start_row, max_row=data_start_row))[0]
        formulas_to_calc = {}  # {列名: 公式}
        
        for cell in formula_row:
            if cell.value and str(cell.value).startswith('='):
                col_letter = get_column_letter(cell.column)
                col_name = headers.get(col_letter)
                if col_name:
                    formulas_to_calc[col_name] = str(cell.value)
                    logger.info(f'公式列 [{col_name}]: {cell.value}')
        
        wb.close()
        
        # 解析并计算每个公式
        for col_name, formula in formulas_to_calc.items():
            try:
                result = self._evaluate_formula(df, formula, headers)
                if result is not None:
                    df[col_name] = result
                    logger.info(f'列 [{col_name}] 计算完成')
            except Exception as e:
                logger.warning(f'列 [{col_name}] 公式计算失败: {e}')
        
        return df
    
    def _evaluate_formula(self, df: 'pd.DataFrame', formula: str, headers: dict) -> 'pd.Series':
        """
        将 Excel 公式转换为 pandas 表达式并计算
        
        支持的公式：
        - 简单算术: =A2*B2, =A2+B2-C2, =A2/B2
        - 带括号: =A2*(1-B2)
        - 幂运算: =A2^2
        - IF 函数: =IF(condition, true_value, false_value)
        - 比较运算: =, <>, <, >, <=, >=
        """
        import re
        import pandas as pd
        import numpy as np
        
        # 移除开头的 =
        expr = formula[1:] if formula.startswith('=') else formula
        original_expr = expr
        
        # 创建列字母到列名的反向映射
        letter_to_name = {letter: name for letter, name in headers.items()}
        
        # 替换单元格引用为 DataFrame 列引用
        # 匹配 $A$2, $A2, A$2, A2, AA2 等格式
        def replace_cell_ref(match):
            col_letter = match.group(1).replace('$', '')
            col_name = letter_to_name.get(col_letter)
            if col_name and col_name in df.columns:
                return f'df["{col_name}"]'
            return match.group(0)
        
        # 替换单元格引用（支持绝对引用 $A$2）
        expr = re.sub(r'\$?([A-Z]+)\$?\d+', replace_cell_ref, expr)
        
        # 替换 Excel 运算符为 Python 等价物
        expr = expr.replace('^', '**')  # 幂运算
        expr = expr.replace('<>', '!=')  # 不等于
        
        # 处理 IF 函数: IF(condition, true_val, false_val) -> np.where(condition, true_val, false_val)
        expr = self._convert_if_function(expr)
        
        logger.debug(f'原始公式: {original_expr}')
        logger.debug(f'转换后表达式: {expr}')
        
        # 安全执行表达式
        try:
            # 创建安全的执行环境
            safe_locals = {'df': df, 'pd': pd, 'np': np}
            result = eval(expr, {"__builtins__": {}}, safe_locals)
            
            # 确保结果是 Series
            if isinstance(result, pd.Series):
                return result
            elif isinstance(result, (int, float)):
                return pd.Series([result] * len(df))
            else:
                return None
        except Exception as e:
            logger.warning(f'公式执行失败: {expr}, error: {e}')
            return None
    
    def _convert_if_function(self, expr: str) -> str:
        """
        将 Excel IF 函数转换为 numpy.where
        
        Excel: IF(condition, true_value, false_value)
        Python: np.where(condition, true_value, false_value)
        
        支持嵌套 IF
        """
        import re
        
        # 匹配 IF(...) 函数，使用递归匹配括号
        def find_if_and_replace(s):
            # 查找 IF( 的位置
            match = re.search(r'\bIF\s*\(', s, re.IGNORECASE)
            if not match:
                return s
            
            start = match.start()
            paren_start = match.end() - 1  # ( 的位置
            
            # 找到匹配的闭括号
            depth = 1
            i = paren_start + 1
            while i < len(s) and depth > 0:
                if s[i] == '(':
                    depth += 1
                elif s[i] == ')':
                    depth -= 1
                i += 1
            
            if depth != 0:
                return s  # 括号不匹配，返回原字符串
            
            paren_end = i - 1  # ) 的位置
            
            # 提取 IF 的参数
            args_str = s[paren_start + 1:paren_end]
            
            # 按逗号分割参数（注意处理嵌套括号）
            args = self._split_formula_args(args_str)
            
            if len(args) != 3:
                logger.warning(f'IF 函数参数数量错误: {args_str}')
                return s
            
            condition, true_val, false_val = args
            
            # 递归处理嵌套的 IF
            condition = find_if_and_replace(condition)
            true_val = find_if_and_replace(true_val)
            false_val = find_if_and_replace(false_val)
            
            # 转换为 np.where
            replacement = f'np.where({condition}, {true_val}, {false_val})'
            
            # 替换原字符串
            result = s[:start] + replacement + s[paren_end + 1:]
            
            # 继续查找其他 IF
            return find_if_and_replace(result)
        
        return find_if_and_replace(expr)
    
    def _split_formula_args(self, args_str: str) -> list:
        """
        按逗号分割公式参数，正确处理嵌套括号
        
        例如: "A2=0, 0, B2/C2" -> ["A2=0", "0", "B2/C2"]
        """
        args = []
        current = ""
        depth = 0
        
        for char in args_str:
            if char == '(':
                depth += 1
                current += char
            elif char == ')':
                depth -= 1
                current += char
            elif char == ',' and depth == 0:
                args.append(current.strip())
                current = ""
            else:
                current += char
        
        if current.strip():
            args.append(current.strip())
        
        return args
    
    def _import_dataframe_to_duckdb(self, df, table_name: str, sheet_name: str = ""):
        """
        将 pandas DataFrame 导入 DuckDB
        
        使用 pyarrow 作为中间层，避免 DuckDB 直接读取 pandas 的类型兼容问题
        """
        import pyarrow as pa
        
        # 非标准表兼容：表头/类型清洗委托到标准化模块
        normalized_df = normalize_dataframe_for_duckdb(
            df, logger, sheet_name=sheet_name
        )
        arrow_table = dataframe_to_arrow_with_fallback(normalized_df, table_name, logger)
        
        # DuckDB 直接从 pyarrow Table 创建表（最高效的方式）
        with self._conn_lock:
            self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')

        if len(arrow_table) == 0:
            # -------------------------------------------------------------------
            # 空表路径：跳过 conn.register()
            # DuckDB C++ 层对零行 Arrow Table 的 register 存在堆损坏 bug
            # （glibc malloc unsorted double linked list corrupted → SIGABRT）
            # 改用 DDL 直接建空结构表，下游 SQL 查询正常可用
            # -------------------------------------------------------------------
            _ARROW_TO_DUCK = {
                pa.int8(): 'TINYINT',   pa.int16(): 'SMALLINT',
                pa.int32(): 'INTEGER',  pa.int64(): 'BIGINT',
                pa.uint8(): 'UTINYINT', pa.uint16(): 'USMALLINT',
                pa.uint32(): 'UINTEGER',pa.uint64(): 'UBIGINT',
                pa.float32(): 'FLOAT',  pa.float64(): 'DOUBLE',
                pa.bool_(): 'BOOLEAN',  pa.date32(): 'DATE',
                pa.utf8(): 'VARCHAR',   pa.large_utf8(): 'VARCHAR',
            }
            col_defs = []
            for field in arrow_table.schema:
                duck_type = _ARROW_TO_DUCK.get(field.type, 'VARCHAR')
                col_defs.append(f'"{field.name}" {duck_type}')
            ddl = ', '.join(col_defs) if col_defs else '"_placeholder" VARCHAR'
            with self._conn_lock:
                self.conn.execute(f'CREATE TABLE "{table_name}" ({ddl})')
            logger.info(f'空表 DDL 建表完成（跳过 Arrow register）: {table_name}')
        else:
            with self._conn_lock:
                self.conn.register('_arrow_table', arrow_table)
                self.conn.execute(f'CREATE TABLE "{table_name}" AS SELECT * FROM _arrow_table')
                self.conn.unregister('_arrow_table')
    
    def load_all_sheets(
        self,
        file_path: str,
        file_id: str,
        progress_callback: Optional[Callable[[str, int, int], None]] = None,
    ) -> Dict[str, Dict[str, Any]]:
        """
        加载 Excel 文件的所有工作表

        progress_callback: (sheet_name, current_index, total) -> None，每完成一个 sheet 调用
        """
        load_lock = self._get_file_load_lock(file_id)
        acquired = load_lock.acquire(blocking=False)

        # 若已有同文件加载任务，等待其完成后直接返回缓存，避免并发导入
        if not acquired:
            logger.info(f'检测到并发 load_all_sheets，请求等待复用已有加载: file_id={file_id}')
            with load_lock:
                pass
            with self._cache_lock:
                sheets = self._loaded_files.get(file_id, {}).get('sheets', {})
                return {name: info for name, info in sheets.items()}

        try:
            sheet_names = self.get_sheet_names(file_path)
            total = len(sheet_names)

            with self._cache_lock:
                loaded_sheets = set(self._loaded_files.get(file_id, {}).get('sheets', {}).keys())

            missing_sheets = [s for s in sheet_names if s not in loaded_sheets]
            if not missing_sheets:
                logger.info(f'所有工作表已加载，跳过重复导入: file_id={file_id}, sheets={sheet_names}')
                with self._cache_lock:
                    sheets = self._loaded_files.get(file_id, {}).get('sheets', {})
                    return {name: info for name, info in sheets.items()}

            results = {}
            failed_sheets: List[str] = []
            for idx, sheet in enumerate(sheet_names):
                # 已加载工作表直接复用缓存
                if sheet in loaded_sheets:
                    with self._cache_lock:
                        cached_info = self._loaded_files.get(file_id, {}).get('sheets', {}).get(sheet)
                    if cached_info:
                        results[sheet] = cached_info
                    if progress_callback:
                        progress_callback(sheet, idx, total)
                    continue

                try:
                    info = self.load_excel(file_path, file_id, sheet)
                    results[sheet] = info
                    self._record_success()
                    if progress_callback:
                        progress_callback(sheet, idx, total)
                except Exception as e:
                    logger.warning(f'加载工作表 {sheet} 失败: {e}')
                    failed_sheets.append(sheet)

            if failed_sheets:
                raise RuntimeError(
                    f"工作表加载失败({len(failed_sheets)}/{total}): {failed_sheets[:5]}"
                )
            return results
        finally:
            load_lock.release()
    
    # ==========================================================================
    # 数据类型转换
    # ==========================================================================
    def _to_json_compatible(self, data: list) -> list:
        """
        将数据转换为 JSON 兼容格式
        
        处理 DuckDB/pandas 返回的特殊类型：
        - Timestamp → ISO 格式字符串
        - date → ISO 格式字符串  
        - Decimal → float
        - bytes → base64 字符串
        - NaN/NaT → None
        """
        import pandas as pd
        from datetime import date, datetime
        from decimal import Decimal
        
        def convert_value(val):
            if val is None:
                return None
            if pd.isna(val):
                return None
            if isinstance(val, (pd.Timestamp, datetime)):
                return val.isoformat()
            if isinstance(val, date):
                return val.isoformat()
            if isinstance(val, Decimal):
                return float(val)
            if isinstance(val, bytes):
                return val.decode('utf-8', errors='replace')
            return val
        
        return [[convert_value(cell) for cell in row] for row in data]
    
    # ==========================================================================
    # 查询操作
    # ==========================================================================
    def _resolve_table_placeholders(
        self, 
        file_id: str, 
        sql: str, 
        default_sheet: Optional[str] = None
    ) -> str:
        """
        解析 SQL 中的表名占位符，支持跨工作表查询和结果表引用
        
        支持的占位符格式：
        - {table} : 引用当前/默认工作表
        - {table:工作表名} : 引用指定工作表（用于跨表 JOIN）
        - {table:结果_xxx} : 引用结果工作表（二次加工）
        
        Args:
            file_id: 文件ID
            sql: 包含占位符的 SQL 语句
            default_sheet: 默认工作表名（用于 {table} 占位符）
            
        Returns:
            替换占位符后的 SQL 语句
        """
        import re
        
        actual_sql = sql
        result_file_key = f"result_{file_id}"
        
        # 1. 替换 {table:sheet_name} 格式（跨表引用，包括结果表）
        pattern = r'\{table:([^}]+)\}'
        matches = re.findall(pattern, sql)
        for ref_sheet in matches:
            ref_sheet = ref_sheet.strip()
            table_name = None
            
            # 首先尝试从结果表中查找
            with self._cache_lock:
                if result_file_key in self._loaded_files:
                    result_sheets = self._loaded_files[result_file_key].get('sheets', {})
                    if ref_sheet in result_sheets:
                        table_name = result_sheets[ref_sheet].get('table_name')
                        logger.debug(f'[跨表查询] 从结果表找到: {ref_sheet} -> {table_name}')
            
            # 如果结果表中没有，从源数据表查找
            if not table_name:
                table_name = self._get_cached_table_name(file_id, ref_sheet)
            
            if not table_name:
                # 尝试自动加载该工作表（仅对源数据有效）
                logger.info(f'[跨表查询] 自动加载工作表: {ref_sheet}')
                file_info = self._loaded_files.get(file_id, {})
                file_path = file_info.get('file_path')
                if file_path:
                    self.load_excel(file_path, file_id, ref_sheet)
                    table_name = self._get_cached_table_name(file_id, ref_sheet)
                if not table_name:
                    raise ValueError(f'工作表未找到: {ref_sheet}（提示：可用的表包括源数据工作表和已生成的结果工作表）')
            
            actual_sql = actual_sql.replace(f'{{table:{ref_sheet}}}', f'"{table_name}"')
            logger.debug(f'[跨表查询] {ref_sheet} -> {table_name}')
        
        # 2. 替换 {table} 格式（当前表）
        if '{table}' in actual_sql:
            table_name = None
            
            # 如果 default_sheet 是结果表，优先从结果表中查找
            if default_sheet and default_sheet.startswith('结果_'):
                with self._cache_lock:
                    if result_file_key in self._loaded_files:
                        result_sheets = self._loaded_files[result_file_key].get('sheets', {})
                        if default_sheet in result_sheets:
                            table_name = result_sheets[default_sheet].get('table_name')
                            logger.debug(f'[{table}] 从结果表找到: {default_sheet} -> {table_name}')
            
            # 如果不是结果表或结果表中没找到，从源数据表查找
            if not table_name:
                table_name = self._get_cached_table_name(file_id, default_sheet)
            
            if not table_name:
                raise ValueError(f'文件未加载: file_id={file_id}, sheet={default_sheet}')
            actual_sql = actual_sql.replace('{table}', f'"{table_name}"')
        
        return actual_sql
    
    def get_loaded_sheets(self, file_id: str) -> List[str]:
        """获取文件已加载的工作表列表"""
        with self._cache_lock:
            file_info = self._loaded_files.get(file_id, {})
            return list(file_info.get('sheets', {}).keys())
    
    def query(
        self, 
        file_id: str, 
        sql: str, 
        sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        执行 SQL 查询，支持跨工作表联合查询
        
        Args:
            file_id: 文件ID
            sql: SQL 查询语句
                - 使用 {table} 引用当前工作表
                - 使用 {table:工作表名} 引用其他工作表（跨表 JOIN）
            sheet_name: 当前工作表名称
            
        Returns:
            查询结果列表
            
        示例：
            SELECT s.*, c.客户名称 
            FROM {table:销售明细} s
            JOIN {table:客户明细} c ON s.客户ID = c.客户ID
            WHERE c.大区 = '华南'
        """
        # 解析所有表名占位符
        actual_sql = self._resolve_table_placeholders(file_id, sql, sheet_name)
        
        try:
            logger.debug(f'执行 SQL: {actual_sql[:200]}...')
            result = self.execute_fetchdf(actual_sql)
            logger.debug(f'查询返回 {len(result)} 行')
            
            # 转换为 JSON 兼容的字典列表
            columns = result.columns.tolist()
            data = self._to_json_compatible(result.values.tolist())
            return [dict(zip(columns, row)) for row in data]
        except Exception as e:
            logger.error(f'查询失败: sql={actual_sql}, error={e}')
            raise
    
    def query_df(
        self, 
        file_id: str, 
        sql: str, 
        sheet_name: Optional[str] = None
    ):
        """
        执行 SQL 查询，返回 DataFrame，支持跨工作表联合查询
        
        占位符格式同 query() 方法
        """
        actual_sql = self._resolve_table_placeholders(file_id, sql, sheet_name)
        return self.execute_fetchdf(actual_sql)
    
    def get_preview(
        self, 
        file_id: str, 
        sheet_name: Optional[str] = None,
        limit: int = 500,
        offset: int = 0
    ) -> Dict[str, Any]:
        """
        获取数据预览
        
        Returns:
            {columns, data, row_count, col_count}
        """
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        safe_limit = max(1, int(limit))
        safe_offset = max(0, int(offset))
        logger.debug(f'获取预览: table={table_name}, limit={safe_limit}, offset={safe_offset}')
        
        # 获取列信息
        columns = self._get_columns(table_name)
        
        # 获取预览数据
        result = self.execute_fetchdf(
            f'SELECT * FROM "{table_name}" LIMIT {safe_limit} OFFSET {safe_offset}'
        )
        
        # 转换为 JSON 兼容格式（处理 Timestamp 等特殊类型）
        data = self._to_json_compatible(result.values.tolist())
        
        # 获取总行数
        total_rows = self.execute_fetchone(f'SELECT COUNT(*) FROM "{table_name}"')[0]
        
        logger.debug(f'预览完成: {len(data)} 行 / 总 {total_rows} 行')
        return {
            'columns': columns,
            'data': data,
            'row_count': total_rows,
            'col_count': len(columns),
            'preview_rows': len(data),
            'offset': safe_offset,
            'limit': safe_limit,
            'has_more': (safe_offset + len(data)) < total_rows,
        }
    
    def get_preview_from_table(
        self,
        table_name: str,
        limit: int = 500,
        offset: int = 0
    ) -> Dict[str, Any]:
        """
        从指定的 DuckDB 表获取预览（用于内存结果）
        
        Args:
            table_name: DuckDB 表名
            limit: 预览行数
            
        Returns:
            {columns, data, row_count, col_count, preview_rows}
        """
        safe_limit = max(1, int(limit))
        safe_offset = max(0, int(offset))
        logger.debug(f'从表获取预览: table={table_name}, limit={safe_limit}, offset={safe_offset}')
        
        # 获取列信息
        columns = self._get_columns(table_name)
        
        # 获取预览数据
        result = self.execute_fetchdf(
            f'SELECT * FROM "{table_name}" LIMIT {safe_limit} OFFSET {safe_offset}'
        )
        
        # 转换为 JSON 兼容格式
        data = self._to_json_compatible(result.values.tolist())
        
        # 获取总行数
        total_rows = self.execute_fetchone(f'SELECT COUNT(*) FROM "{table_name}"')[0]
        
        logger.debug(f'表预览完成: {len(data)} 行 / 总 {total_rows} 行')
        return {
            'columns': columns,
            'data': data,
            'row_count': total_rows,
            'col_count': len(columns),
            'preview_rows': len(data),
            'offset': safe_offset,
            'limit': safe_limit,
            'has_more': (safe_offset + len(data)) < total_rows,
        }
    
    def get_unique_values(
        self, 
        file_id: str, 
        column: str, 
        sheet_name: Optional[str] = None,
        limit: int = 1000
    ) -> List[Any]:
        """获取指定列的唯一值"""
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        logger.debug(f'获取唯一值: table={table_name}, column={column}, limit={limit}')
        result = self.conn.execute(f"""
            SELECT DISTINCT "{column}" 
            FROM "{table_name}" 
            WHERE "{column}" IS NOT NULL
            LIMIT {limit}
        """).fetchall()
        
        values = [row[0] for row in result]
        logger.debug(f'唯一值数量: {len(values)}')
        return values
    
    def get_statistics(
        self, 
        file_id: str, 
        column: str, 
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """获取指定列的统计信息"""
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        logger.debug(f'获取列统计: table={table_name}, column={column}')
        result = self.conn.execute(f"""
            SELECT 
                COUNT(*) as count,
                COUNT(DISTINCT "{column}") as unique_count,
                MIN("{column}") as min_val,
                MAX("{column}") as max_val,
                AVG(TRY_CAST("{column}" AS DOUBLE)) as avg_val,
                SUM(TRY_CAST("{column}" AS DOUBLE)) as sum_val
            FROM "{table_name}"
        """).fetchdf()
        
        stats = result.to_dict('records')[0]
        logger.debug(f'统计结果: count={stats.get("count")}, unique={stats.get("unique_count")}')
        return stats
    
    # ==========================================================================
    # 透视表操作
    # ==========================================================================
    def create_pivot_table(
        self,
        file_id: str,
        row_field: str,
        column_field: str,
        value_field: str,
        agg_func: str = 'SUM',
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        创建数据透视表
        
        Args:
            file_id: 文件ID
            row_field: 行字段
            column_field: 列字段
            value_field: 值字段
            agg_func: 聚合函数 (SUM, COUNT, AVG, MIN, MAX)
            sheet_name: 工作表名称
            
        Returns:
            {columns, data, row_labels, col_labels}
        """
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        try:
            value_expr = (
                f"TRY_CAST("
                f"NULLIF(REGEXP_REPLACE(CAST(\"{value_field}\" AS VARCHAR), '[^0-9.-]', '', 'g'), '') "
                f"AS DOUBLE)"
            )

            # 先做可转换率诊断，避免“执行成功但结果全空”
            stats_sql = f"""
                SELECT
                    COUNT(*) FILTER (
                        WHERE "{row_field}" IS NOT NULL
                          AND "{column_field}" IS NOT NULL
                          AND "{value_field}" IS NOT NULL
                    ) AS total_rows,
                    COUNT(*) FILTER (
                        WHERE "{row_field}" IS NOT NULL
                          AND "{column_field}" IS NOT NULL
                          AND {value_expr} IS NOT NULL
                    ) AS convertible_rows
                FROM "{table_name}"
            """
            total_rows, convertible_rows = self.conn.execute(stats_sql).fetchone()
            logger.info(
                f'[透视表] 值字段 [{value_field}] 可转换性: '
                f'convertible={convertible_rows}, total={total_rows}'
            )

            sample_sql = (
                f'SELECT "{value_field}" '
                f'FROM "{table_name}" '
                f'WHERE "{value_field}" IS NOT NULL '
                f'LIMIT 5'
            )
            sample_result = self.conn.execute(sample_sql).fetchall()
            logger.info(f'[透视表] 值字段样本数据: {sample_result}')

            if not total_rows or not convertible_rows:
                raise ValueError(
                    f'值字段 [{value_field}] 无法转换为数值，'
                    f'请检查源数据格式（例如货币符号、空值或非数字文本）'
                )

            convert_ratio = convertible_rows / max(total_rows, 1)
            if convert_ratio < 0.05:
                raise ValueError(
                    f'值字段 [{value_field}] 数值可转换率过低({convert_ratio:.1%})，'
                    f'无法生成有效透视表'
                )

            # 统一走稳健聚合路径，兼容“数字文本/货币格式/纯数字”
            sql = f"""
                SELECT 
                    "{row_field}",
                    "{column_field}",
                    {agg_func}({value_expr}) AS value
                FROM "{table_name}"
                WHERE "{row_field}" IS NOT NULL
                  AND "{column_field}" IS NOT NULL
                  AND {value_expr} IS NOT NULL
                GROUP BY "{row_field}", "{column_field}"
                ORDER BY "{row_field}", "{column_field}"
            """
            logger.info(f'[透视表] 执行聚合 SQL（稳健数值转换模式）')
            agg_result = self.conn.execute(sql).fetchdf()
            logger.info(f'[透视表] 聚合结果样本:\n{agg_result.head()}')

            pivot_df = agg_result.pivot(
                index=row_field,
                columns=column_field,
                values='value'
            ).fillna(0).reset_index()

            columns = list(pivot_df.columns)
            data = pivot_df.values.tolist()
            
            # 转换为 JSON 兼容格式
            data = self._to_json_compatible(data)
            
            logger.info(f'[透视表] 完成: {len(data)} 行 x {len(columns)} 列')
            return {
                'columns': columns,
                'data': data,
                'row_field': row_field,
                'column_field': column_field,
                'value_field': value_field,
                'agg_func': agg_func,
                'row_count': len(data),
                'col_count': len(columns)
            }
        except Exception as e:
            logger.error(f'创建透视表失败: {e}')
            raise
    
    def group_by(
        self,
        file_id: str,
        group_columns: List[str],
        agg_expressions: Dict[str, str],
        sheet_name: Optional[str] = None,
        having: Optional[str] = None,
        order_by: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        执行分组聚合
        
        Args:
            file_id: 文件ID
            group_columns: 分组列
            agg_expressions: 聚合表达式 {别名: 表达式}，如 {'total': 'SUM(销售额)'}
            sheet_name: 工作表名称
            having: HAVING 条件
            order_by: 排序字段
            
        Returns:
            聚合结果列表
        """
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        # 构建 SELECT 子句
        select_parts = [f'"{col}"' for col in group_columns]
        for alias, expr in agg_expressions.items():
            select_parts.append(f'{expr} AS "{alias}"')
        
        # 构建 GROUP BY 子句
        group_by_clause = ', '.join(f'"{col}"' for col in group_columns)
        
        sql = f"""
            SELECT {', '.join(select_parts)}
            FROM "{table_name}"
            GROUP BY {group_by_clause}
        """
        
        if having:
            sql += f" HAVING {having}"
        if order_by:
            sql += f" ORDER BY {order_by}"
        
        logger.debug(f'分组聚合: group_by={group_columns}, agg={list(agg_expressions.keys())}')
        result = self.conn.execute(sql).fetchdf()
        records = result.to_dict('records')
        logger.debug(f'聚合结果: {len(records)} 行')
        return records
    
    # ==========================================================================
    # 数据修改操作
    # ==========================================================================
    def update_data(
        self,
        file_id: str,
        set_clause: str,
        where_clause: Optional[str] = None,
        sheet_name: Optional[str] = None
    ) -> int:
        """
        更新数据
        
        Args:
            file_id: 文件ID
            set_clause: SET 子句，如 "列名 = 值"
            where_clause: WHERE 条件
            sheet_name: 工作表名称
            
        Returns:
            受影响的行数
        """
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        sql = f'UPDATE "{table_name}" SET {set_clause}'
        if where_clause:
            sql += f' WHERE {where_clause}'
        
        result = self.conn.execute(sql)
        return result.fetchone()[0] if result else 0
    
    def insert_data(
        self,
        file_id: str,
        columns: List[str],
        values: List[List[Any]],
        sheet_name: Optional[str] = None
    ) -> int:
        """
        插入数据
        
        Args:
            file_id: 文件ID
            columns: 列名列表
            values: 值列表（二维数组）
            sheet_name: 工作表名称
            
        Returns:
            插入的行数
        """
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        cols_str = ', '.join(f'"{col}"' for col in columns)
        placeholders = ', '.join(['?' for _ in columns])
        
        sql = f'INSERT INTO "{table_name}" ({cols_str}) VALUES ({placeholders})'
        
        count = 0
        for row in values:
            self.conn.execute(sql, row)
            count += 1
        
        return count
    
    def delete_data(
        self,
        file_id: str,
        where_clause: str,
        sheet_name: Optional[str] = None
    ) -> int:
        """删除数据"""
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        sql = f'DELETE FROM "{table_name}" WHERE {where_clause}'
        result = self.conn.execute(sql)
        return result.fetchone()[0] if result else 0
    
    # ==========================================================================
    # 导出操作
    # ==========================================================================
    def export_to_xlsx(
        self,
        file_id: str,
        output_path: str,
        sheet_name: Optional[str] = None,
        query: Optional[str] = None
    ) -> str:
        """
        导出数据到 Excel 文件
        
        Args:
            file_id: 文件ID
            output_path: 输出文件路径
            sheet_name: 工作表名称
            query: 自定义查询（可选，用于导出查询结果）
            
        Returns:
            输出文件路径
        """
        table_name = self._get_cached_table_name(file_id, sheet_name)
        if not table_name:
            raise ValueError(f'文件未加载: file_id={file_id}')
        
        if query:
            actual_query = query.replace('{table}', f'"{table_name}"')
        else:
            actual_query = f'SELECT * FROM "{table_name}"'
        
        # 使用 COPY 导出
        self.conn.execute(f"""
            COPY ({actual_query}) 
            TO '{output_path}' 
            WITH (FORMAT xlsx, HEADER true)
        """)
        
        logger.info(f'导出完成: {output_path}')
        return output_path
    
    def export_query_to_xlsx(
        self,
        sql: str,
        output_path: str,
        sheet_name: str = 'Sheet1'
    ) -> str:
        """导出 SQL 查询结果到 Excel"""
        self.conn.execute(f"""
            COPY ({sql}) 
            TO '{output_path}' 
            WITH (FORMAT xlsx, HEADER true, sheet='{sheet_name}')
        """)
        
        logger.info(f'查询结果导出完成: {output_path}')
        return output_path
    
    def export_query_to_xlsx_with_placeholder(
        self,
        file_id: str,
        sql: str,
        output_path: str,
        sheet_name: str,
        source_sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        将查询结果导出到 Excel 文件（支持追加到已存在的文件）
        
        如果文件只有占位工作表，直接用 DuckDB COPY 导出覆盖
        如果文件已有数据，导出到临时文件后追加工作表
        
        Args:
            file_id: 文件ID
            sql: SQL 查询语句（包含 {table} 占位符）
            output_path: 输出文件路径
            sheet_name: 工作表名称
            source_sheet: 源工作表名称
            
        Returns:
            {success: bool, row_count: int, col_count: int, columns: List[str]}
        """
        import tempfile
        import shutil
        from openpyxl import load_workbook
        
        # 解析 SQL 占位符
        actual_sql = self._resolve_table_placeholders(file_id, sql, source_sheet)
        
        # 检查输出文件是否存在，以及是否只有占位工作表
        file_exists = os.path.exists(output_path)
        has_only_placeholder = False
        
        if file_exists:
            try:
                wb = load_workbook(output_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                # 如果只有"结果占位"工作表，可以覆盖
                has_only_placeholder = len(sheet_names) == 1 and sheet_names[0] == "结果占位"
            except Exception as e:
                logger.warning(f'检查文件状态失败: {e}，将使用覆盖模式')
                has_only_placeholder = True
        
        if not file_exists or has_only_placeholder:
            # 直接使用 DuckDB COPY 导出（覆盖或创建新文件）
            logger.info(f'[DuckDB] 直接导出到文件: {output_path}, sheet={sheet_name}')
            try:
                # 先验证 SQL 语句是否正确
                try:
                    logger.debug(f'[DuckDB] 验证 SQL 语句（前500字符）: {actual_sql[:500]}...')
                    preview_df = self.conn.execute(f"{actual_sql} LIMIT 1").fetchdf()
                    columns = list(preview_df.columns)
                    logger.debug(f'[DuckDB] SQL 验证成功: 列数={len(columns)}, 列名={columns[:5]}...')
                except Exception as sql_error:
                    logger.error(f'[DuckDB] SQL 语句验证失败: {sql_error}')
                    logger.error(f'[DuckDB] 完整 SQL 语句:\n{actual_sql}')
                    raise Exception(f'SQL 语句错误: {sql_error}')
                
                # 获取总行数
                try:
                    count_result = self.conn.execute(f"SELECT COUNT(*) FROM ({actual_sql})").fetchone()
                    row_count = count_result[0] if count_result else 0
                except Exception as count_error:
                    logger.warning(f'[DuckDB] 获取行数失败: {count_error}，使用估算值')
                    row_count = 0  # 如果无法获取行数，使用0
                
                # 转义路径中的特殊字符
                # DuckDB 的 COPY 命令需要单引号包裹路径
                # 注意：路径中的反斜杠在字符串中需要转义，但 DuckDB 可能不支持 Windows 路径格式
                escaped_path = output_path.replace("'", "''").replace("\\", "/")  # 转义单引号，统一使用正斜杠
                escaped_sheet = sheet_name.replace("'", "''")  # 转义工作表名中的单引号
                
                # 尝试使用 DuckDB COPY 直接导出（性能最优）
                # 注意：某些复杂 SQL（如 JOIN）可能导致 COPY 命令解析失败
                # 如果失败，会回退到 pandas + openpyxl 方式
                try:
                    # 清理 SQL 语句：移除多余的空白字符和换行符
                    import re
                    cleaned_sql = re.sub(r'\s+', ' ', actual_sql.strip())  # 将多个空白字符替换为单个空格
                    
                    # 使用 DuckDB COPY 直接导出（性能最优）
                    # 注意：COPY 命令中的 SQL 需要用括号包裹
                    copy_sql = f"COPY ({cleaned_sql}) TO '{escaped_path}' WITH (FORMAT xlsx, HEADER true, sheet='{escaped_sheet}')"
                    
                    logger.info(f'[DuckDB] 执行 COPY 命令: 路径={escaped_path}, sheet={escaped_sheet}')
                    logger.debug(f'[DuckDB] 原始 SQL 长度: {len(actual_sql)}, 清理后长度: {len(cleaned_sql)}')
                    logger.debug(f'[DuckDB] COPY SQL 语句（前500字符）: {cleaned_sql[:500]}...')
                    
                    self.conn.execute(copy_sql)
                    logger.info(f'[DuckDB] COPY 命令执行成功')
                    
                except Exception as copy_error:
                    # COPY 命令失败（可能是复杂 SQL 或 DuckDB Excel 扩展限制）
                    logger.warning(f'[DuckDB] COPY 命令失败，使用 pandas + openpyxl 批量写入: {copy_error}')
                    logger.debug(f'[DuckDB] COPY 错误详情: {copy_error}')
                    
                    # 回退方案：使用 pandas + openpyxl 批量写入（仍然很快）
                    import pandas as pd
                    
                    # 执行查询获取数据
                    result_df = self.conn.execute(actual_sql).fetchdf()
                    
                    # 使用 pandas 写入 Excel（批量写入，快5-10倍）
                    with pd.ExcelWriter(escaped_path, engine='openpyxl', mode='w') as writer:
                        result_df.to_excel(writer, sheet_name=escaped_sheet, index=False)
                    
                    logger.info(f'[DuckDB] pandas 批量写入完成: {escaped_path}, rows={len(result_df)}')
                    
                    # 更新返回信息
                    columns = list(result_df.columns)
                    row_count = len(result_df)
                
                method_used = 'copy' if copy_success else 'pandas'
                logger.info(f'[DuckDB] 直接导出完成: {output_path}, {row_count} 行 x {len(columns)} 列, method={method_used}')
                
                return {
                    'success': True,
                    'row_count': row_count,
                    'col_count': len(columns),
                    'columns': columns,
                    'method': method_used  # 记录使用的方法
                }
            except Exception as e:
                logger.error(f'[DuckDB] 直接导出失败: {e}，将回退到 openpyxl 方式')
                import traceback
                logger.debug(f'[DuckDB] 错误详情:\n{traceback.format_exc()}')
                raise
        else:
            # 文件已存在且有数据，需要追加工作表
            # 先导出到临时文件，然后读取并追加到目标文件
            logger.info(f'[DuckDB] 文件已存在，导出到临时文件后追加工作表')
            
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_path = tmp_file.name
            
            try:
                # 转义路径和工作表名中的特殊字符
                escaped_tmp_path = tmp_path.replace("'", "''").replace("\\", "/")  # 统一使用正斜杠
                escaped_sheet = sheet_name.replace("'", "''")
                
                # 清理 SQL 语句中的多余空白字符
                cleaned_sql = ' '.join(actual_sql.split())
                
                # 尝试使用 COPY 命令导出到临时文件
                copy_sql = f"COPY ({cleaned_sql}) TO '{escaped_tmp_path}' WITH (FORMAT xlsx, HEADER true, sheet='{escaped_sheet}')"
                logger.info(f'[DuckDB] 执行 COPY 命令（临时文件）: 路径={escaped_tmp_path}, sheet={escaped_sheet}')
                logger.debug(f'[DuckDB] COPY SQL 语句（前500字符）: {cleaned_sql[:500]}...')
                
                try:
                    self.conn.execute(copy_sql)
                    copy_success = True
                except Exception as copy_error:
                    # COPY 失败，使用 pandas 回退方案
                    logger.warning(f'[DuckDB] COPY 命令失败（临时文件），使用 pandas: {copy_error}')
                    import pandas as pd
                    result_df = self.conn.execute(actual_sql).fetchdf()
                    with pd.ExcelWriter(escaped_tmp_path, engine='openpyxl', mode='w') as writer:
                        result_df.to_excel(writer, sheet_name=escaped_sheet, index=False)
                    copy_success = False
                
                # 读取临时文件的工作表并追加到目标文件
                tmp_wb = load_workbook(tmp_path, read_only=True)
                target_wb = load_workbook(output_path, read_only=False)
                
                # 获取临时文件的工作表数据
                tmp_ws = tmp_wb[sheet_name]
                
                # 创建新工作表（确保名称唯一）
                existing_names = list(target_wb.sheetnames)
                final_sheet_name = sheet_name
                if final_sheet_name in existing_names:
                    # 如果名称已存在，添加序号
                    index = 1
                    while f"{sheet_name}_{index:02d}" in existing_names:
                        index += 1
                    final_sheet_name = f"{sheet_name}_{index:02d}"
                target_ws = target_wb.create_sheet(final_sheet_name)
                
                # 复制数据（批量复制以提高性能）
                for row in tmp_ws.iter_rows(values_only=True):
                    target_ws.append(row)
                
                # 如果只有占位工作表，移除它
                if len(existing_names) == 1 and existing_names[0] == "结果占位":
                    target_wb.remove(target_wb[existing_names[0]])
                
                target_wb.save(output_path)
                target_wb.close()
                tmp_wb.close()
                
                # 获取行数和列数
                row_count = tmp_ws.max_row - 1  # 减去表头
                col_count = tmp_ws.max_column
                columns = [tmp_ws.cell(row=1, column=c).value for c in range(1, col_count + 1)]
                
                logger.info(f'[DuckDB] 追加工作表完成: {final_sheet_name}, {row_count} 行 x {col_count} 列')
                
                return {
                    'success': True,
                    'row_count': row_count,
                    'col_count': col_count,
                    'columns': [str(c) if c else f'列{i+1}' for i, c in enumerate(columns)],
                    'sheet_name': final_sheet_name
                }
            finally:
                # 清理临时文件
                try:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
                except Exception as e:
                    logger.warning(f'清理临时文件失败: {e}')
    
    # ==========================================================================
    # 工具方法
    # ==========================================================================
    def get_sheet_names(self, file_path: str) -> List[str]:
        """获取 Excel 文件的所有工作表名称（不含系统内置 __SHEETBOT_META__）"""
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        sheets = [s for s in (wb.sheetnames or []) if s and s != '__SHEETBOT_META__']
        wb.close()
        return sheets
    
    def unload_file(self, file_id: str):
        """卸载文件，释放内存"""
        with self._cache_lock:
            if file_id in self._loaded_files:
                file_info = self._loaded_files[file_id]
                for sheet_info in file_info.get('sheets', {}).values():
                    table_name = sheet_info.get('table_name')
                    if table_name:
                        try:
                            self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                        except:
                            pass
                del self._loaded_files[file_id]
                logger.info(f'文件已卸载: file_id={file_id}')
    
    def clear_cache(self):
        """清空所有缓存"""
        with self._cache_lock:
            file_ids = list(self._loaded_files.keys())
        for file_id in file_ids:
            self.unload_file(file_id)
        logger.info('缓存已清空')
    
    def _get_table_name(self, file_id: str, sheet_name: Optional[str] = None) -> str:
        """生成表名"""
        if sheet_name:
            return f"excel_{file_id}_{sheet_name}".replace('-', '_')
        return f"excel_{file_id}".replace('-', '_')
    
    def _get_cached_table_name(self, file_id: str, sheet_name: Optional[str] = None) -> Optional[str]:
        """获取已缓存的表名"""
        with self._cache_lock:
            if file_id not in self._loaded_files:
                return None
            file_info = self._loaded_files[file_id]
            sheet_key = sheet_name or file_info.get('default_sheet') or 'default'
            return file_info.get('sheets', {}).get(sheet_key, {}).get('table_name')
    
    def _get_table_info(self, table_name: str) -> Dict[str, Any]:
        """获取表信息"""
        # 获取行数
        row_count = self.execute_fetchone(f'SELECT COUNT(*) FROM "{table_name}"')[0]
        
        # 获取列信息
        columns = self._get_columns(table_name)
        
        return {
            'row_count': row_count,
            'col_count': len(columns),
            'columns': columns
        }
    
    def _get_columns(self, table_name: str) -> List[str]:
        """获取表的列名"""
        result = self.execute_fetchdf(f"DESCRIBE \"{table_name}\"")
        return result['column_name'].tolist()
    
    def is_loaded(self, file_id: str, sheet_name: Optional[str] = None) -> bool:
        """检查文件是否已加载"""
        return self._get_cached_table_name(file_id, sheet_name) is not None

    def rename_sheet(self, file_id: str, old_name: str, new_name: str) -> bool:
        """
        重命名已加载的工作表：ALTER TABLE RENAME + 缓存 key 迁移。
        如果该 sheet 未加载到 DuckDB，静默返回 False（不报错）。
        """
        with self._cache_lock:
            if file_id not in self._loaded_files:
                return False
            file_info = self._loaded_files[file_id]
            sheets = file_info.get('sheets', {})
            if old_name not in sheets:
                return False

            old_table = sheets[old_name].get('table_name')
            new_table = self._get_table_name(file_id, new_name)

        # ALTER TABLE 在锁外执行，避免长持锁
        if old_table:
            try:
                self.conn.execute(f'ALTER TABLE "{old_table}" RENAME TO "{new_table}"')
                logger.info(f'DuckDB 表已重命名: {old_table} -> {new_table}')
            except Exception as e:
                logger.error(f'DuckDB 重命名失败: {old_table} -> {new_table}, error={e}')
                return False

        # 迁移缓存 key
        with self._cache_lock:
            sheets = self._loaded_files[file_id].get('sheets', {})
            if old_name in sheets:
                info = sheets.pop(old_name)
                info['table_name'] = new_table
                sheets[new_name] = info
            if self._loaded_files[file_id].get('default_sheet') == old_name:
                self._loaded_files[file_id]['default_sheet'] = new_name

        logger.info(f'工作表重命名完成: file_id={file_id}, {old_name} -> {new_name}')
        return True

    # ==========================================================================
    # 生命周期管理
    # ==========================================================================
    def unload_sheet(self, file_id: str, sheet_name: str) -> bool:
        """
        卸载单个工作表，释放对应的 DuckDB 内存表
        
        Args:
            file_id: 文件ID
            sheet_name: 工作表名称
            
        Returns:
            是否成功卸载
        """
        with self._cache_lock:
            if file_id not in self._loaded_files:
                logger.debug(f'文件未加载，无需卸载: file_id={file_id}')
                return False
            
            file_info = self._loaded_files[file_id]
            sheets = file_info.get('sheets', {})
            
            if sheet_name not in sheets:
                logger.debug(f'工作表未加载，无需卸载: file_id={file_id}, sheet={sheet_name}')
                return False
            
            # 获取并删除 DuckDB 表
            sheet_info = sheets[sheet_name]
            table_name = sheet_info.get('table_name')
            if table_name:
                try:
                    self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                    logger.info(f'DuckDB 表已删除: {table_name}')
                except Exception as e:
                    logger.warning(f'删除 DuckDB 表失败: {table_name}, error={e}')
            
            # 从缓存中移除
            del sheets[sheet_name]
            
            # 如果该文件没有其他工作表，清理整个文件条目
            if not sheets:
                del self._loaded_files[file_id]
                logger.info(f'文件所有工作表已卸载，移除文件缓存: file_id={file_id}')
            
            logger.info(f'工作表已卸载: file_id={file_id}, sheet={sheet_name}')
            return True
    
    def clear_session_cache(self, file_ids: List[str]) -> Dict[str, Any]:
        """
        清空指定会话的所有 DuckDB 内存表
        
        用于导出结果文件后释放内存
        
        Args:
            file_ids: 要清理的文件ID列表
            
        Returns:
            {cleared_files: int, cleared_tables: int}
        """
        cleared_files = 0
        cleared_tables = 0
        
        for file_id in file_ids:
            with self._cache_lock:
                if file_id not in self._loaded_files:
                    continue
                
                file_info = self._loaded_files[file_id]
                sheets = file_info.get('sheets', {})
                
                # 删除所有工作表的 DuckDB 表
                for sheet_name, sheet_info in sheets.items():
                    table_name = sheet_info.get('table_name')
                    if table_name:
                        try:
                            self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                            cleared_tables += 1
                            logger.debug(f'会话清理 - 删除表: {table_name}')
                        except Exception as e:
                            logger.warning(f'会话清理 - 删除表失败: {table_name}, error={e}')
                
                # 移除文件缓存
                del self._loaded_files[file_id]
                cleared_files += 1
        
        logger.info(f'会话缓存已清理: 清理了 {cleared_files} 个文件, {cleared_tables} 个表')
        return {
            'cleared_files': cleared_files,
            'cleared_tables': cleared_tables
        }
    
    def register_result_table(
        self,
        source_file_id: str,
        result_sheet_name: str,
        columns: List[str],
        data: List[List[Any]]
    ) -> Dict[str, Any]:
        """
        注册结果表到 DuckDB，支持二次加工
        
        将查询/透视表等操作的结果注册为 DuckDB 内存表，
        使其可以被后续的 {table:结果工作表名} 语法引用
        
        Args:
            source_file_id: 源文件ID（用于关联）
            result_sheet_name: 结果工作表名称
            columns: 列名列表
            data: 数据行（二维数组）
            
        Returns:
            {table_name, row_count, col_count}
        """
        import pandas as pd
        import pyarrow as pa
        
        # 生成结果表名（使用特殊前缀区分）
        table_name = f"result_{source_file_id}_{result_sheet_name}".replace('-', '_').replace(' ', '_')
        
        try:
            # 创建 DataFrame
            df = pd.DataFrame(data, columns=columns)
            
            # 转换为 pyarrow Table 并导入 DuckDB
            arrow_table = pa.Table.from_pandas(df, preserve_index=False)
            
            self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            self.conn.register('_result_arrow_table', arrow_table)
            self.conn.execute(f'CREATE TABLE "{table_name}" AS SELECT * FROM _result_arrow_table')
            self.conn.unregister('_result_arrow_table')
            
            # 注册到缓存（使用特殊的 result 前缀）
            result_file_key = f"result_{source_file_id}"
            with self._cache_lock:
                if result_file_key not in self._loaded_files:
                    self._loaded_files[result_file_key] = {
                        'file_path': None,  # 结果表无文件路径
                        'sheets': {},
                        'default_sheet': result_sheet_name,
                        'is_result': True  # 标记为结果表
                    }
                self._loaded_files[result_file_key]['sheets'][result_sheet_name] = {
                    'table_name': table_name,
                    'row_count': len(data),
                    'col_count': len(columns),
                    'columns': columns
                }
            
            logger.info(f'结果表已注册: table={table_name}, rows={len(data)}, cols={len(columns)}')
            return {
                'table_name': table_name,
                'row_count': len(data),
                'col_count': len(columns)
            }
        except Exception as e:
            logger.error(f'注册结果表失败: {e}')
            raise
    
    def get_result_table_name(self, source_file_id: str, result_sheet_name: str) -> Optional[str]:
        """
        获取结果表的 DuckDB 表名
        
        Args:
            source_file_id: 源文件ID
            result_sheet_name: 结果工作表名称
            
        Returns:
            DuckDB 表名，如果不存在则返回 None
        """
        result_file_key = f"result_{source_file_id}"
        with self._cache_lock:
            if result_file_key not in self._loaded_files:
                return None
            return self._loaded_files[result_file_key].get('sheets', {}).get(result_sheet_name, {}).get('table_name')
    
    def list_available_tables(self, file_id: str) -> List[Dict[str, Any]]:
        """
        列出文件相关的所有可用表（包括源数据表和结果表）
        
        用于 AI Agent 了解可操作的数据源
        
        Args:
            file_id: 文件ID
            
        Returns:
            [{name, type, row_count, columns}]
        """
        tables = []
        result_file_key = f"result_{file_id}"
        
        with self._cache_lock:
            # 源数据表
            if file_id in self._loaded_files:
                file_info = self._loaded_files[file_id]
                for sheet_name, sheet_info in file_info.get('sheets', {}).items():
                    tables.append({
                        'name': sheet_name,
                        'table_name': sheet_info.get('table_name'),
                        'type': 'source',
                        'row_count': sheet_info.get('row_count', 0),
                        'columns': sheet_info.get('columns', []),
                        'syntax': f'{{table:{sheet_name}}}'
                    })
            
            # 结果表
            if result_file_key in self._loaded_files:
                result_info = self._loaded_files[result_file_key]
                for sheet_name, sheet_info in result_info.get('sheets', {}).items():
                    tables.append({
                        'name': sheet_name,
                        'table_name': sheet_info.get('table_name'),
                        'type': 'result',
                        'row_count': sheet_info.get('row_count', 0),
                        'columns': sheet_info.get('columns', []),
                        'syntax': f'{{table:{sheet_name}}}'
                    })
        
        return tables


# ==============================================================================
# 全局实例
# ==============================================================================
duckdb_manager = DuckDBExcelManager()
