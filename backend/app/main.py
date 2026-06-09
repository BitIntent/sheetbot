# backend/app/main.py
"""
FastAPI Main Application
Excel AI Assistant Backend Entry Point
"""
import os
import asyncio
import uuid
from datetime import datetime
from contextlib import asynccontextmanager
from typing import Optional
from fastapi import FastAPI, WebSocket, UploadFile, File, HTTPException, Request, Query, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

# 核心导入
from .core.database import init_db, close_db, get_db, async_session_maker
from .core.config import settings
from .core.dependencies import get_current_user, get_optional_user, get_optional_user_info
from .core.quota import QuotaGuard, get_user_quota, get_user_plan_name
from .core.usage_service import increment_usage
from .core.security import verify_token
from .utils.logger import get_logger, bind_log_user_tag
from .utils.access_logger import write_access_log
from .auth.models import User, PasswordResetToken  # noqa: F401 - 确保 ORM 建表
from .files.models import UserFile
from .files.models import UserPptx, BatchWordExport  # noqa: F401 — 确保 ORM 建表
from .files import service as file_service
from .connect.models import Connector, SyncJob  # noqa: F401 - 确保 ORM 建表
from .config.models import UserPreferences, PlatformSetting  # noqa: F401 - 确保 ORM 建表
from .formula.models import CustomFormula  # noqa: F401 - 确保 ORM 建表
from .skill.models import Skill  # noqa: F401 - 确保 ORM 建表
from .business_inquiry_models import BusinessInquiry  # noqa: F401 - 确保 ORM 建表
from .plans.models import SubscriptionPlan, UserSubscription, UsageRecord, SystemAnnouncement  # noqa: F401

# 路由导入
from .auth.router import router as auth_router
from .excel.router import router as excel_router
from .analyze.router import router as analyze_router
from .files.router import router as files_router, folder_router
from .report.router import router as report_router, public_router as report_public_router
from .notification.router import router as notification_router
from .pptx.router import router as pptx_router
from .collect.router import router as collect_router, public_router as collect_public_router
from .connect.router import router as connect_router, webhook_router as connect_webhook_router
from .config.router import router as config_router
from .formula.router import router as formula_router
from .skill.router import router as skill_router
from .batch_word.router import router as batch_word_router
from .business_inquiry_router import router as business_inquiry_router
from .plans.router import router as plans_router
from .connect.scheduler import sync_scheduler
from .report.task_manager import report_task_manager

# WebSocket/SSE 处理器
from .websocket_handler import websocket_endpoint, connection_manager
from .sse_handler import sse_endpoint, sse_connection_manager
from .agent.excel_agent import agent_manager
from .large_file.large_file_agent import large_file_agent_manager
from .large_file.storage import large_file_storage
from .large_file.schemas import LARGE_FILE_THRESHOLD_BYTES, PREVIEW_ROW_COUNT, FileStatus

# 应用日志
logger = get_logger('app')
large_file_log = get_logger('large_file.api')
app_log = logger  # 兼容旧代码


def _resolve_client_ip(request: Request) -> str:
    """优先从代理头中提取真实 IP。"""
    forwarded_for = request.headers.get("x-forwarded-for")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    real_ip = request.headers.get("x-real-ip")
    if real_ip:
        return real_ip.strip()
    if request.client and request.client.host:
        return request.client.host
    return "-"


def _resolve_session_id(request: Request) -> str:
    """尽量提取 session_id，缺失时返回空字符串。"""
    path_params = request.path_params or {}
    if path_params.get("session_id"):
        return str(path_params["session_id"])
    if request.query_params.get("session_id"):
        return str(request.query_params.get("session_id"))
    if request.headers.get("x-session-id"):
        return request.headers.get("x-session-id", "")
    return ""


def _is_large_file_info_only_command(command: str) -> bool:
    """
    判定是否为“仅查看信息”请求（不要求创建新结果工作表）。
    只要包含分析语义关键词，就不算 info-only。
    """
    text_raw = str(command or "").strip()
    if not text_raw:
        return False
    text = text_raw.lower()

    analysis_keywords = [
        "分析", "统计", "汇总", "分组", "透视", "筛选", "过滤", "排序", "去重",
        "求和", "平均", "最大", "最小", "总计", "导出", "生成结果", "结果表", "新工作表",
        "sql", "select", "group by", "order by", "where", "having", "sum(", "avg(", "count(", "max(", "min(",
    ]
    if any(k in text for k in analysis_keywords):
        return False

    info_only_keywords = [
        "文件信息", "工作表信息", "有哪些工作表", "列标题", "列名", "字段", "预览", "查看前",
        "show sheets", "sheet list", "columns", "headers", "preview", "file info",
    ]
    if any(k in text for k in info_only_keywords):
        return True

    # 保守策略：未知请求默认按“需要导出结果表”处理，避免漏约束。
    return False


def _should_force_large_file_export(command: str) -> bool:
    """大文件模式是否强制要求本轮生成新结果工作表。"""
    return not _is_large_file_info_only_command(command)


LARGE_FILE_INTENT_ANALYSIS = "analysis"
LARGE_FILE_INTENT_INFO = "info"
LARGE_FILE_INTENT_PLANNING = "planning"
LARGE_FILE_ALLOWED_INTENTS = {
    LARGE_FILE_INTENT_ANALYSIS,
    LARGE_FILE_INTENT_INFO,
    LARGE_FILE_INTENT_PLANNING,
}


def _resolve_large_file_intent_type(payload: dict, command: str) -> str:
    """
    解析大文件请求意图：
    - 新客户端：显式传入 intent_type（analysis/info/planning）
    - 旧客户端：回退到历史规则推断
    """
    raw = payload.get("intent_type")
    intent = str(raw or "").strip().lower()
    if intent in LARGE_FILE_ALLOWED_INTENTS:
        return intent
    return LARGE_FILE_INTENT_ANALYSIS if _should_force_large_file_export(command) else LARGE_FILE_INTENT_INFO


async def _resolve_user_identity(request: Request) -> tuple[str, str]:
    """
    从 Authorization 中解析 user_id / username。
    解析失败时返回空字符串，不影响主流程。
    """
    auth_header = request.headers.get("authorization", "")
    if not auth_header.startswith("Bearer "):
        return "", ""
    token = auth_header[7:].strip()
    if not token:
        return "", ""

    payload = verify_token(token, token_type="access")
    if not payload:
        return "", ""

    user_id = str(payload.get("sub") or "")
    if not user_id:
        return "", ""

    username = ""
    try:
        async with async_session_maker() as db:
            result = await db.execute(select(User).where(User.id == user_id))
            user = result.scalar_one_or_none()
            if user:
                username = user.username or ""
    except Exception:
        username = ""

    return user_id, username

# 验证必要的环境变量
if not settings.ANTHROPIC_CREDENTIAL:
    logger.warning("ANTHROPIC_API_KEY / ANTHROPIC_AUTH_TOKEN 环境变量未设置")


@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期管理"""
    logger.info(f"Excel AI Assistant 后端正在启动... (v{settings.APP_VERSION})")
    
    # 初始化数据库
    try:
        await init_db()
        logger.info("数据库初始化成功")
    except Exception as e:
        logger.error(f"数据库初始化失败: {e}")

    try:
        from .config.platform_runtime import seed_platform_defaults
        async with async_session_maker() as db:
            await seed_platform_defaults(db)
            await db.commit()
        logger.info("platform_settings 默认项已校验")
    except Exception as e:
        logger.warning(f"platform_settings 种子写入跳过: {e}")
    
    # 启动 Agent 管理器
    await large_file_agent_manager.start()
    
    # 启动异步报表任务管理器
    await report_task_manager.startup()
    logger.info("异步报表任务管理器已启动")
    
    # 启动连接器同步调度器
    await sync_scheduler.start()

    # ---- 默认套餐种子数据（幂等） ----
    try:
        from .plans.seed import seed_default_plans
        async with async_session_maker() as db:
            await seed_default_plans(db)
        logger.info("默认套餐种子数据已就绪")
    except Exception as e:
        logger.warning(f"默认套餐种子写入失败（不影响主系统运行）: {e}")

    # ---- 存量用户补写免费套餐订阅（一次性幂等修复） ----
    try:
        async with async_session_maker() as db:
            await db.execute(
                text("""
                    INSERT INTO user_subscriptions (id, user_id, plan_code, status, started_at, created_at, updated_at)
                    SELECT UUID(), u.id, 'free', 'active', NOW(), NOW(), NOW()
                    FROM users u
                    WHERE u.is_active = 1
                      AND NOT EXISTS (
                          SELECT 1 FROM user_subscriptions us
                          WHERE us.user_id = u.id AND us.status = 'active'
                      )
                """)
            )
            await db.commit()
        logger.info("存量用户免费套餐订阅补写完成")
    except Exception as e:
        logger.warning(f"存量用户套餐补写失败（不影响主系统运行）: {e}")

    yield
    
    # 清理资源
    logger.info("正在关闭...")
    await sync_scheduler.stop()
    await agent_manager.close_all()
    await large_file_agent_manager.stop()
    await close_db()


# 创建 FastAPI 应用
app = FastAPI(
    title=settings.APP_NAME,
    description="使用 Claude Agent SDK 的 AI 驱动 Excel 操作助手 - 支持用户认证与数据分析",
    version=settings.APP_VERSION,
    lifespan=lifespan
)

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:3000",
        "http://127.0.0.1:5173",
        "http://localhost:80",
        "http://127.0.0.1:80",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)


@app.middleware("http")
async def access_log_middleware(request: Request, call_next):
    """统一 access log。"""
    request_id = request.headers.get("x-request-id") or f"req_{uuid.uuid4().hex[:16]}"
    start_time = datetime.now().astimezone()
    user_id = ""
    username = ""

    response = None
    status_code = 500
    try:
        user_id, username = await _resolve_user_identity(request)
        bind_log_user_tag(username or user_id)
        response = await call_next(request)
        status_code = response.status_code
        response.headers["X-Request-ID"] = request_id
        return response
    finally:
        end_time = datetime.now().astimezone()
        duration_ms = int((end_time - start_time).total_seconds() * 1000)
        route = request.scope.get("route")
        route_path = getattr(route, "path", request.url.path)

        access_payload = {
            "ts": end_time.isoformat(timespec="milliseconds"),
            "level": "INFO",
            "type": "access",
            "request_id": request_id,
            "session_id": _resolve_session_id(request),
            "user_id": user_id,
            "username": username,
            "client_ip": _resolve_client_ip(request),
            "method": request.method,
            "path": request.url.path,
            "route": route_path,
            "query": request.url.query or "",
            "status": status_code,
            "duration_ms": duration_ms,
            "user_agent": request.headers.get("user-agent", ""),
            "referer": request.headers.get("referer", ""),
            "app": "excel-ai-backend",
            "env": os.getenv("APP_ENV", "prod"),
        }
        write_access_log(access_payload)

# ============================================================
# 注册路由
# ============================================================

# 认证路由（公开访问）
app.include_router(auth_router)

# Excel 编辑路由（可选认证）
app.include_router(excel_router)

# 数据分析路由（可选认证）
app.include_router(analyze_router)

# 文件/文件夹管理路由（需要认证）
app.include_router(files_router)
app.include_router(folder_router)

# 报表路由（需要认证 + 公开分享端点）
app.include_router(report_router)
app.include_router(report_public_router)

# PPTX 汇报路由
app.include_router(pptx_router)

# 表单收集路由（需要认证 + 公开提交端点）
app.include_router(collect_router)
app.include_router(collect_public_router)

# 外部系统连接器路由（需要认证 + Webhook 公开端点）
app.include_router(connect_router)
app.include_router(connect_webhook_router)

# 通知路由（需要认证）
app.include_router(notification_router)

# 系统配置路由（需要认证）
app.include_router(config_router)

# 自定义公式
app.include_router(formula_router)

# 技能库
app.include_router(skill_router)

# 商务咨询公开端点
app.include_router(business_inquiry_router)

# 套餐公开信息（landing 定价区动态刷新）
from .plans.public_router import router as public_plans_router
app.include_router(public_plans_router)

# 套餐只读（当前订阅信息，不含支付）
app.include_router(plans_router)

# 批量转 Word
app.include_router(batch_word_router)

# ============================================================
# REST API Endpoints
# ============================================================

@app.get("/")
async def root():
    """根路径 - 健康检查"""
    return {
        "status": "ok",
        "service": settings.APP_NAME,
        "version": settings.APP_VERSION,
        "database": "connected"
    }


@app.get("/health")
async def health_check():
    """健康检查端点"""
    return {
        "status": "healthy",
        "api_key_configured": bool(settings.ANTHROPIC_CREDENTIAL),
        "database": "ok"
    }


@app.post("/api/client-log")
async def client_log(request: Request):
    """
    接收前端诊断日志（用于排查 workspace -> landing 快速回跳）。
    """
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    event = str(payload.get("event") or "unknown")
    href = str(payload.get("href") or "")
    detail = payload.get("detail") if isinstance(payload.get("detail"), dict) else {}
    client_ip = _resolve_client_ip(request)

    logger.warning(
        "CLIENT_DIAG event=%s ip=%s href=%s detail=%s",
        event,
        client_ip,
        href,
        detail,
    )
    return {"ok": True}


# Demo 文件路径
from pathlib import Path
_demo_file_path = Path(__file__).resolve().parent.parent.parent / settings.UPLOAD_DIR / 'demo' / '某公司产品销售表.xlsx'


@app.get("/api/demo/file")
async def get_demo_file():
    """获取 Demo Excel 文件，供首次登录用户快速体验"""
    if not _demo_file_path.exists():
        raise HTTPException(status_code=404, detail="Demo 文件不存在，请确保 uploads/demo/某公司产品销售表.xlsx 已放置")
    return FileResponse(
        _demo_file_path,
        filename="某公司产品销售表.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# 注意：/api/excel/* 和 /api/analyze/* 端点已移至模块化路由
# 参见 app/excel/router.py 和 app/analyze/router.py


# ============================================================
# WebSocket Endpoint
# ============================================================

@app.websocket("/ws/{session_id}")
async def websocket_route(websocket: WebSocket, session_id: str):
    """WebSocket 连接端点"""
    await websocket_endpoint(websocket, session_id)


@app.get("/sse/{session_id}")
async def sse_route(session_id: str):
    """SSE 连接端点"""
    return await sse_endpoint(session_id)


# ============================================================
# Large File API Endpoints (独立模块，与现有架构隔离)
# ============================================================

_prepared_user_file_map: dict[str, str] = {}
_user_active_prepare: dict[str, str] = {}  # user.id -> user_file_id (tracks active prepare per user)
PREPARE_LOAD_TIMEOUT_SECONDS = int(os.getenv("PREPARE_LOAD_TIMEOUT_SECONDS", "180"))


async def _run_blocking_with_timeout(func, timeout_seconds: int, timeout_message: str):
    """在线程池执行阻塞任务并施加超时熔断。"""
    loop = asyncio.get_running_loop()
    try:
        return await asyncio.wait_for(
            loop.run_in_executor(None, func),
            timeout=timeout_seconds,
        )
    except asyncio.TimeoutError as exc:
        raise TimeoutError(timeout_message) from exc


def _prepare_status_payload(file_id: str):
    metadata = large_file_storage.get_metadata(file_id)
    if not metadata:
        return None

    duckdb_load_time_seconds = None
    if metadata.duckdb_ready and metadata.duckdb_load_started_at and metadata.duckdb_load_finished_at:
        duckdb_load_time_seconds = (metadata.duckdb_load_finished_at - metadata.duckdb_load_started_at).total_seconds()

    return {
        "file_id": metadata.file_id,
        "status": metadata.status.value,
        "original_name": metadata.original_name,
        "file_size": metadata.file_size,
        "sheet_names": metadata.sheet_names,
        "sheet_row_counts": metadata.sheet_row_counts,
        "row_count": metadata.row_count,
        "col_count": metadata.col_count,
        "duckdb_ready": metadata.duckdb_ready,
        "first_page_ready": metadata.first_page_ready,
        "duckdb_load_time_seconds": duckdb_load_time_seconds,
        "duckdb_load_stage": metadata.duckdb_load_stage,
        "duckdb_load_progress": metadata.duckdb_load_progress,
        "result_files": large_file_storage.get_result_files(metadata.file_id),
    }


@app.post("/api/large-file/prepare")
async def prepare_large_file_from_user_file(
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    基于文件管理中已上传的文件，准备 DuckDB 分析上下文。
    target_view='analyze' 时额外检查 large_file_rows 配额。
    """
    payload = await request.json()
    user_file_id = payload.get("user_file_id")
    target_view = payload.get("target_view", "")

    # ---- 大文件分析功能门禁（仅 analyze 视图需要）----
    if target_view == "analyze":
        quotas = await get_user_quota(user.id, db)
        limit = quotas.get("large_file_rows")
        if limit is not None and limit == 0:
            plan_name = await get_user_plan_name(user.id, db)
            raise HTTPException(403, detail={
                "code": "feature_disabled",
                "key": "large_file_rows",
                "feature": "大文件分析",
                "plan_name": plan_name,
                "message": f"您当前为「{plan_name}」，未开通「大文件分析」功能，请升级套餐后使用。",
            })
    if not user_file_id:
        raise HTTPException(status_code=400, detail="user_file_id 不能为空")

    result = await db.execute(
        select(UserFile).where(
            UserFile.id == user_file_id,
            UserFile.user_id == user.id,
            UserFile.status == "active",
        )
    )
    user_file = result.scalar_one_or_none()
    if not user_file:
        raise HTTPException(status_code=404, detail="文件不存在或无权限访问")

    resolved_user_file_path = file_service.resolve_storage_path(user_file.storage_path)
    if not os.path.exists(resolved_user_file_path):
        raise HTTPException(status_code=404, detail="文件内容不存在")

    lower_name = (user_file.file_name or "").lower()
    if not lower_name.endswith((".xlsx", ".xls", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 Excel 文件用于分析准备")

    cached_file_id = _prepared_user_file_map.get(user_file.id)
    if cached_file_id:
        cached_payload = _prepare_status_payload(cached_file_id)
        if cached_payload:
            # 校验缓存上下文是否包含全部源工作表；缺失时补加载，避免“只加载首表”
            try:
                expected_sheets = [
                    s for s in (cached_payload.get("sheet_names") or [])
                    if s and s != "__SHEETBOT_META__"
                ]
                if expected_sheets:
                    from .large_file.large_file_duckdb import duckdb_manager
                    loaded_tables = duckdb_manager.list_available_tables(cached_file_id)
                    loaded_source_sheets = {
                        t.get("name")
                        for t in loaded_tables
                        if t.get("type") == "source" and t.get("name")
                    }
                    missing_sheets = [s for s in expected_sheets if s not in loaded_source_sheets]
                    if missing_sheets:
                        # 若当前文件仍在后台加载中，禁止前台补加载，避免同一 file_id 并发导入导致 DuckDB 连接冲突
                        is_loading = not bool(cached_payload.get("duckdb_ready"))
                        if is_loading:
                            logger.info(
                                f"检测到后台仍在加载，跳过并发补加载: file_id={cached_file_id}, missing={missing_sheets}"
                            )
                            _user_active_prepare[str(user.id)] = user_file_id
                            return cached_payload
                        logger.info(
                            f"缓存上下文缺少工作表，补加载全部: file_id={cached_file_id}, missing={missing_sheets}"
                        )
                        file_path = large_file_storage.get_file_path(cached_file_id)
                        if file_path and os.path.exists(file_path):
                            def _on_cached_sheet(sn: str, idx: int, total: int):
                                pct = 30 + int((idx + 1) / max(1, total) * 70)
                                large_file_storage.update_duckdb_load_stage(
                                    cached_file_id, f"正在加载工作表 ({idx + 1}/{total})...", pct
                                )

                            await _run_blocking_with_timeout(
                                lambda: duckdb_manager.load_all_sheets(
                                    str(file_path), cached_file_id, progress_callback=_on_cached_sheet
                                ),
                                PREPARE_LOAD_TIMEOUT_SECONDS,
                                f"DuckDB 加载超时（>{PREPARE_LOAD_TIMEOUT_SECONDS}秒）",
                            )
                            large_file_storage.set_duckdb_ready(cached_file_id, True)
                            large_file_storage.update_duckdb_load_stage(cached_file_id, "已完成", 100)
                            refreshed_payload = _prepare_status_payload(cached_file_id)
                            if refreshed_payload:
                                cached_payload = refreshed_payload
                    else:
                        # 所有工作表都已加载，无需重复加载
                        logger.info(
                            f"所有工作表已在内存中，跳过重复加载: file_id={cached_file_id}, sheets={expected_sheets}"
                        )
            except Exception as e:
                logger.warning(f"缓存上下文校验失败（忽略）: file_id={cached_file_id}, error={e}")
            _user_active_prepare[str(user.id)] = user_file_id
            return cached_payload
        _prepared_user_file_map.pop(user_file.id, None)

    # ---- 读取文件内容（内存操作，通常很快） ----
    with open(resolved_user_file_path, "rb") as f:
        content = f.read()

    # ---- 预注册占位元数据：立即分配 file_id 供前端轮询 ----
    placeholder = large_file_storage.register_placeholder(
        user_file.file_name, len(content), source_file_id=user_file.id,
    )
    _prepared_user_file_map[user_file.id] = placeholder.file_id
    _user_active_prepare[str(user.id)] = user_file_id

    # ---- 后台全流程：写入文件 → 解析结构 → 加载到内存 ----
    from .large_file.large_file_duckdb import duckdb_manager

    _bg_fid = placeholder.file_id

    async def _background_prepare():
        try:
            # 阶段 1: 写入 + 解析文件结构
            meta = await large_file_storage.save_file(
                content, user_file.file_name,
                source_file_id=user_file.id,
                preload_duckdb=False,
                file_id_override=_bg_fid,
            )

            # 阶段 2: 逐表加载到内存（首表完成即可渲染）
            def _on_sheet_loaded(sheet_name: str, idx: int, total: int):
                pct = 30 + int((idx + 1) / max(1, total) * 70)
                large_file_storage.update_duckdb_load_stage(
                    _bg_fid, f"正在加载工作表 ({idx + 1}/{total})...", pct,
                )
                if idx == 0:
                    large_file_storage.set_first_page_ready(_bg_fid)

            await _run_blocking_with_timeout(
                lambda: duckdb_manager.load_all_sheets(
                    str(meta.file_path), _bg_fid,
                    progress_callback=_on_sheet_loaded,
                ),
                PREPARE_LOAD_TIMEOUT_SECONDS,
                f"DuckDB 加载超时（>{PREPARE_LOAD_TIMEOUT_SECONDS}秒）",
            )
            large_file_storage.set_duckdb_ready(_bg_fid, True)
            large_file_storage.update_duckdb_load_stage(_bg_fid, "已完成", 100)
        except Exception as exc:
            logger.warning(f"后台准备文件失败: file_id={_bg_fid}, error={exc}")
            large_file_storage.set_duckdb_ready(_bg_fid, False)
            large_file_storage.update_status(_bg_fid, FileStatus.ERROR, str(exc))
            large_file_storage.update_duckdb_load_stage(
                _bg_fid,
                f"处理出错: {str(exc)[:80]}",
                0,
            )
            # 异常时清掉映射，确保前端下次 prepare 能走新上下文，而不是卡在坏状态
            _prepared_user_file_map.pop(user_file.id, None)
            if _user_active_prepare.get(str(user.id)) == user_file.id:
                _user_active_prepare.pop(str(user.id), None)

    asyncio.create_task(_background_prepare())

    return _prepare_status_payload(_bg_fid)


@app.get("/api/large-file/prepare-status/{user_file_id}")
async def get_prepare_status(
    user_file_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    查询 prepare 状态（按文件管理 user_file_id 查询）。
    """
    result = await db.execute(
        select(UserFile).where(
            UserFile.id == user_file_id,
            UserFile.user_id == user.id,
            UserFile.status == "active",
        )
    )
    user_file = result.scalar_one_or_none()
    if not user_file:
        raise HTTPException(status_code=404, detail="文件不存在或无权限访问")

    file_id = _prepared_user_file_map.get(user_file_id)
    if not file_id:
        raise HTTPException(status_code=404, detail="文件尚未准备，请先调用 prepare")

    payload = _prepare_status_payload(file_id)
    if not payload:
        _prepared_user_file_map.pop(user_file_id, None)
        raise HTTPException(status_code=404, detail="准备上下文不存在，请重新 prepare")

    return payload


@app.post("/api/large-file/upload")
async def upload_large_file(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _quota=Depends(QuotaGuard("storage_mb")),
):
    """
    上传大型Excel文件

    文件将保存到 uploads/YYYY-MM-DD/ 目录
    返回文件ID和预览数据
    """
    large_file_log.info(f"收到文件上传请求: {file.filename}")

    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        large_file_log.warning(f"文件类型不支持: {file.filename}")
        raise HTTPException(status_code=400, detail="只支持 Excel 文件格式（.xlsx, .xls, .xlsm）")

    try:
        large_file_log.debug(f"开始读取文件内容: {file.filename}")
        content = await file.read()
        file_size = len(content)
        file_size_mb = file_size / 1024 / 1024

        # ---- 单文件大小限制 ----
        quotas = await get_user_quota(user.id, db)
        plan_name = await get_user_plan_name(user.id, db)
        limit_mb = quotas.get("file_size_mb")
        if limit_mb is not None and limit_mb != -1 and file_size_mb > limit_mb:
            raise HTTPException(413, detail={
                "code": "quota_exceeded",
                "key": "file_size_mb",
                "feature": "单文件大小",
                "plan_name": plan_name,
                "limit": limit_mb,
                "current": round(file_size_mb, 1),
                "unit": "MB",
                "message": (
                    f"您当前为「{plan_name}」，"
                    f"单文件大小上限为 {limit_mb}MB，"
                    f"当前文件 {file_size_mb:.1f}MB 超出限制。"
                    f"请升级套餐以上传更大文件。"
                ),
            })

        large_file_log.info(f"文件读取完成: {file.filename}, 大小: {file_size_mb:.2f} MB ({file_size:,} 字节)")
        
        # 检查文件大小（可选：警告小文件）
        if file_size < LARGE_FILE_THRESHOLD_BYTES:
            large_file_log.info(f"上传文件 {file.filename} 小于 50MB ({file_size_mb:.2f}MB)，建议使用普通模式")
        
        # 保存文件
        large_file_log.debug(f"开始保存文件: {file.filename}")
        metadata = await large_file_storage.save_file(content, file.filename)
        large_file_log.info(f"文件保存成功: file_id={metadata.file_id}, 路径={metadata.file_path}")
        
        # 获取预览（上传时使用快速模式，不读取样式，大幅提升大文件性能）
        large_file_log.debug(f"开始生成预览: file_id={metadata.file_id}")
        preview = await large_file_storage.get_preview(metadata.file_id, include_styles=False)
        if preview:
            large_file_log.info(f"预览生成成功: file_id={metadata.file_id}, 工作表={metadata.sheet_names}, 预览行数={preview.get('preview_rows', 0)}")
        else:
            large_file_log.warning(f"预览生成失败: file_id={metadata.file_id}, 将返回空预览")
        
        result = {
            "file_id": metadata.file_id,
            "original_name": metadata.original_name,
            "file_size": metadata.file_size,
            "sheet_names": metadata.sheet_names,
            "sheet_row_counts": metadata.sheet_row_counts,
            "row_count": metadata.row_count,
            "col_count": metadata.col_count,
            "preview": preview,
            "duckdb_ready": metadata.duckdb_ready,
            "duckdb_load_stage": metadata.duckdb_load_stage,
            "duckdb_load_progress": metadata.duckdb_load_progress,
            "message": f"文件上传成功，共 {metadata.row_count} 行，前端显示前 {PREVIEW_ROW_COUNT} 行预览"
        }
        
        large_file_log.info(f"文件上传完成: file_id={metadata.file_id}, 文件名={metadata.original_name}")
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        large_file_log.error(f"文件上传失败: {file.filename}, 错误: {str(e)}")
        large_file_log.debug(f"错误详情:\n{error_details}")
        app_log.error(f"文件上传失败: {e}")
        raise HTTPException(status_code=500, detail="文件上传失败，请稍后重试。")


@app.get("/api/large-file/preview/{file_id}")
async def get_large_file_preview(
    file_id: str,
    sheet_name: str = None,
    include_styles: bool = False,
    offset: int = Query(0, ge=0, description="数据偏移量（仅 include_styles=false 时生效）"),
    limit: int = Query(PREVIEW_ROW_COUNT, ge=1, le=5000, description="分页大小（仅 include_styles=false 时生效）"),
):
    """
    获取大型文件的预览数据（TOP 500 行）
    
    优先使用 DuckDB 快速预览（性能提升 20-120倍），如果失败则回退到 openpyxl 方式
    
    Args:
        file_id: 文件ID
        sheet_name: 工作表名称
        include_styles: 是否包含样式（默认 False 以提升性能）
    """
    large_file_log.debug(
        f"获取预览请求: file_id={file_id}, sheet_name={sheet_name}, include_styles={include_styles}, offset={offset}, limit={limit}"
    )
    
    # 优先使用 DuckDB 快速预览（不需要样式时，性能提升 20-120倍）
    if not include_styles:
        try:
            preview = await large_file_storage.get_preview_fast(
                file_id,
                sheet_name,
                max_rows=limit,
                offset=offset
            )
            if preview:
                large_file_log.info(f"预览获取成功（DuckDB快速模式）: file_id={file_id}, 预览行数={preview.get('preview_rows', 0)}")
                return preview
        except Exception as e:
            large_file_log.warning(f"DuckDB 快速预览失败，回退到 openpyxl: file_id={file_id}, error={e}")
            # 继续使用 openpyxl 方式
    
    # 回退到传统方式（需要样式时，或 DuckDB 失败时）
    preview = await large_file_storage.get_preview(file_id, sheet_name, max_rows=limit, include_styles=include_styles)
    if not preview:
        large_file_log.warning(f"预览获取失败: file_id={file_id}, sheet_name={sheet_name}")
        raise HTTPException(status_code=404, detail="文件不存在或无法读取")
    
    large_file_log.info(f"预览获取成功（openpyxl模式）: file_id={file_id}, 预览行数={preview.get('preview_rows', 0)}, 包含样式={include_styles}")
    return preview


@app.get("/api/large-file/status/{file_id}")
async def get_large_file_status(file_id: str):
    """
    获取文件状态
    """
    large_file_log.debug(f"获取文件状态: file_id={file_id}")
    metadata = large_file_storage.get_metadata(file_id)
    if not metadata:
        large_file_log.warning(f"文件不存在: file_id={file_id}")
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 获取结果文件列表
    import json
    import time as _time
    import os as _os
    def _write_debug_log(data: dict):
        try:
            log_path = "/usr1/python/excel-ai/logs/debug.log"
            _os.makedirs(_os.path.dirname(log_path), exist_ok=True)
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(json.dumps(data, ensure_ascii=False) + '\n')
        except Exception:
            if _os.name == "nt":
                try:
                    with open(r'd:\dev\python\excel-ai\.cursor\debug.log', 'a', encoding='utf-8') as f:
                        f.write(json.dumps(data, ensure_ascii=False) + '\n')
                except Exception:
                    pass
    _write_debug_log({
        "sessionId": "debug-session",
        "runId": "run1",
        "hypothesisId": "A",
        "location": "main.py:324",
        "message": "调用 get_result_files 前",
        "data": {"file_id": file_id, "has_method": hasattr(large_file_storage, 'get_result_files')},
        "timestamp": _time.time() * 1000
    })
    try:
        result_files = large_file_storage.get_result_files(file_id)
        _write_debug_log({
            "sessionId": "debug-session",
            "runId": "run1",
            "hypothesisId": "A",
            "location": "main.py:327",
            "message": "调用 get_result_files 成功",
            "data": {"file_id": file_id, "result_count": len(result_files)},
            "timestamp": _time.time() * 1000
        })
    except AttributeError as e:
        _write_debug_log({
            "sessionId": "debug-session",
            "runId": "run1",
            "hypothesisId": "A",
            "location": "main.py:330",
            "message": "AttributeError 异常",
            "data": {"file_id": file_id, "error": str(e), "methods": dir(large_file_storage)[:20]},
            "timestamp": _time.time() * 1000
        })
        result_files = []  # 临时返回空列表避免崩溃
    
    # 计算 DuckDB 加载耗时
    duckdb_load_time_seconds = None
    if metadata.duckdb_ready and metadata.duckdb_load_started_at and metadata.duckdb_load_finished_at:
        duckdb_load_time_seconds = (metadata.duckdb_load_finished_at - metadata.duckdb_load_started_at).total_seconds()
    
    large_file_log.debug(f"文件状态: file_id={file_id}, status={metadata.status.value}, duckdb_ready={metadata.duckdb_ready}, load_time={duckdb_load_time_seconds}s, result_files={len(result_files)}")
    return {
        "file_id": metadata.file_id,
        "status": metadata.status.value,
        "original_name": metadata.original_name,
        "file_size": metadata.file_size,
        "sheet_names": metadata.sheet_names,
        "sheet_row_counts": metadata.sheet_row_counts,  # 每个工作表的行数
        "row_count": metadata.row_count,  # 所有工作表行数总和
        "col_count": metadata.col_count,
        "created_at": metadata.created_at.isoformat(),
        "last_accessed": metadata.last_accessed.isoformat(),
        "duckdb_ready": metadata.duckdb_ready,
        "first_page_ready": metadata.first_page_ready,
        "duckdb_load_time_seconds": duckdb_load_time_seconds,
        "duckdb_load_stage": metadata.duckdb_load_stage,
        "duckdb_load_progress": metadata.duckdb_load_progress,
        "result_files": result_files,
    }


@app.get("/api/large-file/results/{file_id}")
async def get_result_files(file_id: str):
    """
    获取源文件关联的结果文件列表
    """
    large_file_log.debug(f"获取结果文件列表: file_id={file_id}")
    result_files = large_file_storage.get_result_files(file_id)
    return {
        "result_files": result_files
    }


@app.get("/api/large-file/memory-tables/{file_id}")
async def list_memory_tables(file_id: str):
    """
    列出当前会话可用的内存表（源表 + 结果表）
    """
    try:
        from .large_file.large_file_duckdb import duckdb_manager
        tables = duckdb_manager.list_available_tables(file_id)
        return {
            "file_id": file_id,
            "tables": tables
        }
    except Exception as e:
        large_file_log.warning(f"获取内存表失败: file_id={file_id}, error={e}")
        return {
            "file_id": file_id,
            "tables": [],
            "error": str(e)
        }


@app.get("/api/large-file/session-info/{file_id}")
async def get_session_info(file_id: str):
    """
    获取当前会话的内存占用概况（供前端可观测面板使用）

    Returns:
        source_tables: 源数据表列表及行数
        result_tables: 结果表列表及行数
        total_rows: 全部行数合计
        table_count: 表总数
    """
    from .large_file.large_file_duckdb import duckdb_manager

    source_tables = []
    result_tables = []

    try:
        all_tables = duckdb_manager.list_available_tables(file_id)
        for t in all_tables:
            entry = {
                "name": t.get("name", ""),
                "table_name": t.get("table_name", ""),
                "row_count": t.get("row_count", 0),
                "columns": t.get("columns", []),
            }
            if t.get("type") == "result":
                result_tables.append(entry)
            else:
                source_tables.append(entry)
    except Exception as e:
        large_file_log.warning(f"获取会话信息失败: file_id={file_id}, error={e}")

    memory_results = large_file_storage.list_memory_results(file_id)

    total_rows = sum(t["row_count"] for t in source_tables) + sum(t["row_count"] for t in result_tables)
    return {
        "file_id": file_id,
        "source_tables": source_tables,
        "result_tables": result_tables,
        "memory_results": memory_results,
        "table_count": len(source_tables) + len(result_tables),
        "total_rows": total_rows,
    }


@app.get("/api/large-file/result-preview/{file_id}")
async def preview_result_file(file_id: str, sheet_name: Optional[str] = None):
    """
    预览结果文件（实际是预览源文件的某个工作表）
    """
    large_file_log.debug(f"预览结果文件: file_id={file_id}, sheet_name={sheet_name}")
    preview = await large_file_storage.get_preview(file_id, sheet_name, include_styles=True)
    if not preview:
        raise HTTPException(status_code=404, detail="文件不存在或无法读取")
    return preview


@app.delete("/api/large-file/results/{file_id}/sheet")
async def delete_result_sheet(file_id: str, sheet_name: str = Query(..., description="工作表名称")):
    """
    删除结果文件中的指定工作表
    
    使用查询参数而不是路径参数，避免工作表名称中的特殊字符导致 URI 解析错误
    """
    large_file_log.info(f"删除结果工作表: file_id={file_id}, sheet={sheet_name}")
    result = await large_file_storage.remove_sheet_from_result_file(file_id, sheet_name)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("message", "删除失败"))
    return result


@app.patch("/api/large-file/{file_id}/rename-sheet")
async def rename_large_file_sheet(
    file_id: str,
    request: Request,
    user: User = Depends(get_current_user),
):
    """
    重命名大文件工作表（串联 xlsx + meta + DuckDB + 缓存）

    请求体: {"old_name": str, "new_name": str}
    """
    from .large_file.large_file_duckdb import duckdb_manager

    payload = await request.json()
    old_name = payload.get("old_name", "").strip()
    new_name = payload.get("new_name", "").strip()

    if not old_name or not new_name:
        raise HTTPException(400, "old_name 和 new_name 不能为空")
    if old_name == new_name:
        raise HTTPException(400, "新旧名称相同，无需重命名")

    meta = large_file_storage.get_metadata(file_id)
    if not meta:
        raise HTTPException(404, "文件不存在")

    # 1. openpyxl 改名 + 保存 xlsx
    import openpyxl
    from pathlib import Path

    file_path = Path(meta.file_path)
    if not file_path.exists():
        raise HTTPException(404, "物理文件不存在")

    wb = openpyxl.load_workbook(str(file_path))
    if old_name not in wb.sheetnames:
        wb.close()
        raise HTTPException(400, f'工作表 "{old_name}" 不存在')
    if new_name in wb.sheetnames:
        wb.close()
        raise HTTPException(400, f'工作表 "{new_name}" 已存在')

    wb[old_name].title = new_name
    wb.save(str(file_path))
    wb.close()

    # 2. 更新 FileMetadata
    meta.sheet_names = [new_name if n == old_name else n for n in meta.sheet_names]
    if old_name in meta.sheet_row_counts:
        meta.sheet_row_counts[new_name] = meta.sheet_row_counts.pop(old_name)

    # 3. DuckDB 内存表重命名（未加载时静默跳过）
    duckdb_manager.rename_sheet(file_id, old_name, new_name)

    large_file_log.info(f"工作表重命名成功: file_id={file_id}, {old_name} -> {new_name}")
    return {
        "success": True,
        "old_name": old_name,
        "new_name": new_name,
        "sheet_names": meta.sheet_names,
    }


@app.post("/api/large-file/operation/stream")
async def execute_large_file_operation_stream(
    request: Request,
    _quota_feature=Depends(QuotaGuard("large_file_rows")),
    _quota_daily=Depends(QuotaGuard("ai_daily")),
    user_info=Depends(get_optional_user_info),
    db: AsyncSession = Depends(get_db),
):
    """
    执行大型文件操作（流式返回，实时反馈进度）
    
    使用 Server-Sent Events 格式返回，前端可实时接收：
    - thinking: AI 正在思考
    - tool_call: 工具调用开始
    - tool_result: 工具执行结果
    - text: AI 文本回复
    - preview: 操作完成后的预览数据
    - done: 操作完成
    - error: 操作出错
    """
    user_id, username = user_info
    if user_id:
        await increment_usage(user_id, "ai_count", db)
    payload = await request.json()
    file_id = payload.get("file_id")
    command = payload.get("command")
    raw_session_id = payload.get("session_id") or str(uuid.uuid4())
    session_id = f"large_file_{raw_session_id}"
    active_sheet = payload.get("active_sheet")
    large_file_log.info(f"[流式] 收到操作请求: file_id={file_id}, session_id={session_id}, command={command[:100]}...")
    
    # 参数验证
    if not file_id:
        raise HTTPException(status_code=400, detail="file_id 不能为空")
    if not command:
        raise HTTPException(status_code=400, detail="command 不能为空")
    intent_type = _resolve_large_file_intent_type(payload, command)
    require_export_sheet = intent_type == LARGE_FILE_INTENT_ANALYSIS
    large_file_log.info(
        f"[流式] 意图判定: intent_type={intent_type}, require_export_sheet={require_export_sheet}"
    )
    
    metadata = large_file_storage.get_metadata(file_id)
    if not metadata:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    async def event_generator():
        """SSE 事件生成器（带并发心跳）"""
        import json
        import asyncio
        
        LARGE_FILE_LOG_PREFIX = "[流式-DEBUG]"
        
        def debug_log(msg: str, **kwargs):
            """调试日志，包含追踪信息"""
            extra = ', '.join([f"{k}={v}" for k, v in kwargs.items()])
            large_file_log.info(f"{LARGE_FILE_LOG_PREFIX} {msg} ({extra})")

        # region agent log helper
        def _debug_log_to_file(hypothesis_id: str, location: str, message: str, data: dict) -> None:
            try:
                import time as _time
                import os as _os
                payload = {
                    "sessionId": "debug-session",
                    "runId": "run1",
                    "hypothesisId": hypothesis_id,
                    "location": location,
                    "message": message,
                    "data": data,
                    "timestamp": int(_time.time() * 1000)
                }
                # 优先写入远程环境约定路径，其次写本地调试路径
                try:
                    log_path = "/usr1/python/excel-ai/logs/debug.log"
                    _os.makedirs(_os.path.dirname(log_path), exist_ok=True)
                    with open(log_path, "a", encoding="utf-8") as f:
                        f.write(json.dumps(payload, ensure_ascii=False) + "\n")
                except Exception:
                    pass
                if _os.name == "nt":
                    try:
                        with open(r"d:\dev\python\excel-ai\.cursor\debug.log", "a", encoding="utf-8") as f:
                            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
                    except Exception:
                        pass
            except Exception:
                pass
        # endregion
        
        def sse_event(event_type: str, data: dict) -> str:
            """格式化 SSE 事件"""
            debug_log("SSE事件生成", event_type=event_type, data_size=len(str(data)))
            return f"event: {event_type}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"
        
        def sse_heartbeat() -> str:
            """心跳事件（保持连接活跃）"""
            debug_log("心跳事件")
            return ": heartbeat\n\n"
        
        # 使用队列来协调消息和心跳
        message_queue = asyncio.Queue()
        processing_done = asyncio.Event()
        has_error = False
        text_content = ""
        
        async def heartbeat_task():
            """后台心跳任务"""
            heartbeat_interval = 30.0  # 每 30 秒发送心跳
            while not processing_done.is_set():
                try:
                    await asyncio.wait_for(processing_done.wait(), timeout=heartbeat_interval)
                except asyncio.TimeoutError:
                    # 超时，发送心跳
                    await message_queue.put(("heartbeat", None))
        
        async def process_task():
            """处理命令的任务"""
            nonlocal has_error, text_content
            try:
                large_file_log.info(f"[流式] 开始处理命令...")
                async for msg in agent.process_command(
                    command,
                    require_export_sheet=require_export_sheet,
                ):
                    await message_queue.put(("message", msg))
                large_file_log.info(f"[流式] 命令处理完成")
            except asyncio.TimeoutError:
                has_error = True
                large_file_log.error("[流式] AI 处理超时")
                await message_queue.put(("error", "AI 处理超时，请稍后重试"))
            except Exception as e:
                has_error = True
                large_file_log.error(f"[流式] 处理命令错误: {e}")
                await message_queue.put(("error", str(e)))
            finally:
                processing_done.set()
                await message_queue.put(("done", None))
        
        try:
            large_file_log.info(f"[流式] event_generator 开始执行")
            debug_log("event_generator 入口", file_id=file_id, session_id=session_id, command_length=len(command))
            # region agent log H1
            _debug_log_to_file(
                "H1",
                "main.py:event_generator:entry",
                "event_generator_entry",
                {"file_id": file_id, "session_id": session_id, "command_length": len(command)}
            )
            # endregion
            
            # 更新状态
            large_file_storage.update_status(file_id, FileStatus.PROCESSING)
            large_file_log.info(f"[流式] 即将 yield status=processing")
            yield sse_event("status", {"status": "processing", "message": "开始处理..."})
            large_file_log.info(f"[流式] yield status=processing 完成")
            
            # 获取 Agent
            agent = None
            try:
                large_file_log.info(f"[流式] 正在获取 Agent...")
                agent = await large_file_agent_manager.get_or_create_agent(session_id, file_id, active_sheet)
                if username and hasattr(agent, 'log'):
                    agent.log.set_user_tag(f'@{username}')
                large_file_log.info(f"[流式] Agent 获取成功, 即将 yield status=ready")
                yield sse_event("status", {"status": "ready", "message": "AI Agent 已就绪"})
                large_file_log.info(f"[流式] yield status=ready 完成")
            except Exception as e:
                large_file_log.error(f"[流式] Agent 创建失败: {e}")
                has_error = True
                try:
                    yield sse_event("error", {"message": f"Agent 创建失败: {str(e)}"})
                    yield sse_event("done", {"success": False, "message": f"Agent 创建失败: {str(e)}"})
                except GeneratorExit:
                    large_file_log.warning(f"[流式] 生成器被关闭（Agent创建失败后）: file_id={file_id}")
                    return
                except Exception as yield_error:
                    large_file_log.error(f"[流式] 发送错误/完成事件失败: {yield_error}")
                try:
                    large_file_storage.update_status(file_id, FileStatus.ERROR, str(e))
                except Exception as status_error:
                    large_file_log.error(f"[流式] 更新错误状态失败: {status_error}")
                return
            
            # 启动心跳和处理任务
            heartbeat = asyncio.create_task(heartbeat_task())
            processor = asyncio.create_task(process_task())
            
            msg_count = 0
            loop_exited_normally = False
            try:
                while True:
                    try:
                        # 增加超时时间到 1800 秒（30分钟），支持大文件长时间操作
                        msg_type, msg_data = await asyncio.wait_for(message_queue.get(), timeout=1800.0)
                    except asyncio.TimeoutError:
                        large_file_log.warning("[流式] 消息队列超时，结束循环")
                        # region agent log H2
                        _debug_log_to_file(
                            "H2",
                            "main.py:event_generator:queue_timeout",
                            "message_queue_timeout",
                            {"msg_count": msg_count, "loop_exited_normally": loop_exited_normally}
                        )
                        # endregion
                        break
                    
                    if msg_type == "done":
                        loop_exited_normally = True
                        # region agent log H2
                        _debug_log_to_file(
                            "H2",
                            "main.py:event_generator:done_received",
                            "message_done_received",
                            {"msg_count": msg_count, "loop_exited_normally": loop_exited_normally}
                        )
                        # endregion
                        break
                    elif msg_type == "heartbeat":
                        try:
                            large_file_log.debug("[流式] 发送心跳")
                            yield sse_heartbeat()
                        except Exception as e:
                            large_file_log.error(f"[流式] 发送心跳失败: {e}")
                    elif msg_type == "error":
                        try:
                            has_error = True
                            yield sse_event("error", {"message": msg_data})
                        except Exception as e:
                            large_file_log.error(f"[流式] 发送错误事件失败: {e}")
                    elif msg_type == "message":
                        try:
                            msg = msg_data
                            agent_msg_type = msg.get("type", "unknown")
                            msg_count += 1
                            large_file_log.info(f"[流式] 收到消息 #{msg_count}: type={agent_msg_type}")
                            
                            if agent_msg_type == "thinking":
                                yield sse_event("thinking", {"content": msg.get("content", "思考中...")})
                            
                            elif agent_msg_type == "tool_use":
                                tool_name = msg.get("tool_name", "unknown")
                                tool_input = msg.get("tool_input", {})
                                display_input = {k: str(v)[:100] + "..." if len(str(v)) > 100 else v 
                                                 for k, v in tool_input.items()} if tool_input else {}
                                yield sse_event("tool_call", {
                                    "tool_name": tool_name,
                                    "tool_input": display_input,
                                    "message": f"正在执行: {tool_name}"
                                })
                            
                            elif agent_msg_type == "tool_result":
                                tool_name = msg.get("tool_name", "unknown")
                                result = msg.get("result", "")
                                result_str = str(result) if not isinstance(result, str) else result
                                display_result = result_str[:200] + "..." if len(result_str) > 200 else result_str
                                tool_result_data = {
                                    "tool_name": tool_name,
                                    "result": display_result,
                                    "success": not msg.get("is_error", False)
                                }
                                # 解析工具结果，提取结果文件、工作表信息和执行进度
                                try:
                                    if isinstance(result, str):
                                        result_obj = json.loads(result)
                                    else:
                                        result_obj = result
                                    if isinstance(result_obj, dict):
                                        # 提取结果文件信息
                                        data = result_obj.get("data", {})
                                        if data.get("result_file_id") and data.get("sheet_name"):
                                            tool_result_data["new_file"] = {
                                                "file_id": data["result_file_id"],
                                                "filename": data.get("filename"),
                                                "sheet_name": data.get("sheet_name"),
                                                "row_count": data.get("row_count", 0),
                                                "col_count": data.get("col_count", 0),
                                                "can_reprocess": data.get("can_reprocess", False)  # 是否支持二次加工
                                            }
                                        
                                        # 发送后端操作进度（包含 SQL、执行时间等详细信息）
                                        steps = result_obj.get("steps", [])
                                        execution_time_ms = result_obj.get("execution_time_ms")
                                        sql_executed = result_obj.get("sql_executed")
                                        
                                        if steps or execution_time_ms or sql_executed:
                                            progress_data = {
                                                "tool_name": tool_name,
                                                "steps": steps,  # 执行步骤列表
                                                "message": result_obj.get("message", ""),
                                                "success": result_obj.get("success", True)
                                            }
                                            if execution_time_ms is not None:
                                                progress_data["execution_time_ms"] = execution_time_ms
                                            if sql_executed:
                                                progress_data["sql_executed"] = sql_executed
                                            
                                            # 发送后端进度事件
                                            yield sse_event("backend_progress", progress_data)
                                            large_file_log.debug(f"[流式] 发送后端进度: steps={len(steps)}, time={execution_time_ms}ms")
                                except (json.JSONDecodeError, TypeError, KeyError) as parse_error:
                                    large_file_log.debug(f"[流式] 解析工具结果失败: {parse_error}")
                                yield sse_event("tool_result", tool_result_data)
                            
                            elif agent_msg_type == "text":
                                content = msg.get("content", "")
                                text_content += content
                                large_file_log.info(f"[流式] 发送文本: {content[:100] if content else '(空)'}...")
                                yield sse_event("text", {"content": content})
                            
                            elif agent_msg_type == "error":
                                has_error = True
                                error_msg = msg.get("content", "未知错误")
                                large_file_log.error(f"[流式] Agent 报告错误: {error_msg}")
                                yield sse_event("error", {"message": error_msg})
                        except Exception as e:
                            large_file_log.error(f"[流式] 处理消息时出错: {e}")
                            # 继续处理，不中断流
                
                large_file_log.info(f"[流式] 命令处理完成，共收到 {msg_count} 条消息，正常退出={loop_exited_normally}")
                # region agent log H2
                _debug_log_to_file(
                    "H2",
                    "main.py:event_generator:loop_exit",
                    "loop_exit",
                    {"msg_count": msg_count, "loop_exited_normally": loop_exited_normally}
                )
                # endregion
            
            finally:
                # 确保任务被清理
                heartbeat.cancel()
                processor.cancel()
                try:
                    await heartbeat
                except asyncio.CancelledError:
                    pass
                except Exception as e:
                    large_file_log.error(f"[流式] 心跳任务清理错误: {e}")
                try:
                    await processor
                except asyncio.CancelledError:
                    pass
                except Exception as e:
                    large_file_log.error(f"[流式] 处理任务清理错误: {e}")
            
            # 更新状态
            try:
                if has_error:
                    large_file_storage.update_status(file_id, FileStatus.READY)  # 错误后仍可继续操作
                else:
                    large_file_storage.update_status(file_id, FileStatus.READY)
            except Exception as e:
                large_file_log.error(f"[流式] 更新状态失败: {e}")
            
            # 获取更新后的预览（即使有错误也尝试获取）
            try:
                yield sse_event("status", {"status": "loading_preview", "message": "正在加载预览..."})
                preview = await large_file_storage.get_preview(file_id, sheet_name=active_sheet, include_styles=True)
                
                if preview:
                    large_file_log.info(f"[流式] 发送预览事件: rows={preview.get('preview_rows')}, has_styles={bool(preview.get('styles'))}")
                    yield sse_event("preview", {"preview": preview})
                else:
                    large_file_log.warning("[流式] 预览数据为空，跳过发送")
            except Exception as e:
                large_file_log.error(f"[流式] 获取预览失败: {e}")
                try:
                    yield sse_event("warning", {"message": f"获取预览失败: {str(e)}"})
                except Exception as yield_error:
                    large_file_log.error(f"[流式] 发送警告事件失败: {yield_error}")
            
            # 发送完成事件（确保总是发送）
            done_sent = False
            try:
                done_message = text_content or ("操作完成（有错误发生）" if has_error else "操作完成")
                large_file_log.info(f"[流式] 准备发送完成事件: file_id={file_id}, has_error={has_error}, message_length={len(done_message)}")
                # region agent log H3
                _debug_log_to_file(
                    "H3",
                    "main.py:event_generator:done_before_yield",
                    "done_before_yield",
                    {"has_error": has_error, "message_length": len(done_message)}
                )
                # endregion
                yield sse_event("done", {
                    "success": not has_error,
                    "message": done_message
                })
                done_sent = True
                large_file_log.info(f"[流式] 完成事件已发送: file_id={file_id}, has_error={has_error}")
                # region agent log H3
                _debug_log_to_file(
                    "H3",
                    "main.py:event_generator:done_after_yield",
                    "done_after_yield",
                    {"done_sent": done_sent, "has_error": has_error}
                )
                # endregion
            except GeneratorExit:
                large_file_log.warning(f"[流式] 生成器被关闭（客户端断开连接）: file_id={file_id}, done_sent={done_sent}")
                return
            except Exception as e:
                large_file_log.error(f"[流式] 发送完成事件失败: {e}, done_sent={done_sent}")
                import traceback
                large_file_log.debug(traceback.format_exc())
                # 即使失败也记录日志，但不抛出异常，确保生成器正常结束
            
            large_file_log.info(f"[流式] 生成器即将结束: file_id={file_id}, done_sent={done_sent}, has_error={has_error}")
            
        except asyncio.CancelledError:
            large_file_log.warning(f"[流式] 请求被取消或客户端断开: file_id={file_id}")
            return
        except Exception as e:
            import traceback
            large_file_log.error(f"[流式] 严重错误: {str(e)}")
            large_file_log.debug(traceback.format_exc())
            # region agent log H4
            _debug_log_to_file(
                "H4",
                "main.py:event_generator:outer_exception",
                "outer_exception",
                {"error": str(e)}
            )
            # endregion
            try:
                large_file_storage.update_status(file_id, FileStatus.ERROR, str(e))
            except Exception as status_error:
                large_file_log.error(f"[流式] 更新错误状态失败: {status_error}")
            try:
                yield sse_event("error", {"message": f"操作失败: {str(e)}"})
                yield sse_event("done", {"success": False, "message": f"操作失败: {str(e)}"})
            except Exception as yield_error:
                large_file_log.error(f"[流式] 发送错误事件失败: {yield_error}")
                # 即使失败也记录日志，确保生成器不会提前结束
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",  # 禁用 nginx 缓冲
        }
    )


@app.post("/api/large-file/operation")
async def execute_large_file_operation(
    request: Request,
    _quota_feature=Depends(QuotaGuard("large_file_rows")),
    _quota_daily=Depends(QuotaGuard("ai_daily")),
    user_info=Depends(get_optional_user_info),
    db: AsyncSession = Depends(get_db),
):
    """
    执行大型文件操作（批量返回，兼容旧版本）
    
    请求体：
    {
        "file_id": "xxx",
        "command": "用户指令",
        "intent_type": "analysis|info|planning (可选，默认 analysis)",
        "session_id": "可选，用于保持会话",
        "active_sheet": "可选，当前活动工作表"
    }
    """
    user_id, username = user_info
    if user_id:
        await increment_usage(user_id, "ai_count", db)
    payload = await request.json()
    file_id = payload.get("file_id")
    command = payload.get("command")
    raw_session_id = payload.get("session_id") or str(uuid.uuid4())
    session_id = f"large_file_{raw_session_id}"
    active_sheet = payload.get("active_sheet")
    large_file_log.info(f"收到操作请求: file_id={file_id}, session_id={session_id}, active_sheet={active_sheet}, command={command[:100]}...")
    
    if not file_id:
        large_file_log.warning("操作请求缺少 file_id")
        raise HTTPException(status_code=400, detail="file_id 不能为空")
    if not command:
        large_file_log.warning(f"操作请求缺少 command: file_id={file_id}")
        raise HTTPException(status_code=400, detail="command 不能为空")
    intent_type = _resolve_large_file_intent_type(payload, command)
    require_export_sheet = intent_type == LARGE_FILE_INTENT_ANALYSIS
    large_file_log.info(
        f"意图判定: intent_type={intent_type}, require_export_sheet={require_export_sheet}"
    )
    
    metadata = large_file_storage.get_metadata(file_id)
    if not metadata:
        large_file_log.warning(f"文件不存在: file_id={file_id}")
        raise HTTPException(status_code=404, detail="文件不存在")
    
    large_file_storage.update_status(file_id, FileStatus.PROCESSING)
    
    try:
        agent = await large_file_agent_manager.get_or_create_agent(session_id, file_id, active_sheet)
        if username and hasattr(agent, 'log'):
            agent.log.set_user_tag(f'@{username}')
        
        responses = []
        async for msg in agent.process_command(
            command,
            require_export_sheet=require_export_sheet,
        ):
            responses.append(msg)
            msg_type = msg.get("type", "unknown")
            if msg_type == "tool_use":
                large_file_log.info(f"工具调用: {msg.get('tool_name', 'unknown')}")
        
        text_content = "".join(r.get("content", "") for r in responses if r.get("type") == "text")
        has_agent_error = any(r.get("type") == "error" for r in responses)
        has_new_sheet = False
        for r in responses:
            if r.get("type") != "tool_result":
                continue
            result = r.get("result")
            if isinstance(result, dict):
                data = result.get("data") or {}
                if data.get("result_file_id") and data.get("sheet_name"):
                    has_new_sheet = True
                    break
        
        large_file_storage.update_status(file_id, FileStatus.READY)
        preview = await large_file_storage.get_preview(file_id, sheet_name=active_sheet, include_styles=True)
        
        large_file_log.info(f"操作完成: file_id={file_id}")
        return {
            "success": (not has_agent_error),
            "message": text_content or ("操作完成" if not has_agent_error else "操作失败"),
            "preview": preview,
            "has_new_sheet": has_new_sheet,
            "responses": responses,
        }
        
    except Exception as e:
        import traceback
        large_file_storage.update_status(file_id, FileStatus.ERROR, str(e))
        large_file_log.error(f"大文件操作失败: {str(e)}")
        large_file_log.debug(traceback.format_exc())
        raise HTTPException(status_code=500, detail="操作失败，请稍后重试。")


@app.get("/api/large-file/preview-result/{file_id}/{sheet_name}")
async def preview_memory_result(file_id: str, sheet_name: str):
    """
    预览内存中的结果（从 DuckDB 内存表读取，极快）
    
    Args:
        file_id: 源文件ID
        sheet_name: 结果工作表名称
    """
    large_file_log.debug(f"预览内存结果: file_id={file_id}, sheet_name={sheet_name}")
    
    result_info = large_file_storage.get_memory_result(file_id, sheet_name)
    if not result_info:
        raise HTTPException(status_code=404, detail="内存结果不存在")
    
    try:
        from .large_file.large_file_duckdb import duckdb_manager
        
        table_name = result_info['table_name']
        columns = result_info['columns']
        row_count = result_info['row_count']
        
        # 从 DuckDB 内存表获取预览（极快）
        preview = duckdb_manager.get_preview_from_table(table_name, limit=500)
        data = preview['data']
        
        preview_data = {
            'file_id': file_id,
            'sheet_name': sheet_name,
            'headers': columns,
            'data': data,
            'total_rows': row_count,
            'total_cols': len(columns),
            'preview_rows': len(data),
            'in_memory': True
        }
        
        large_file_log.info(f"内存结果预览成功: file_id={file_id}, sheet={sheet_name}, 预览行数={len(data)}")
        return preview_data
        
    except Exception as e:
        large_file_log.error(f"预览内存结果失败: {e}")
        raise HTTPException(status_code=500, detail="预览失败，请稍后重试。")


@app.get("/api/large-file/list-results/{file_id}")
async def list_memory_results(file_id: str):
    """
    列出源文件的所有内存结果
    """
    results = large_file_storage.list_memory_results(file_id)
    return {
        "file_id": file_id,
        "results": results
    }


@app.post("/api/large-file/save-result/{file_id}")
async def save_memory_result(file_id: str, filename: str = None):
    """
    保存内存结果到文件（下载时触发）
    
    返回详细的进度信息，前端可以显示保存进度
    
    Args:
        file_id: 源文件ID
        filename: 文件名（可选）
    """
    large_file_log.info(f"保存内存结果到文件: file_id={file_id}, filename={filename}")
    
    # 检查内存结果
    memory_results = large_file_storage.list_memory_results(file_id)
    if not memory_results:
        # 兼容场景：当前会话内存已清理，但历史结果文件仍存在
        existing_result_id = large_file_storage.get_result_file_id(file_id)
        if existing_result_id:
            existing_meta = large_file_storage.get_metadata(existing_result_id)
            existing_path = large_file_storage.get_file_path(existing_result_id)
            if existing_meta and existing_path and existing_path.exists():
                return {
                    "success": True,
                    "file_id": existing_meta.file_id,
                    "filename": existing_meta.original_name,
                    "message": "当前没有新的内存结果，已返回现有结果文件",
                    "total_sheets": len(existing_meta.sheet_names or []),
                    "total_rows": existing_meta.row_count or 0,
                    "progress": [],
                    "existing_result": True
                }
        raise HTTPException(status_code=404, detail="没有内存结果需要保存")
    
    total_sheets = len(memory_results)
    total_rows = sum(r['row_count'] for r in memory_results)
    
    large_file_log.info(f"开始保存: {total_sheets} 个工作表, 共 {total_rows:,} 行数据")
    
    # 保存文件（带进度回调）
    progress_messages = []
    
    def progress_callback(stage: str, message: str, progress: float = None):
        """进度回调"""
        progress_info = {
            "stage": stage,
            "message": message,
            "progress": progress
        }
        progress_messages.append(progress_info)
        progress_percent = f"{progress*100:.1f}%" if progress is not None else "N/A"
        large_file_log.info(f"[保存进度] {stage}: {message} (进度: {progress_percent})")
    
    result_meta = await large_file_storage.save_memory_results_to_file(
        file_id, 
        filename,
        progress_callback=progress_callback
    )
    
    if not result_meta:
        raise HTTPException(status_code=500, detail="保存文件失败")
    
    return {
        "success": True,
        "file_id": result_meta.file_id,
        "filename": result_meta.original_name,
        "message": f"结果已保存到文件: {total_sheets} 个工作表, {total_rows:,} 行数据",
        "total_sheets": total_sheets,
        "total_rows": total_rows,
        "progress": progress_messages  # 返回进度信息，前端可以显示
    }


@app.delete("/api/large-file/result/{file_id}/{sheet_name}")
async def close_memory_result(file_id: str, sheet_name: str):
    """
    关闭内存结果（销毁内存中的数据，释放 DuckDB 内存表）
    
    Args:
        file_id: 源文件ID
        sheet_name: 结果工作表名称
    """
    large_file_log.info(f"关闭内存结果: file_id={file_id}, sheet_name={sheet_name}")
    
    result = large_file_storage.remove_memory_result(file_id, sheet_name)
    if not result.get('success'):
        raise HTTPException(status_code=404, detail=result.get('message', '内存结果不存在'))
    
    return {
        "success": True,
        "message": result.get('message', f'内存结果 [{sheet_name}] 已释放'),
        "remaining_sheets": result.get('remaining_sheets', [])
    }


@app.post("/api/large-file/close-sheet")
async def close_result_sheet(request: Request):
    """
    关闭结果工作表（释放对应的 DuckDB 内存表）
    
    当用户关闭某个结果工作表标签时调用此接口，
    同时从结果文件和 DuckDB 内存中删除该工作表
    
    请求体：
    {
        "source_file_id": "源文件ID",
        "sheet_name": "工作表名称"
    }
    """
    payload = await request.json()
    source_file_id = payload.get("source_file_id")
    sheet_name = payload.get("sheet_name")
    
    if not source_file_id:
        raise HTTPException(status_code=400, detail="source_file_id 不能为空")
    if not sheet_name:
        raise HTTPException(status_code=400, detail="sheet_name 不能为空")
    
    large_file_log.info(f"关闭结果工作表: source_file_id={source_file_id}, sheet_name={sheet_name}")
    
    # 1. 释放内存结果（DuckDB 内存表）
    memory_result = large_file_storage.remove_memory_result(source_file_id, sheet_name)
    large_file_log.debug(f"内存结果释放: {memory_result}")
    
    # 2. 从结果文件中删除工作表（如果存在）
    file_result = await large_file_storage.remove_sheet_from_result_file(source_file_id, sheet_name)
    large_file_log.debug(f"文件工作表删除: {file_result}")
    
    # 3. 释放 DuckDB 中的结果表（如果存在）
    from .large_file.large_file_duckdb import duckdb_manager
    duckdb_result = duckdb_manager.unload_sheet(f"result_{source_file_id}", sheet_name)
    large_file_log.debug(f"DuckDB 表释放: {duckdb_result}")
    
    return {
        "success": True,
        "message": f"工作表 [{sheet_name}] 已关闭，内存已释放",
        "memory_released": memory_result.get('success', False),
        "file_updated": file_result.get('success', False),
        "result_file_deleted": file_result.get('result_deleted', False),
        "remaining_sheets": file_result.get('sheet_names', [])
    }


@app.post("/api/large-file/clear-session")
async def clear_session_memory(request: Request):
    """
    清空会话的所有 DuckDB 内存（导出文件后调用）
    
    请求体：
    {
        "source_file_id": "源文件ID"
    }
    """
    payload = await request.json()
    source_file_id = payload.get("source_file_id")
    
    if not source_file_id:
        raise HTTPException(status_code=400, detail="source_file_id 不能为空")
    
    large_file_log.info(f"清空会话内存: source_file_id={source_file_id}")
    
    # 清空会话内存
    result = large_file_storage.clear_session_memory(source_file_id)

    # 失效 prepare 映射，避免后续命中脏缓存（已清会话却返回缓存就绪态）
    stale_user_file_ids = [
        user_file_id for user_file_id, prepared_file_id in _prepared_user_file_map.items()
        if prepared_file_id == source_file_id or user_file_id == source_file_id
    ]
    for stale_user_file_id in stale_user_file_ids:
        _prepared_user_file_map.pop(stale_user_file_id, None)
    if stale_user_file_ids:
        for uid, active_user_file_id in list(_user_active_prepare.items()):
            if active_user_file_id in stale_user_file_ids:
                _user_active_prepare.pop(uid, None)
    
    return {
        "success": True,
        "message": f"会话内存已清空",
        "cleared_sheets": result.get('cleared_sheets', 0),
        "cleared_tables": result.get('cleared_tables', 0),
        "invalidated_prepare_mappings": len(stale_user_file_ids)
    }


@app.get("/api/large-file/download/{file_id}")
async def download_large_file(
    file_id: str, 
    clear_memory: bool = Query(False, description="下载后是否清空会话内存（导出分析结果后使用）")
):
    """
    下载文件（支持内存结果，下载时自动保存）
    
    注意：如果检测到内存结果，会先保存到文件再下载。
    建议前端流程：
    1. 先调用 POST /api/large-file/save-result/{file_id} 显示保存进度
    2. 保存完成后，再调用此接口下载文件
    
    Args:
        file_id: 源文件ID（如果有内存结果）或结果文件ID
        clear_memory: 下载后是否清空会话内存（导出分析结果后建议设为 True）
    """
    large_file_log.info(f"收到下载请求: file_id={file_id}, clear_memory={clear_memory}")
    
    # 记录原始的 source_file_id（用于后续清理内存）
    source_file_id = file_id
    
    # 检查是否是内存结果（通过检查是否有内存结果缓存）
    memory_results = large_file_storage.list_memory_results(file_id)
    if memory_results:
        # 有内存结果，先保存到文件
        total_sheets = len(memory_results)
        total_rows = sum(r['row_count'] for r in memory_results)
        large_file_log.info(f"检测到内存结果，正在保存到文件: {total_sheets} 个工作表, {total_rows:,} 行数据")
        large_file_log.info(f"提示：建议前端先调用 save-result 接口显示保存进度，然后再下载")
        
        # 保存文件（同步保存，因为下载需要文件已存在）
        # 注意：这里没有进度回调，因为下载接口是同步的
        # 前端应该先调用 save-result 接口显示进度
        result_meta = await large_file_storage.save_memory_results_to_file(file_id)
        if result_meta:
            file_id = result_meta.file_id  # 使用保存后的文件ID
            large_file_log.info(f"内存结果已保存，准备下载: file_id={file_id}")
        else:
            raise HTTPException(status_code=500, detail="保存内存结果到文件失败")
    
    file_path = large_file_storage.get_file_path(file_id)
    if not file_path or not file_path.exists():
        large_file_log.warning(f"文件不存在: file_id={file_id}, path={file_path}")
        raise HTTPException(status_code=404, detail="文件不存在")
    
    metadata = large_file_storage.get_metadata(file_id)
    filename = metadata.original_name if metadata else f"{file_id}.xlsx"
    
    # 确保文件名有 .xlsx 扩展名
    if not filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        filename = f"{filename}.xlsx"
    
    file_size = file_path.stat().st_size if file_path.exists() else 0
    large_file_log.info(f"开始下载文件: file_id={file_id}, 文件名={filename}, 大小={file_size / 1024 / 1024:.2f} MB")
    
    # 如果设置了 clear_memory=True，在下载完成后清空会话内存
    # 注意：FileResponse 会自动处理文件流，这里通过 background task 清理
    if clear_memory:
        large_file_log.info(f"下载完成后将清空会话内存: source_file_id={source_file_id}")
        
        # 使用背景任务清理内存（下载完成后执行）
        from fastapi import BackgroundTasks
        from starlette.background import BackgroundTask
        
        def cleanup_session_memory():
            """后台清理会话内存"""
            try:
                result = large_file_storage.clear_session_memory(source_file_id)
                large_file_log.info(f"会话内存已清空: source_file_id={source_file_id}, cleared={result}")
            except Exception as e:
                large_file_log.error(f"清空会话内存失败: source_file_id={source_file_id}, error={e}")
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=BackgroundTask(cleanup_session_memory)
        )
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.delete("/api/large-file/{file_id}")
async def delete_large_file(file_id: str):
    """
    删除大型文件
    """
    success = await large_file_storage.delete_file(file_id)
    if not success:
        raise HTTPException(status_code=404, detail="文件不存在或删除失败")
    
    return {"success": True, "message": "文件已删除"}


# 旧报表端点已迁移至 /api/report/ 路由（backend/app/report/router.py）


@app.get("/api/large-file/list")
async def list_large_files():
    """
    列出所有大型文件
    """
    files = large_file_storage.list_files()
    return {
        "files": [
            {
                "file_id": f.file_id,
                "original_name": f.original_name,
                "file_size": f.file_size,
                "status": f.status.value,
                "row_count": f.row_count,
                "created_at": f.created_at.isoformat(),
            }
            for f in files
        ]
    }


# ============================================================
# Error Handlers
# ============================================================

@app.exception_handler(HTTPException)
async def http_exception_handler(request, exc):
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail}
    )


@app.exception_handler(Exception)
async def general_exception_handler(request, exc):
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc)}
    )
