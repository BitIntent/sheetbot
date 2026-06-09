# ============================================================================
# 表单收集 - 业务逻辑
# CRUD / 提交处理 / 同步到工作表
# ============================================================================
from __future__ import annotations

import json
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import select, func, update, or_
from sqlalchemy.ext.asyncio import AsyncSession

from ..utils.logger import get_logger
from ..files import service as file_service
from .models import Form, FormSubmission

logger = get_logger("collect.service")
PROJECT_ROOT = Path(__file__).resolve().parents[3]


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


# ── 表单 CRUD ──────────────────────────────────────────────

async def create_form(
    db: AsyncSession,
    user_id: str,
    *,
    title: str,
    description: str = "",
    sheet_name: Optional[str] = None,
    file_id: Optional[str] = None,
    form_config: Dict[str, Any],
    max_submissions: Optional[int] = None,
    expires_at: Optional[datetime] = None,
) -> Form:
    """创建新表单"""
    if file_id:
        await file_service.assert_active_file_owned(db, user_id, file_id)

    form = Form(
        id=str(uuid.uuid4()),
        user_id=user_id,
        file_id=file_id,
        sheet_name=sheet_name,
        title=title,
        description=description,
        share_token=uuid.uuid4().hex,
        form_config=json.dumps(form_config, ensure_ascii=False),
        status="active",
        max_submissions=max_submissions,
        expires_at=expires_at,
    )
    db.add(form)
    await db.flush()
    return form


async def list_forms(
    db: AsyncSession,
    user_id: str,
) -> List[Form]:
    """列出用户所有表单（按创建时间倒序）"""
    result = await db.execute(
        select(Form)
        .where(Form.user_id == user_id)
        .order_by(Form.created_at.desc())
    )
    return list(result.scalars().all())


async def get_form(
    db: AsyncSession,
    form_id: str,
    user_id: str,
) -> Optional[Form]:
    """获取用户的表单"""
    result = await db.execute(
        select(Form).where(Form.id == form_id, Form.user_id == user_id)
    )
    return result.scalar_one_or_none()


async def assert_form_owned(
    db: AsyncSession,
    form_id: str,
    user_id: str,
) -> Form:
    """断言表单归属于当前用户。"""
    form = await get_form(db, form_id, user_id)
    if not form:
        raise ValueError("表单不存在")
    return form


async def update_form(
    db: AsyncSession,
    form: Form,
    allow_none_keys: Optional[set[str]] = None,
    **kwargs: Any,
) -> Form:
    """更新表单字段"""
    allow_none_keys = allow_none_keys or set()
    for key, value in kwargs.items():
        if value is None and key not in allow_none_keys:
            continue
        if key == "form_config" and isinstance(value, dict):
            setattr(form, key, json.dumps(value, ensure_ascii=False))
        else:
            setattr(form, key, value)
    await db.flush()
    return form


async def delete_form(db: AsyncSession, form: Form) -> None:
    await db.delete(form)
    await db.flush()


# ── 公开表单 ──────────────────────────────────────────────

async def get_form_by_token(
    db: AsyncSession,
    share_token: str,
) -> Optional[Form]:
    """通过 share_token 获取表单（公开端点，无需 user_id）"""
    result = await db.execute(
        select(Form).where(Form.share_token == share_token)
    )
    return result.scalar_one_or_none()


def is_form_accepting(form: Form) -> bool:
    """判断表单是否还接受新提交"""
    if form.status != "active":
        return False
    if form.max_submissions and form.submission_count >= form.max_submissions:
        return False
    if form.expires_at and _utc_now() > form.expires_at:
        return False
    return True


async def submit_form(
    db: AsyncSession,
    form: Form,
    data: Dict[str, Any],
    ip_address: Optional[str] = None,
    user_agent: Optional[str] = None,
) -> FormSubmission:
    """创建一条提交记录并递增计数器"""
    submission = FormSubmission(
        id=str(uuid.uuid4()),
        form_id=form.id,
        data=json.dumps(data, ensure_ascii=False),
        ip_address=ip_address,
        user_agent=user_agent,
    )
    db.add(submission)
    form.submission_count = (form.submission_count or 0) + 1
    await db.flush()
    return submission


# ── 提交列表 ──────────────────────────────────────────────

async def list_submissions(
    db: AsyncSession,
    form_id: str,
    page: int = 1,
    page_size: int = 20,
) -> Tuple[List[FormSubmission], int]:
    """分页获取提交记录"""
    total_q = await db.execute(
        select(func.count()).select_from(FormSubmission)
        .where(FormSubmission.form_id == form_id)
    )
    total = total_q.scalar() or 0

    offset = (page - 1) * page_size
    result = await db.execute(
        select(FormSubmission)
        .where(FormSubmission.form_id == form_id)
        .order_by(FormSubmission.submitted_at.desc())
        .offset(offset)
        .limit(page_size)
    )
    items = list(result.scalars().all())
    return items, total


# ── 同步到 xlsx ──────────────────────────────────────────

async def sync_submissions_to_xlsx(
    db: AsyncSession,
    form: Form,
    upload_dir: str,
) -> Dict[str, Any]:
    """
    将未同步的提交批量追加到关联的 xlsx 文件。
    返回同步结果：
    {
        "success": bool,
        "synced_count": int,
        "code": str,
        "message": str,
    }
    """
    logger.info("[sync] start form=%s file_id=%s sheet=%s", form.id, form.file_id, form.sheet_name)

    if not form.file_id:
        logger.warning("同步失败：表单未绑定工作簿 form=%s", form.id)
        return {
            "success": False,
            "synced_count": 0,
            "code": "NO_FILE_BINDING",
            "message": "表单未绑定工作簿，请先在“修改”中选择映射工作簿",
        }

    try:
        user_file = await file_service.assert_active_file_owned(db, form.user_id, form.file_id)
        logger.info("[sync] storage_path=%s", user_file.storage_path)
    except ValueError:
        logger.warning("关联文件不存在: file_id=%s", form.file_id)
        return {
            "success": False,
            "synced_count": 0,
            "code": "FILE_NOT_FOUND",
            "message": "映射工作簿不存在或无权限访问，请重新选择工作簿",
        }

    # 兼容两种 storage_path:
    # 1) 相对 uploads 根目录: "2026-02-26/xxx.xlsx"
    # 2) 已包含 uploads 前缀: "uploads/2026-02-26/xxx.xlsx"
    upload_root = Path(upload_dir)
    if not upload_root.is_absolute():
        upload_root = PROJECT_ROOT / upload_root
    raw_storage_path = Path(str(user_file.storage_path).strip())
    if raw_storage_path.is_absolute():
        xlsx_path = raw_storage_path
    else:
        upload_root_name = upload_root.name
        if raw_storage_path.parts and raw_storage_path.parts[0] == upload_root_name:
            # storage_path 已含 "uploads/" 前缀（如 "uploads/2026-02-26/xxx.xlsx"）
            # 必须从 PROJECT_ROOT 拼，不能直接用相对路径
            xlsx_path = PROJECT_ROOT / raw_storage_path
        else:
            xlsx_path = upload_root / raw_storage_path

    logger.info("[sync] xlsx_path=%s exists=%s", xlsx_path, xlsx_path.exists())

    if not xlsx_path.exists():
        logger.warning("xlsx 文件不存在: %s", xlsx_path)
        return {
            "success": False,
            "synced_count": 0,
            "code": "XLSX_MISSING",
            "message": "映射工作簿文件不存在，请重新选择工作簿",
        }

    # 诊断：全量 synced 分布
    all_q = await db.execute(
        select(FormSubmission.id, FormSubmission.synced)
        .where(FormSubmission.form_id == form.id)
    )
    all_rows = all_q.all()
    logger.info("[sync] all_submissions=%d synced_dist=%s", len(all_rows),
                [(str(r.id)[-8:], r.synced) for r in all_rows])

    # 获取未同步提交
    unsync_q = await db.execute(
        select(FormSubmission)
        .where(
            FormSubmission.form_id == form.id,
            # 兼容历史脏数据：synced 可能为 NULL，前端会按“待同步”显示
            # 同步时应将 false / null 都视为未同步
            or_(FormSubmission.synced.is_(False), FormSubmission.synced.is_(None)),
        )
        .order_by(FormSubmission.submitted_at.asc())
    )
    unsync_items = list(unsync_q.scalars().all())
    logger.info("[sync] pending_count=%d", len(unsync_items))
    if not unsync_items:
        logger.info("无待同步提交: form=%s", form.id)
        return {
            "success": True,
            "synced_count": 0,
            "code": "NO_PENDING_SUBMISSIONS",
            "message": "没有待同步的提交数据",
        }

    # 解析表单字段顺序
    config = json.loads(form.form_config) if form.form_config else {}
    fields = config.get("fields", [])
    field_keys = [f["key"] for f in fields]

    # 使用 openpyxl 追加行
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path))
    target_sheet = form.sheet_name or wb.sheetnames[0]
    if target_sheet not in wb.sheetnames:
        wb.close()
        logger.warning(
            "目标工作表已被重命名或删除: sheet=%s, file_id=%s",
            target_sheet, form.file_id,
        )
        return {
            "success": False,
            "synced_count": 0,
            "code": "TARGET_SHEET_MISSING",
            "message": "目标工作表已被重命名或删除，请在“修改”中重新选择映射工作表",
        }
    ws = wb[target_sheet]

    synced_count = 0
    for sub in unsync_items:
        row_data = json.loads(sub.data) if isinstance(sub.data, str) else sub.data
        row_values = [row_data.get(k, "") for k in field_keys]
        ws.append(row_values)
        sub.synced = True
        synced_count += 1

    wb.save(str(xlsx_path))
    wb.close()

    # 更新文件行数
    user_file.row_count = (user_file.row_count or 0) + synced_count
    await db.flush()

    logger.info("同步完成: form=%s, rows=%d", form.id, synced_count)
    return {
        "success": True,
        "synced_count": synced_count,
        "code": "SYNC_OK",
        "message": f"同步完成，共写入 {synced_count} 条",
    }
