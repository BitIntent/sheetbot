# backend/app/files/router.py
"""
文件/文件夹管理 API
"""
import os
import json
import uuid
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status, Header
from fastapi.responses import FileResponse as FastAPIFileResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from ..core.database import get_db
from ..core.dependencies import get_current_user
from ..core.quota import QuotaGuard, get_user_quota, get_user_plan_name
from ..core.usage_service import get_storage_mb
from ..auth.models import User
from ..utils.logger import get_logger
from . import service
from ..large_file.storage import large_file_storage
from .schemas import (
    FolderCreate, FolderRename, FolderMoveRequest, FolderResponse,
    FileResponse, FileRename, FileMoveRequest, FileStarResponse,
    FileUploadResponse,
)

logger = get_logger('files_router')

PROJECT_ROOT = Path(__file__).resolve().parents[3]
UPLOAD_ROOT = PROJECT_ROOT / "uploads"

router = APIRouter(prefix="/api/files", tags=["files"])
folder_router = APIRouter(prefix="/api/folders", tags=["folders"])


def _parse_client_updated_at(value: Optional[str]) -> Optional[datetime]:
    """解析前端传入的更新时间（ISO 8601）。"""
    if not value:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    if txt.endswith("Z"):
        txt = f"{txt[:-1]}+00:00"
    try:
        dt = datetime.fromisoformat(txt)
    except ValueError:
        raise HTTPException(status_code=400, detail="X-Expected-Updated-At 格式非法")
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _parse_client_content_version(value: Optional[str]) -> Optional[datetime]:
    """解析前端传入的内容版本戳（ISO 8601）。"""
    if not value:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    if txt.endswith("Z"):
        txt = f"{txt[:-1]}+00:00"
    try:
        dt = datetime.fromisoformat(txt)
    except ValueError:
        raise HTTPException(status_code=400, detail="X-Expected-Content-Version 格式非法")
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _to_utc(dt: datetime) -> datetime:
    """统一转换为 UTC aware datetime。"""
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _normalize_version_ts(dt: datetime) -> datetime:
    """
    版本戳比较统一到秒级，避免 DB 精度（秒）与客户端序列化精度（微秒）不一致导致伪冲突。
    """
    utc = _to_utc(dt)
    return utc.replace(microsecond=0)


# ==================== File Endpoints ====================

@router.get("", response_model=List[FileResponse])
async def list_files(
    folder_id: Optional[str] = Query(None, description="文件夹 ID，不传则返回全部"),
    starred: Optional[bool] = Query(None, description="仅返回星标文件"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取文件列表"""
    fid = folder_id if folder_id is not None else "__unset__"
    files = await service.list_files(db, user.id, folder_id=fid, starred=starred)
    return files


@router.get("/search", response_model=List[FileResponse])
async def search_files(
    q: str = Query(..., min_length=1, description="搜索关键词"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """搜索文件"""
    files = await service.search_files(db, user.id, q)
    return files


@router.get("/storage-usage")
async def storage_usage(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """当前用户的存储用量 + 套餐上限（全部来自 DB，无硬编码）"""
    used = await get_storage_mb(user.id, db)
    quotas = await get_user_quota(user.id, db)
    plan_name = await get_user_plan_name(user.id, db)
    total = quotas.get("storage_mb")
    file_limit = quotas.get("file_size_mb")

    plan_row = (await db.execute(
        text(
            "SELECT us.plan_code "
            "FROM user_subscriptions us "
            "WHERE us.user_id=:uid AND us.status='active'"
        ),
        {"uid": user.id},
    )).mappings().first()

    return {
        "used_mb": used,
        "total_mb": total if total and total != -1 else None,
        "file_size_limit_mb": file_limit if file_limit and file_limit != -1 else None,
        "plan_code": plan_row["plan_code"] if plan_row else "free",
        "plan_name": plan_name,
    }


@router.get("/tree")
async def get_file_tree(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取完整文件树（文件夹 + 文件）"""
    data = await service.get_file_tree(db, user.id)
    folders = [FolderResponse.model_validate(f) for f in data["folders"]]
    files = [FileResponse.model_validate(f) for f in data["files"]]
    return {"folders": [f.model_dump() for f in folders], "files": [f.model_dump() for f in files]}


@router.post("/upload", response_model=FileUploadResponse, status_code=201)
async def upload_file(
    file: UploadFile = File(...),
    folder_id: Optional[str] = Query(None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _quota=Depends(QuotaGuard("storage_mb")),
):
    """统一文件上传入口"""
    if not file.filename:
        raise HTTPException(400, "文件名不能为空")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls", ".xlsm", ".csv"):
        raise HTTPException(400, f"不支持的文件格式: {ext}")

    date_dir = datetime.now().strftime("%Y-%m-%d")
    upload_dir_abs = UPLOAD_ROOT / "excel_files" / date_dir
    upload_dir_abs.mkdir(parents=True, exist_ok=True)

    file_id = str(uuid.uuid4())
    save_rel_path = Path("uploads") / "excel_files" / date_dir / f"{file_id}{ext}"
    save_path = str((PROJECT_ROOT / save_rel_path).resolve())

    content = await file.read()
    file_size = len(content)
    file_size_mb = file_size / (1024 * 1024)

    # ---- 单文件大小限制（即时型检查，非累计型） ----
    quotas = await get_user_quota(user.id, db)
    plan_name = await get_user_plan_name(user.id, db)
    limit_mb = quotas.get("file_size_mb")
    if limit_mb is not None and limit_mb != -1 and file_size_mb > limit_mb:
        raise HTTPException(413, detail={
            "code": "quota_exceeded",
            "key": "file_size_mb",
            "feature": "单文件大小",
            "plan_name": plan_name,
            "limit": limit_mb,
            "current": round(file_size_mb, 1),
            "unit": "MB",
            "message": (
                f"您当前为「{plan_name}」，"
                f"单文件大小上限为 {limit_mb}MB，"
                f"当前文件 {file_size_mb:.1f}MB 超出限制。"
                f"请升级套餐以上传更大文件。"
            ),
        })

    with open(save_path, "wb") as f:
        f.write(content)
    sheet_names_str = None

    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(save_path, read_only=True)
            sheet_names_str = json.dumps(wb.sheetnames, ensure_ascii=False)
            wb.close()
        except Exception as e:
            logger.warning(f"解析工作表名失败: {e}")

    try:
        record = await service.upload_file_record(
            db=db,
            user_id=user.id,
            file_name=file.filename,
            file_size=file_size,
            storage_path=save_rel_path.as_posix(),
            folder_id=folder_id,
            file_format=ext.lstrip("."),
            sheet_names=sheet_names_str,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return FileUploadResponse(
        id=record.id,
        file_name=record.file_name,
        file_size=record.file_size,
        sheet_names=record.sheet_names,
        updated_at=record.updated_at,
        message="上传成功",
    )


@router.patch("/{file_id}/rename", response_model=FileResponse)
async def rename_file(
    file_id: str,
    body: FileRename,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """重命名文件"""
    f = await service.rename_file(db, user.id, file_id, body.name)
    if not f:
        raise HTTPException(404, "文件不存在")
    return f


@router.patch("/{file_id}/star", response_model=FileStarResponse)
async def toggle_star(
    file_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """切换文件星标"""
    f = await service.toggle_star(db, user.id, file_id)
    if not f:
        raise HTTPException(404, "文件不存在")
    return FileStarResponse(id=f.id, is_starred=f.is_starred)


@router.patch("/{file_id}/move")
async def move_file(
    file_id: str,
    body: FileMoveRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """移动文件到目标文件夹"""
    try:
        f = await service.move_file(db, user.id, file_id, body.folder_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not f:
        raise HTTPException(404, "文件不存在")
    return {"id": f.id, "folder_id": f.folder_id, "message": "移动成功"}


@router.get("/{file_id}/download")
async def download_file(
    file_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """下载用户文件（用于前端打开并渲染）"""
    file_record = await service.get_file_by_id(db, user.id, file_id)
    if not file_record:
        raise HTTPException(404, "文件不存在")
    resolved_path = service.resolve_storage_path(file_record.storage_path)
    if not os.path.exists(resolved_path):
        raise HTTPException(404, "文件内容不存在")

    return FastAPIFileResponse(
        path=str(resolved_path),
        filename=file_record.file_name,
        media_type="application/octet-stream",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@router.get("/{file_id}")
async def open_file(
    file_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """兼容旧前端：直接 GET /api/files/{id} 视为下载。"""
    return await download_file(file_id=file_id, user=user, db=db)


@router.put("/{file_id}/content", response_model=FileUploadResponse)
async def save_file_content(
    file_id: str,
    file: UploadFile = File(...),
    expected_content_version: Optional[str] = Header(None, alias="X-Expected-Content-Version"),
    expected_updated_at: Optional[str] = Header(None, alias="X-Expected-Updated-At"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """覆盖保存文件内容（普通模式）"""
    if not file.filename:
        raise HTTPException(400, "文件名不能为空")

    file_record = await service.get_file_by_id(db, user.id, file_id)
    if not file_record:
        raise HTTPException(404, "文件不存在")

    # ============================================================================
    # 并发保护策略（已收敛）：
    # - 仅依赖 accessed_at（内容版本戳），不再用 updated_at 兜底做并发判定，
    #   updated_at 会被元数据更新（rename/star/folder_move...）触发漂移，是误报的主要来源。
    # - 兼容旧前端：旧前端可能把 updated_at 当作 X-Expected-Content-Version 传上来，
    #   若其值恰等于服务端 updated_at，则视为旧客户端的 fallback，跳过版本校验。
    # - 409 时记录详细日志，便于排查。
    # ============================================================================
    client_expected_content_dt = _parse_client_content_version(expected_content_version)
    if client_expected_content_dt is not None:
        server_content_dt = _normalize_version_ts(file_record.accessed_at)
        server_updated_dt = _normalize_version_ts(file_record.updated_at)
        client_content_dt = _normalize_version_ts(client_expected_content_dt)
        if client_content_dt != server_content_dt:
            if client_content_dt == server_updated_dt:
                logger.info(
                    "保存：旧前端 fallback 行为（X-Expected-Content-Version=updated_at），"
                    f"跳过并发校验 file_id={file_id}"
                )
            else:
                logger.warning(
                    f"保存被并发校验拦截 file_id={file_id} "
                    f"client_content={client_content_dt.isoformat()} "
                    f"server_content={server_content_dt.isoformat()} "
                    f"server_updated={server_updated_dt.isoformat()}"
                )
                raise HTTPException(
                    status_code=409,
                    detail={
                        "message": "文件内容已被其他操作更新，请先刷新后再保存，已阻止覆盖保存。",
                        "server_content_version": server_content_dt.isoformat(),
                        "server_updated_at": server_updated_dt.isoformat(),
                        "client_content_version": client_content_dt.isoformat(),
                    },
                )

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls", ".xlsm"):
        raise HTTPException(400, f"不支持的文件格式: {ext}")

    old_path_rel = service.normalize_storage_path(file_record.storage_path)
    old_abs_path = service.resolve_storage_path(file_record.storage_path)
    old_ext = os.path.splitext(old_path_rel)[1].lower()
    target_rel_path = old_path_rel
    if old_ext != ext:
        target_rel_path = f"{os.path.splitext(old_path_rel)[0]}{ext}"
    if not target_rel_path.startswith("uploads/"):
        date_dir = datetime.now().strftime("%Y-%m-%d")
        target_rel_path = f"uploads/excel_files/{date_dir}/{file_id}{ext}"

    target_abs_path = PROJECT_ROOT / target_rel_path

    os.makedirs(os.path.dirname(target_abs_path), exist_ok=True)
    content = await file.read()
    with open(target_abs_path, "wb") as f:
        f.write(content)
    if str(target_abs_path) != str(old_abs_path) and os.path.exists(old_abs_path):
        try:
            os.remove(old_abs_path)
        except Exception:
            logger.warning(f"旧文件删除失败: {old_abs_path}")

    file_size = len(content)
    sheet_names = None
    row_count = 0
    col_count = 0
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(target_abs_path), read_only=True, data_only=True)
        names = wb.sheetnames
        sheet_names = json.dumps(names, ensure_ascii=False)
        for name in names:
            ws = wb[name]
            row_count += (ws.max_row or 0)
            col_count = max(col_count, ws.max_column or 0)
        wb.close()
    except Exception as e:
        logger.warning(f"保存后解析工作表失败: {e}")

    updated = await service.update_file_content(
        db=db,
        user_id=user.id,
        file_id=file_id,
        storage_path=target_rel_path,
        file_size=file_size,
        file_format=ext.lstrip("."),
        sheet_names=sheet_names,
        row_count=row_count,
        col_count=col_count,
    )
    if not updated:
        raise HTTPException(404, "文件不存在")

    return FileUploadResponse(
        id=updated.id,
        file_name=updated.file_name,
        file_size=updated.file_size,
        sheet_names=updated.sheet_names,
        updated_at=updated.updated_at,
        accessed_at=updated.accessed_at,
        message="保存成功",
    )


@router.delete("/{file_id}", status_code=200)
async def delete_file(
    file_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """软删除文件"""
    ok = await service.soft_delete_file(db, user.id, file_id)
    if not ok:
        raise HTTPException(404, "文件不存在")
    cleanup = {}
    try:
        cleanup = await large_file_storage.purge_by_source_file_id(file_id)
    except Exception as e:
        logger.warning(f"清理大文件会话失败: file_id={file_id}, error={e}")
        cleanup = {"source_file_id": file_id, "cleanup_error": str(e)}
    return {"message": "文件已删除", "cleanup": cleanup}


# ==================== Folder Endpoints ====================

@folder_router.get("", response_model=List[FolderResponse])
async def list_folders(
    parent_id: Optional[str] = Query(None, description="父文件夹 ID"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取文件夹列表"""
    folders = await service.list_folders(db, user.id, parent_id)
    return folders


@folder_router.post("", response_model=FolderResponse, status_code=201)
async def create_folder(
    body: FolderCreate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """创建文件夹"""
    try:
        folder = await service.create_folder(db, user.id, body.name, body.parent_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return folder


@folder_router.patch("/{folder_id}", response_model=FolderResponse)
async def rename_folder(
    folder_id: str,
    body: FolderRename,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """重命名文件夹"""
    folder = await service.rename_folder(db, user.id, folder_id, body.name)
    if not folder:
        raise HTTPException(404, "文件夹不存在")
    return folder


@folder_router.patch("/{folder_id}/move", response_model=FolderResponse)
async def move_folder(
    folder_id: str,
    body: FolderMoveRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """移动文件夹到新的父文件夹"""
    try:
        folder = await service.move_folder(db, user.id, folder_id, body.parent_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not folder:
        raise HTTPException(404, "文件夹不存在")
    return folder


@folder_router.delete("/{folder_id}", status_code=200)
async def delete_folder(
    folder_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """删除文件夹"""
    ok = await service.delete_folder(db, user.id, folder_id)
    if not ok:
        raise HTTPException(404, "文件夹不存在")
    return {"message": "文件夹已删除"}
