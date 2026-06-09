# backend/app/excel/chart_inject.py
"""
原生 Excel 图表注入模块
将前端图表元数据（chartType / dataRange / title / row / col / width / height）
转换为 openpyxl 原生图表对象，嵌入到 xlsx 文件中。

图表类型映射：
  column   -> BarChart(barDir='col')
  bar      -> BarChart(barDir='bar', 水平)
  line     -> LineChart
  area     -> AreaChart
  pie      -> PieChart
  doughnut -> DoughnutChart
  scatter  -> ScatterChart
  radar    -> RadarChart
"""
import io
import re
import json
from typing import Any

import openpyxl
from openpyxl.chart import (
    BarChart, LineChart, AreaChart, PieChart, DoughnutChart,
    ScatterChart, RadarChart, Reference, Series,
)
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.utils import get_column_letter, column_index_from_string

from ..utils.logger import get_logger

logger = get_logger('chart_inject')

# ==================== A1 范围解析 ====================

_RANGE_RE = re.compile(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', re.IGNORECASE)
_COL_RE   = re.compile(r'([A-Z]+):([A-Z]+)', re.IGNORECASE)


def _parse_range(range_str: str, ws) -> dict | None:
    """
    解析 A1:B10 格式，返回 {startRow, startCol, endRow, endCol}（全部 1-based）
    支持整列写法 A:B（自动取工作表数据末行）
    """
    if not range_str:
        return None
    # 去掉工作表前缀（Sheet1!A1:B10）
    clean = range_str.split('!')[-1].replace('$', '').strip()

    m = _RANGE_RE.search(clean)
    if m:
        return {
            'startRow': int(m.group(2)),
            'startCol': column_index_from_string(m.group(1)),
            'endRow':   int(m.group(4)),
            'endCol':   column_index_from_string(m.group(3)),
        }
    # 整列写法
    m2 = _COL_RE.match(clean)
    if m2:
        max_row = ws.max_row or 1
        return {
            'startRow': 1,
            'startCol': column_index_from_string(m2.group(1)),
            'endRow':   max_row,
            'endCol':   column_index_from_string(m2.group(2)),
        }
    return None


def _anchor(col_idx: int, row_idx: int) -> str:
    """将 1-based col/row 转为 Excel 锚点字符串，如 'E3'"""
    return f"{get_column_letter(col_idx)}{row_idx}"


# 匹配汇总行标签的正则，与前端 excelOperations.js 保持一致
import re as _re
_TOTAL_ROW_RE = _re.compile(
    r'^(总计|合计|小计|汇总|total|grand\s+total|subtotal|sub\s+total|sum)$',
    _re.IGNORECASE,
)


def _trim_empty_rows(rng: dict, ws) -> dict:
    """裁剪前导/尾部空行，避免空行被纳入图表数据区。"""
    rng = dict(rng)
    max_skip = 5
    for _ in range(max_skip):
        if rng['startRow'] >= rng['endRow']:
            break
        if all(ws.cell(rng['startRow'], c).value in (None, '')
               for c in range(rng['startCol'], rng['endCol'] + 1)):
            rng['startRow'] += 1
        else:
            break
    while rng['endRow'] > rng['startRow'] + 1:
        if all(ws.cell(rng['endRow'], c).value in (None, '')
               for c in range(rng['startCol'], rng['endCol'] + 1)):
            rng['endRow'] -= 1
        else:
            break
    return rng


def _trim_empty_cols(rng: dict, ws) -> dict:
    """裁剪尾部空列：LLM 有时 endCol 多传 1 列，导致导出图表生成幽灵系列。"""
    rng = dict(rng)
    while rng['endCol'] > rng['startCol'] + 1:
        col = rng['endCol']
        if all(ws.cell(r, col).value in (None, '')
               for r in range(rng['startRow'], rng['endRow'] + 1)):
            rng['endCol'] -= 1
        else:
            break
    return rng


def _trim_total_rows(rng: dict, ws) -> dict:
    """
    从 rng 尾部倒序剔除标签列命中"总计/合计"关键词的行。
    同时扫描中间行，返回需要跳过的行号集合 skip_rows。

    饼图/柱图纳入"总计"行会导致数据严重失真（总计占比 50%、柱高翻倍）。
    规则适用于所有图表类型，写进永久记忆。
    """
    rng = dict(rng)
    cat_col = rng['startCol']

    def is_total_label(r: int) -> bool:
        val = ws.cell(r, cat_col).value
        if val is None:
            return False
        return bool(_TOTAL_ROW_RE.match(str(val).strip()))

    # 从尾部倒删
    while rng['endRow'] > rng['startRow'] + 1 and is_total_label(rng['endRow']):
        rng['endRow'] -= 1

    # 收集中间夹着的总计行
    skip_rows: set[int] = set()
    for r in range(rng['startRow'] + 1, rng['endRow']):
        if is_total_label(r):
            skip_rows.add(r)

    rng['_skip_rows'] = skip_rows
    return rng


def _to_number(value: Any) -> float | None:
    """尽量将单元格值解析为数值，失败返回 None"""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(',', '')
        if not text:
            return None
        if text.endswith('%'):
            text = text[:-1]
        try:
            return float(text)
        except Exception:
            return None
    return None


def _has_header_row(ws, rng: dict) -> bool:
    """
    基于首行内容判断 data_range 是否包含表头。
    规则：若首行值列是文本，且后续行主要是数值，则判定有表头。
    """
    start_row = rng['startRow']
    end_row = rng['endRow']
    start_col = rng['startCol']
    end_col = rng['endCol']
    if end_row <= start_row:
        return False

    value_cols = list(range(start_col + 1, end_col + 1))
    # 单值列范围（如 B1:B10）也要尝试识别表头
    if not value_cols:
        value_cols = [start_col]

    probe_end_row = min(end_row, start_row + 8)
    header_signal = False
    numeric_signal = False

    for col in value_cols:
        first_value = ws.cell(start_row, col).value
        first_is_numeric = _to_number(first_value) is not None
        if not first_is_numeric and str(first_value or '').strip():
            header_signal = True

        numeric_count = 0
        sample_count = 0
        for row in range(start_row + 1, probe_end_row + 1):
            sample_value = ws.cell(row, col).value
            if sample_value is None or str(sample_value).strip() == '':
                continue
            sample_count += 1
            if _to_number(sample_value) is not None:
                numeric_count += 1
        if sample_count > 0 and (numeric_count / sample_count) >= 0.6:
            numeric_signal = True

    return header_signal and numeric_signal


def _score_category_col(ws, col: int, start_row: int, end_row: int) -> float:
    """
    对候选类别列做质量打分：
    - 文本覆盖率越高越好
    - 数值密度越低越好（类别轴通常不应是纯数值）
    """
    if col < 1:
        return -1.0
    text_count = 0
    numeric_count = 0
    non_empty = 0
    for row in range(start_row, end_row + 1):
        value = ws.cell(row, col).value
        if value is None or str(value).strip() == '':
            continue
        non_empty += 1
        if _to_number(value) is not None:
            numeric_count += 1
        else:
            text_count += 1
    if non_empty == 0:
        return -1.0
    text_ratio = text_count / non_empty
    numeric_ratio = numeric_count / non_empty
    return text_ratio - 0.4 * numeric_ratio


def _resolve_category_col(ws, rng: dict, has_header: bool) -> int | None:
    """
    解析类别列：
    - 优先使用 data_range 首列
    - 若首列质量差（空/高数值密度），自动回退到左侧一列（若存在）
    """
    default_col = rng['startCol']
    data_start = rng['startRow'] + (1 if has_header else 0)
    data_end = rng['endRow']
    if data_start > data_end:
        return default_col

    default_score = _score_category_col(ws, default_col, data_start, data_end)
    left_col = default_col - 1
    left_score = _score_category_col(ws, left_col, data_start, data_end) if left_col >= 1 else -1.0

    # 左侧列明显更像“类别标签”时，启用自愈
    if left_score > default_score + 0.2:
        return left_col
    if default_score < 0:
        return left_col if left_score >= 0 else None
    return default_col


def _resolve_value_start_col(rng: dict, cat_col: int | None) -> int:
    """
    解析值列起点：
    - 若类别列就是 data_range 首列，值列从下一列开始
    - 若类别列来自左侧补列或不存在，值列从 data_range 首列开始
    """
    if cat_col is not None and cat_col == rng['startCol']:
        return rng['startCol'] + 1
    return rng['startCol']


def _set_categories(chart, ws, cat_col: int, min_row: int, max_row: int) -> None:
    """
    为类别型图表设置分类轴标签。
    使用标准 openpyxl set_categories(Reference) 生成 numRef，
    这是 Excel 兼容性最好的方式（numRef 对文字/数值类别均有效）。
    """
    if cat_col is None or min_row > max_row:
        return
    cat_ref = Reference(ws, min_col=cat_col, max_col=cat_col, min_row=min_row, max_row=max_row)
    chart.set_categories(cat_ref)


def _set_category_axis_title(chart, ws, rng: dict, has_header: bool, cat_col: int | None) -> None:
    """
    将类别列表头单元格（如"渠道""品类"）设为 x 轴标题，
    避免图表中首列标题信息完全丢失。
    """
    if not has_header or cat_col is None:
        return
    cell = ws.cell(row=rng['startRow'], column=cat_col)
    title_text = str(cell.value or '').strip()
    if not title_text:
        return
    try:
        chart.x_axis.title = title_text
    except Exception:
        pass


def _fix_axes(chart, horizontal: bool = False) -> None:
    """
    修正坐标轴位置与可见性。

    openpyxl 默认对两条轴都设 axPos="l"，且不显式写 <delete val="0"/>。
    Excel 遇到缺失 delete 时可能将轴隐藏，axPos 全为 "l" 时两轴互相遮挡。

    正确配置：
      column（竖柱）: x_axis(CatAx) axPos="b"，y_axis(ValAx) axPos="l"
      bar（横柱）   : x_axis(CatAx) axPos="l"，y_axis(ValAx) axPos="b"
    """
    if horizontal:
        # 水平条形图：类别轴在左(l)，值轴在底部(b)
        chart.x_axis.axPos = "l"
        chart.y_axis.axPos = "b"
    else:
        # 竖向柱状图：类别轴在底部(b)，值轴在左(l)
        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"
    # 显式写 delete=False，避免 Excel 默认隐藏轴
    chart.x_axis.delete = False
    chart.y_axis.delete = False


# ==================== 图表类型构建器 ====================

def _make_bar_chart(rng: dict, ws, chart_meta: dict, horizontal: bool) -> BarChart:
    chart = BarChart()
    chart.type  = 'bar' if horizontal else 'col'
    chart.grouping = 'clustered'

    has_header = _has_header_row(ws, rng)
    data_min_row = rng['startRow']
    data_max_row = rng['endRow']
    cat_col      = _resolve_category_col(ws, rng, has_header)
    val_start    = _resolve_value_start_col(rng, cat_col)
    val_end      = rng['endCol']

    if val_start > val_end:
        # 只有一列：把唯一列当值列，无标签列
        val_start = rng['startCol']
        cat_col   = None

    data_ref = Reference(
        ws,
        min_col=val_start, max_col=val_end,
        min_row=data_min_row, max_row=data_max_row,
    )
    chart.add_data(data_ref, titles_from_data=has_header)

    if cat_col:
        _set_categories(
            chart, ws,
            cat_col=cat_col,
            min_row=data_min_row + (1 if has_header else 0),
            max_row=data_max_row,
        )
    _set_category_axis_title(chart, ws, rng, has_header, cat_col)

    # 修正轴位置与可见性（Excel 默认可能将轴隐藏或位置错误）
    _fix_axes(chart, horizontal=horizontal)
    return chart


def _make_line_chart(rng: dict, ws, chart_meta: dict) -> LineChart:
    chart = LineChart()
    has_header = _has_header_row(ws, rng)
    cat_col    = _resolve_category_col(ws, rng, has_header)
    val_start  = _resolve_value_start_col(rng, cat_col)
    val_end    = rng['endCol']

    if val_start > val_end:
        val_start = rng['startCol']
        cat_col   = None

    data_ref = Reference(
        ws,
        min_col=val_start, max_col=val_end,
        min_row=rng['startRow'], max_row=rng['endRow'],
    )
    chart.add_data(data_ref, titles_from_data=has_header)

    if cat_col:
        _set_categories(
            chart, ws,
            cat_col=cat_col,
            min_row=rng['startRow'] + (1 if has_header else 0),
            max_row=rng['endRow'],
        )
    _set_category_axis_title(chart, ws, rng, has_header, cat_col)

    # 折线图：类别轴在底部，值轴在左
    _fix_axes(chart, horizontal=False)
    return chart


def _make_area_chart(rng: dict, ws, chart_meta: dict) -> AreaChart:
    chart = AreaChart()
    has_header = _has_header_row(ws, rng)
    cat_col    = _resolve_category_col(ws, rng, has_header)
    val_start  = _resolve_value_start_col(rng, cat_col)
    val_end    = rng['endCol']

    if val_start > val_end:
        val_start = rng['startCol']
        cat_col   = None

    data_ref = Reference(
        ws,
        min_col=val_start, max_col=val_end,
        min_row=rng['startRow'], max_row=rng['endRow'],
    )
    chart.add_data(data_ref, titles_from_data=has_header)

    if cat_col:
        _set_categories(
            chart, ws,
            cat_col=cat_col,
            min_row=rng['startRow'] + (1 if has_header else 0),
            max_row=rng['endRow'],
        )
    _set_category_axis_title(chart, ws, rng, has_header, cat_col)

    # 面积图：类别轴在底部，值轴在左
    _fix_axes(chart, horizontal=False)
    return chart


def _make_pie_chart(rng: dict, ws, chart_meta: dict, is_doughnut: bool = False) -> PieChart | DoughnutChart:
    chart = DoughnutChart() if is_doughnut else PieChart()

    has_header = _has_header_row(ws, rng)
    cat_col    = _resolve_category_col(ws, rng, has_header)
    val_col    = rng['endCol']

    if cat_col == val_col:
        # 单列时：值列就是唯一列，无标签
        cat_col = None

    data_start = rng['startRow']
    data_end   = rng['endRow']

    data_ref = Reference(
        ws,
        min_col=val_col, max_col=val_col,
        min_row=data_start, max_row=data_end,
    )
    chart.add_data(data_ref, titles_from_data=has_header)

    if cat_col:
        _set_categories(
            chart, ws,
            cat_col=cat_col,
            min_row=data_start + (1 if has_header else 0),
            max_row=data_end,
        )

    return chart


def _make_scatter_chart(rng: dict, ws, chart_meta: dict) -> ScatterChart:
    chart = ScatterChart()
    chart.scatterStyle = 'marker'

    has_header = _has_header_row(ws, rng)
    data_start = rng['startRow'] + (1 if has_header else 0)
    data_end   = rng['endRow']
    x_col      = rng['startCol']

    x_ref = Reference(ws, min_col=x_col, max_col=x_col, min_row=data_start, max_row=data_end)

    for col in range(x_col + 1, rng['endCol'] + 1):
        y_ref = Reference(ws, min_col=col, max_col=col, min_row=data_start, max_row=data_end)
        series = Series(y_ref, x_ref)
        if has_header:
            # XYSeries.tx 需要 SeriesLabel 对象，不能直接赋字符串
            title_val = ws.cell(rng['startRow'], col).value
            if title_val:
                series.title = SeriesLabel(v=str(title_val))
        chart.series.append(series)

    # 散点图：X轴（底部）+ Y轴（左侧），显式可见
    _fix_axes(chart, horizontal=False)
    return chart


def _make_radar_chart(rng: dict, ws, chart_meta: dict) -> RadarChart:
    chart = RadarChart()
    chart.radarStyle = 'marker'

    has_header = _has_header_row(ws, rng)
    cat_col    = _resolve_category_col(ws, rng, has_header)
    val_start  = _resolve_value_start_col(rng, cat_col)
    val_end    = rng['endCol']

    if val_start > val_end:
        val_start = rng['startCol']
        cat_col   = None

    data_ref = Reference(
        ws,
        min_col=val_start, max_col=val_end,
        min_row=rng['startRow'], max_row=rng['endRow'],
    )
    chart.add_data(data_ref, titles_from_data=has_header)

    if cat_col:
        _set_categories(
            chart, ws,
            cat_col=cat_col,
            min_row=rng['startRow'] + (1 if has_header else 0),
            max_row=rng['endRow'],
        )

    # 雷达图：类别轴在底部，值轴在左
    _fix_axes(chart, horizontal=False)
    return chart


# ==================== 尺寸换算 ====================
# openpyxl 图表尺寸单位：EMU (English Metric Units)
# 1 px ≈ 9525 EMU (96 dpi)

_PX_TO_EMU = 9525


def _px_to_emu(px: int | float) -> int:
    return int(px * _PX_TO_EMU)


# ==================== 核心注入函数 ====================

def inject_native_charts(xlsx_bytes: bytes, charts_by_sheet: dict[str, list[dict]]) -> bytes:
    """
    在 xlsx 文件中注入 openpyxl 原生图表。

    Args:
        xlsx_bytes: 前端 ExcelJS 生成的 xlsx 二进制
        charts_by_sheet: { sheetName: [chartMeta, ...] }
            chartMeta 字段：chartType / dataRange / title / row / col / width / height

    Returns:
        注入图表后的 xlsx 二进制
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    for sheet_name, chart_list in charts_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f'[ChartInject] 工作表不存在，跳过: {sheet_name}')
            continue

        ws = wb[sheet_name]

        for meta in chart_list:
            try:
                _inject_one_chart(ws, meta)
            except Exception as e:
                logger.error(f'[ChartInject] 注入图表失败: sheet={sheet_name} meta={meta} err={e}', exc_info=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _inject_one_chart(ws, meta: dict[str, Any]) -> None:
    """将单个图表元数据转换为 openpyxl 图表并添加到工作表"""
    chart_type = str(meta.get('chartType') or 'column').lower()
    data_range = meta.get('dataRange') or ''
    title      = meta.get('title') or ''
    # row/col 是 1-based 锚点
    anchor_row = int(meta.get('row') or 1)
    anchor_col = int(meta.get('col') or 1)
    width_px   = float(meta.get('width')  or 600)
    height_px  = float(meta.get('height') or 400)

    rng = _parse_range(data_range, ws)
    if not rng:
        logger.warning(f'[ChartInject] dataRange 解析失败，跳过: {data_range}')
        return

    # 裁剪前导/尾部空行 —— dataRange 可能包含表头上方或数据下方的空行
    rng = _trim_empty_rows(rng, ws)

    # 裁剪尾部空列 —— LLM 有时 endCol 多传 1 列，导致幽灵系列
    rng = _trim_empty_cols(rng, ws)

    # 剔除"总计/合计"汇总行 —— 纳入饼图/柱图会导致严重失真（永久规则）
    rng = _trim_total_rows(rng, ws)
    skipped = rng.pop('_skip_rows', set())
    if skipped:
        # 中间总计行难以在 openpyxl Reference 中精确剔除，记录 warning 供人工核查
        logger.warning(f'[ChartInject] 发现中间总计行(已记录未剔除): rows={sorted(skipped)} title={title!r}; 建议 Agent 规划时避免将总计行纳入数据范围')

    # 根据类型构建图表对象
    if chart_type == 'bar':
        chart = _make_bar_chart(rng, ws, meta, horizontal=True)
    elif chart_type in ('pie', 'doughnut', 'donut'):
        chart = _make_pie_chart(rng, ws, meta, is_doughnut=(chart_type in ('doughnut', 'donut')))
    elif chart_type == 'line':
        chart = _make_line_chart(rng, ws, meta)
    elif chart_type == 'area':
        chart = _make_area_chart(rng, ws, meta)
    elif chart_type == 'scatter':
        chart = _make_scatter_chart(rng, ws, meta)
    elif chart_type == 'radar':
        chart = _make_radar_chart(rng, ws, meta)
    else:
        # 默认：竖向柱状图
        chart = _make_bar_chart(rng, ws, meta, horizontal=False)

    if title:
        chart.title = title

    chart.width  = width_px  / 96 * 2.54  # px -> cm (96 dpi)
    chart.height = height_px / 96 * 2.54

    ws.add_chart(chart, _anchor(anchor_col, anchor_row))
    logger.info(f'[ChartInject] 注入成功: type={chart_type} anchor={_anchor(anchor_col, anchor_row)} title={title!r}')
