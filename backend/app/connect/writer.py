# ============================================================================
# 外部系统连接器 - 数据写入
# 将适配器拉取的数据按字段映射追加到 Excel 工作表
# 复用 collect 模块的 openpyxl 追加行模式
# ============================================================================
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from ..utils.logger import get_logger

logger = get_logger("connect.writer")


def write_rows_to_xlsx(
    xlsx_path: str | Path,
    sheet_name: Optional[str],
    field_mapping: Dict[str, str],
    rows: List[Dict[str, Any]],
    primary_key: Optional[str] = None,
    deduplicate: bool = False,
) -> int:
    """
    将数据行追加到 Excel 工作表。

    field_mapping: {"外部字段名": "Excel列名", ...}
      例: {"order_number": "订单号", "total_price": "总价"}

    返回实际写入行数。
    """
    if not rows or not field_mapping:
        return 0

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        logger.warning("工作表文件不存在: %s", xlsx_path)
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path))

    # 定位目标工作表
    target = sheet_name or wb.sheetnames[0]
    if target not in wb.sheetnames:
        ws = wb.create_sheet(target)
        _write_header(ws, field_mapping)
    else:
        ws = wb[target]
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            _write_header(ws, field_mapping)

    # 构建列索引映射: Excel列名 -> 列号
    col_index = _get_or_create_col_index(ws, field_mapping)
    dedup_col_num = None
    existing_keys: set[str] = set()
    if deduplicate and primary_key:
        dedup_col_name = field_mapping.get(primary_key, primary_key)
        dedup_col_num = col_index.get(dedup_col_name)
        if dedup_col_num:
            for row_num in range(2, ws.max_row + 1):
                val = ws.cell(row=row_num, column=dedup_col_num).value
                if val is not None:
                    existing_keys.add(str(val))

    # 追加数据行
    written = 0
    for row_data in rows:
        if dedup_col_num and primary_key:
            row_key = row_data.get(primary_key)
            if row_key is not None and str(row_key) in existing_keys:
                continue

        next_row = ws.max_row + 1
        for ext_field, excel_col in field_mapping.items():
            col_num = col_index.get(excel_col)
            if col_num is None:
                continue
            value = row_data.get(ext_field, "")
            ws.cell(row=next_row, column=col_num, value=value)
        if dedup_col_num and primary_key:
            row_key = row_data.get(primary_key)
            if row_key is not None:
                existing_keys.add(str(row_key))
        written += 1

    wb.save(str(xlsx_path))
    wb.close()

    logger.info("写入完成: file=%s, sheet=%s, rows=%d", xlsx_path, target, written)
    return written


def _write_header(ws, field_mapping: Dict[str, str]) -> None:
    """写入表头行"""
    for col_idx, excel_col in enumerate(field_mapping.values(), 1):
        ws.cell(row=1, column=col_idx, value=excel_col)


def _get_or_create_col_index(
    ws, field_mapping: Dict[str, str],
) -> Dict[str, int]:
    """
    读取首行获取列名 -> 列号映射。
    如果映射中的 Excel 列名不存在，追加到末尾。
    """
    col_map: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            col_map[str(val)] = col_idx

    next_col = ws.max_column + 1
    for excel_col in field_mapping.values():
        if excel_col not in col_map:
            ws.cell(row=1, column=next_col, value=excel_col)
            col_map[excel_col] = next_col
            next_col += 1

    return col_map
